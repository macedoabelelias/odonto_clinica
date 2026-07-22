import os

from django.conf import settings

from django.http import (
    HttpResponse,
    JsonResponse
)

from django.shortcuts import (
    render,
    redirect,
    get_object_or_404
)

from django.contrib.auth import (
    authenticate,
    login,
    logout
)

from django.contrib.auth.decorators import (
    login_required
)

from django.utils import timezone

from django.db.models.deletion import ProtectedError

from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    HRFlowable
)

from .models import (
    Auditoria,
    Convenio,
    EvolucaoClinica,
    ProntuarioClinico,
    DocumentoClinico,
    TemplateDocumento,
    Paciente,
    Tratamento,
    Anamnese,
    Procedimento,
    Orcamento,
    ItemOrcamento,
    AnexoPaciente,
    ConfiguracaoClinica,
    Medicamento,
    Receita,
    ModeloReceita,
    Exame,
    SolicitacaoExame,
    Fornecedor,
    Produto,
    Compra,
    ItemCompra,
    MovimentacaoEstoque,
    LoteProduto,
    ContaPagar,
    ContaReceber,
    Caixa, 
    Perfil,
    PerfilUsuario,
    Modulo,
    Permissao,
)

from .permissoes_padrao import aplicar_permissoes_padrao
from .permissions import tem_permissao
from .permissions import filtrar_escopo
from .permissions import permissao_required

from .forms import (
    ProcedimentoForm,
    ItemOrcamentoForm,
    ConvenioForm,
    PerfilForm
)

from .services import registrar_auditoria

from datetime import timedelta

from django.contrib.auth.models import User
from django.contrib import messages

from .permissions import perfil_required
from django.contrib.auth import update_session_auth_hash

from decimal import Decimal

from django.db.models import (
    Sum,
    Max,
    Min,
)

from django.db.models import Count, Sum, F, DecimalField, ExpressionWrapper
from django.db.models.functions import ExtractMonth

from agenda.models import Profissional
from agenda.models import Agendamento

from django.db import transaction

from django.core.paginator import Paginator

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from openpyxl.styles import Border, Side

from django.contrib.auth.models import User

# =========================================
# LOGIN
# =========================================

def login_view(request):

    erro = None

    if request.method == "POST":

        username = request.POST.get("username")

        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password,
        )

        print("=" * 60)
        print("Username digitado:", username)

        if user:
            print("Usuário autenticado:", user.username)
        else:
            print("Falha na autenticação")

        print("=" * 60)
        if user is not None:

            login(request, user)

            # =========================================
            # AUDITORIA
            # =========================================

            registrar_auditoria(

                request=request,

                usuario=user,

                modulo="Autenticação",

                acao="login",

                descricao=f"Login realizado pelo usuário '{user.username}'.",

                nivel="info",

            )

            return redirect("dashboard")

        else:

            erro = "Usuário ou senha inválidos."

    return render(

        request,

        "accounts/login.html",

        {

            "erro": erro,

        },

    )

from .permissions import permissao_required

# =========================================
# DASHBOARD
# =========================================

@login_required(login_url="/")
@permissao_required("Dashboard")
def dashboard_view(request):

    from decimal import Decimal
    from collections import defaultdict

    from django.db.models import Sum
    from django.db.models.functions import ExtractMonth
    from django.utils import timezone

    hoje = timezone.now().date()

    # =========================================
    # PERFIL DO USUÁRIO
    # =========================================

    perfil_usuario = getattr(request.user, "perfil", None)

    perfil_acesso = getattr(perfil_usuario, "perfil_acesso", None)

    perfil_nome = perfil_acesso.nome if perfil_acesso else ""

    # =========================================
    # TIPO DE DASHBOARD
    # =========================================

    TIPOS_DASHBOARD = {
        "Administrador": "admin",
        "Gestor": "gestor",
        "Dentista": "dentista",
        "Secretária": "secretaria",
        "Auxiliar de Saúde Bucal": "acd",
        "Contabilidade": "contabilidade",
        "Marketing": "marketing",
        "Auditoria": "auditoria",
    }

    dashboard_tipo = TIPOS_DASHBOARD.get(
        perfil_nome,
        "admin"
    )

        # =========================================
    # CONFIGURAÇÃO DO DASHBOARD
    # =========================================

    dashboard_config = {

        # Cards Superiores
        "mostrar_pacientes": True,
        "mostrar_receber": True,
        "mostrar_pagar": True,
        "mostrar_saldo_caixa": True,

        # Segunda linha
        "mostrar_recebimentos": True,
        "mostrar_pagamentos": True,
        "mostrar_lucro": True,
        "mostrar_fornecedores": True,

        # Conteúdo
        "mostrar_faturamento": True,
        "mostrar_consultas": True,
        "mostrar_aniversariantes": True,
        "mostrar_ranking": True,
        "mostrar_movimentacoes": True,

    }

    # =========================================
    # PERSONALIZAÇÃO POR PERFIL
    # =========================================

    # Administrador e Gestor utilizam o Dashboard completo.

    # =========================================
    # DENTISTA
    # =========================================

    if dashboard_tipo == "dentista":

        dashboard_config.update({

            # Cards
            "mostrar_pacientes": True,
            "mostrar_receber": True,
            "mostrar_pagar": False,
            "mostrar_saldo_caixa": False,

            # Segunda linha
            "mostrar_recebimentos": False,
            "mostrar_pagamentos": False,
            "mostrar_lucro": False,
            "mostrar_fornecedores": False,

            # Conteúdo
            "mostrar_faturamento": False,
            "mostrar_consultas": True,
            "mostrar_aniversariantes": True,
            "mostrar_ranking": True,
            "mostrar_movimentacoes": False,

        })

    # =========================================
    # PACIENTES
    # =========================================

    total_pacientes = Paciente.objects.filter(
        ativo=True
    ).count()

    # =========================================
    # FORNECEDORES
    # =========================================

    fornecedores_ativos = Fornecedor.objects.filter(
        ativo=True
    ).count()

    # =========================================
    # CONTAS A RECEBER
    # =========================================

    contas_receber = ContaReceber.objects.filter(
        status="PENDENTE"
    )

    receber_pendente = (
        contas_receber.aggregate(
            total=Sum("valor")
        )["total"]
        or Decimal("0.00")
    )

    # =========================================
    # CONTAS A PAGAR
    # =========================================

    contas_pagar = ContaPagar.objects.filter(
        status="PENDENTE"
    )

    pagar_pendente = (
        contas_pagar.aggregate(
            total=Sum("valor")
        )["total"]
        or Decimal("0.00")
    )

    # =========================================
    # CAIXA
    # =========================================

    caixa_entradas = Caixa.objects.filter(
        tipo="ENTRADA"
    )

    caixa_saidas = Caixa.objects.filter(
        tipo="SAIDA"
    )

    total_entradas = (
        caixa_entradas.aggregate(
            total=Sum("valor")
        )["total"]
        or Decimal("0.00")
    )

    total_saidas = (
        caixa_saidas.aggregate(
            total=Sum("valor")
        )["total"]
        or Decimal("0.00")
    )

    saldo_caixa = total_entradas - total_saidas

        # =========================================
    # MOVIMENTAÇÃO HOJE
    # =========================================

    movimentacao_entradas_hoje = Caixa.objects.filter(
        tipo="ENTRADA",
        data=hoje
    )

    movimentacao_saidas_hoje = Caixa.objects.filter(
        tipo="SAIDA",
        data=hoje
    )

    entradas_hoje = (
        movimentacao_entradas_hoje.aggregate(
            total=Sum("valor")
        )["total"]
        or Decimal("0.00")
    )

    saidas_hoje = (
        movimentacao_saidas_hoje.aggregate(
            total=Sum("valor")
        )["total"]
        or Decimal("0.00")
    )

    lucro_hoje = entradas_hoje - saidas_hoje

    # =========================================
    # ÚLTIMAS MOVIMENTAÇÕES
    # =========================================

    movimentacoes = Caixa.objects.order_by(
        "-data",
        "-id"
    )

    ultimas_movimentacoes = movimentacoes[:20]

        # =========================================
    # GRÁFICO FINANCEIRO
    # =========================================

    meses = [
        'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
        'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'
    ]

    entradas_mes = [0] * 12
    saidas_mes = [0] * 12

    caixa_entradas_grafico = Caixa.objects.filter(
        tipo='ENTRADA'
    )

    caixa_saidas_grafico = Caixa.objects.filter(
        tipo='SAIDA'
    )

    entradas_grafico = (
        caixa_entradas_grafico
        .annotate(mes=ExtractMonth('data'))
        .values('mes')
        .annotate(total=Sum('valor'))
    )

    for item in entradas_grafico:

        if item['mes']:

            entradas_mes[item['mes'] - 1] = float(item['total'])

    saidas_grafico = (
        caixa_saidas_grafico
        .annotate(mes=ExtractMonth('data'))
        .values('mes')
        .annotate(total=Sum('valor'))
    )

    for item in saidas_grafico:

        if item['mes']:

            saidas_mes[item['mes'] - 1] = float(item['total'])

    # =========================================
    # PRÓXIMAS CONSULTAS
    # =========================================

    consultas = (
        Agendamento.objects.filter(
            status__in=[
                'agendado',
                'confirmado',
                'atendimento'
            ],
            data__gte=hoje
        )
        .select_related(
            'paciente',
            'procedimento'
        )
        .order_by(
            'data',
            'hora_inicio'
        )
    )

        # =========================================
    # FILTRO DO DENTISTA
    # =========================================

    print("=" * 60)
    print("Dashboard Tipo:", dashboard_tipo)
    print("Perfil:", perfil_nome)
    print("Usuário:", request.user.username)
    print("Profissional:", getattr(request.user, "profissional", None))
    print("=" * 60)

    if dashboard_tipo == "dentista":

        print(">>> ENTROU NO FILTRO DO DENTISTA <<<")

        profissional = getattr(
            request.user,
            "profissional",
            None
        )

        if profissional:

            print("Profissional encontrado:", profissional)

            consultas = consultas.filter(
                profissional=profissional
            )

            print("Total consultas:", consultas.count())

        else:

            print("Profissional NÃO encontrado")

            consultas = consultas.none()

    proximas_consultas = consultas[:5]

    # =========================================
    # ANIVERSARIANTES DO MÊS
    # =========================================

    aniversariantes = (
        Paciente.objects.filter(
            ativo=True,
            nascimento__month=hoje.month
        )
        .order_by(
            "nascimento__day",
            "nome"
        )
    )

        # =========================================
    # RANKING DOS DENTISTAS
    # =========================================

    status_validos = [
        "aprovado",
        "em_execucao",
        "finalizado",
    ]

    ranking_dict = defaultdict(Decimal)

    tratamentos = (
        Tratamento.objects.filter(
            dentista__isnull=False
        )
        .select_related(
            "dentista"
        )
        .prefetch_related(
            "orcamentos"
        )
    )

    # =========================================
    # FILTRO POR PERFIL
    # =========================================

    if dashboard_tipo == "dentista":

        tratamentos = tratamentos.filter(
            dentista=request.user
        )

    # =========================================
    # CÁLCULO DA PRODUÇÃO
    # =========================================

    for tratamento in tratamentos:

        orcamentos = tratamento.orcamentos.filter(
            status__in=status_validos
        )

        for orcamento in orcamentos:

            ranking_dict[
                tratamento.dentista
            ] += orcamento.total

    ranking_dentistas = sorted(
        ranking_dict.items(),
        key=lambda item: item[1],
        reverse=True,
    )

    # Administrador e Gestor veem o Top 5.
    if dashboard_tipo != "dentista":

        ranking_dentistas = ranking_dentistas[:5]

        # =========================================
    # CONTEXT
    # =========================================

    context = {

        # =========================================
        # PERFIL
        # =========================================

        "perfil_nome": perfil_nome,
        "dashboard_tipo": dashboard_tipo,
        "dashboard": dashboard_config,

        # =========================================
        # INDICADORES
        # =========================================

        "total_pacientes": total_pacientes,
        "fornecedores_ativos": fornecedores_ativos,

        "receber_pendente": receber_pendente,
        "pagar_pendente": pagar_pendente,

        "saldo_caixa": saldo_caixa,

        "total_entradas": total_entradas,
        "total_saidas": total_saidas,

        "entradas_hoje": entradas_hoje,
        "saidas_hoje": saidas_hoje,

        "lucro_hoje": lucro_hoje,

        # =========================================
        # DASHBOARD
        # =========================================

        "ultimas_movimentacoes": ultimas_movimentacoes,
        "proximas_consultas": proximas_consultas,
        "ranking_dentistas": ranking_dentistas,
        "aniversariantes": aniversariantes,

        # =========================================
        # GRÁFICOS
        # =========================================

        "meses": meses,
        "entradas_mes": entradas_mes,
        "saidas_mes": saidas_mes,

    }

    return render(
        request,
        "accounts/dashboard.html",
        context,
    )

# =========================================
# PACIENTES
# =========================================

@login_required(login_url='/')
@permissao_required("pacientes", "visualizar")
def pacientes_view(request):

    # =========================================
    # CADASTRAR PACIENTE
    # =========================================

    if request.method == 'POST':

        Paciente.objects.create(

            # FOTO
            foto=request.FILES.get('foto'),

            # DADOS PESSOAIS
            nome=request.POST.get('nome'),
            cpf=request.POST.get('cpf'),
            rg=request.POST.get('rg'),
            nascimento=request.POST.get('nascimento') or None,
            genero=request.POST.get('genero'),
            estado_civil=request.POST.get('estado_civil'),
            profissao=request.POST.get('profissao'),

            # CONTATO
            telefone=request.POST.get('telefone'),
            whatsapp=request.POST.get('whatsapp'),
            email=request.POST.get('email'),

            # ENDEREÇO
            cep=request.POST.get('cep'),
            endereco=request.POST.get('endereco'),
            numero=request.POST.get('numero'),
            complemento=request.POST.get('complemento'),
            bairro=request.POST.get('bairro'),
            cidade=request.POST.get('cidade'),
            estado=request.POST.get('estado'),

            # CLÍNICO
            convenio=request.POST.get('convenio'),
            carteirinha=request.POST.get('carteirinha'),
            alergias=request.POST.get('alergias'),
            medicamentos=request.POST.get('medicamentos'),
            observacoes=request.POST.get('observacoes'),

            # RESPONSÁVEL
            responsavel=request.POST.get('responsavel'),

            cpf_responsavel=request.POST.get(
                'cpf_responsavel'
            ),

            telefone_responsavel=request.POST.get(
                'telefone_responsavel'
            ),

            dentista_id=request.POST.get('dentista') or None,

        )

        return redirect('pacientes')

    # =========================================
    # LISTAGEM
    # =========================================

    perfil = request.user.perfil.tipo_usuario

    if perfil == 'dentista':

        pacientes = Paciente.objects.filter(
            dentista=request.user
        ).order_by('-id')

    else:

        pacientes = Paciente.objects.all().order_by('-id')

    convenios = Convenio.objects.filter(
        ativo=True
    ).order_by('nome')

    dentistas = User.objects.filter(
        perfil__tipo_usuario='dentista'
    ).order_by('first_name')

    context = {
        'pacientes': pacientes,
        'convenios': convenios,
        'dentistas': dentistas,
    }

    return render(

        request,

        'accounts/pacientes.html',

        context

    )

# =========================================
# PERFIL DO PACIENTE
# =========================================

@login_required(login_url='/')
@permissao_required("pacientes", "visualizar")
def perfil_paciente(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    agendamentos = list(

        paciente.agendamentos.filter(
            status__in=[
                'agendado',
                'confirmado',
                'atendimento'
            ]
        ).order_by(
            '-data',
            '-hora_inicio'
        )[:4]

    )

    agendamentos.reverse()

    return render(

        request,

        'accounts/perfil.html',

        {

            'paciente': paciente,
            'agendamentos': agendamentos,

        }

    )
# =========================================
# LOGOUT
# =========================================

def logout_view(request):

    logout(request)

    return redirect('/')

# =========================================
# NOVO PACIENTE
# =========================================

@login_required(login_url='/')
@permissao_required("pacientes", "inserir")
def novo_paciente(request):

    if request.method == 'POST':

        dentista_id = request.POST.get('dentista')

        dentista = None

        if dentista_id:

            dentista = User.objects.get(
                id=dentista_id
            )

        paciente = Paciente.objects.create(

            dentista=dentista,

            foto=request.FILES.get('foto'),

            nome=request.POST.get('nome'),
            cpf=request.POST.get('cpf'),
            rg=request.POST.get('rg'),
            nascimento=request.POST.get('nascimento') or None,

            genero=request.POST.get('genero'),
            estado_civil=request.POST.get('estado_civil'),
            profissao=request.POST.get('profissao'),

            telefone=request.POST.get('telefone'),
            whatsapp=request.POST.get('whatsapp'),
            email=request.POST.get('email'),

            cep=request.POST.get('cep'),
            endereco=request.POST.get('endereco'),
            numero=request.POST.get('numero'),
            complemento=request.POST.get('complemento'),
            bairro=request.POST.get('bairro'),
            cidade=request.POST.get('cidade'),
            estado=request.POST.get('estado'),

            convenio=request.POST.get('convenio'),
            carteirinha=request.POST.get('carteirinha'),

            alergias=request.POST.get('alergias'),
            medicamentos=request.POST.get('medicamentos'),
            observacoes=request.POST.get('observacoes'),

            responsavel=request.POST.get('responsavel'),
            cpf_responsavel=request.POST.get('cpf_responsavel'),
            telefone_responsavel=request.POST.get(
                'telefone_responsavel'
            )

        )

                # =========================================
        # AUDITORIA
        # =========================================

        print("========== CHEGOU NA AUDITORIA ==========")

        registrar_auditoria(

            request=request,

            modulo="Pacientes",

            acao="cadastro",

            descricao=(
                f"Paciente '{paciente.nome}' cadastrado."
            ),

            objeto_id=paciente.id,

            nivel="info",

        )

        print(
            "========== AUDITORIA GRAVADA =========="
        )

        print(
            'PACIENTE CRIADO:',
            paciente.id
        )

        return redirect(

            'perfil_paciente',

            id=paciente.id

        )

    convenios = Convenio.objects.filter(
        ativo=True
    ).order_by('nome')

    # =========================================
    # DENTISTAS
    # =========================================

    dentistas = User.objects.filter(

        perfil__tipo_usuario='dentista',

        perfil__ativo=True

    ).order_by(

        'first_name'

    )

    return render(

        request,

        'accounts/paciente_form.html',

        {

            'convenios': convenios,

            'dentistas': dentistas,

            'modo': 'novo'

        }

    )


# =========================================
# EDITAR PACIENTE
# =========================================

@login_required(login_url='/')
@permissao_required("pacientes", "editar")
def editar_paciente(request, id):

    paciente = get_object_or_404(

        Paciente,

        id=id

    )

    if request.method == 'POST':

        # FOTO
        if request.FILES.get('foto'):

            paciente.foto = request.FILES.get(
                'foto'
            )

        # DADOS PESSOAIS
        paciente.nome = request.POST.get('nome')
        paciente.cpf = request.POST.get('cpf')
        paciente.rg = request.POST.get('rg')

        paciente.nascimento = (
            request.POST.get('nascimento')
            or None
        )

        paciente.genero = request.POST.get(
            'genero'
        )

        paciente.estado_civil = request.POST.get(
            'estado_civil'
        )

        paciente.profissao = request.POST.get(
            'profissao'
        )

        # CONTATO
        paciente.telefone = request.POST.get(
            'telefone'
        )

        paciente.whatsapp = request.POST.get(
            'whatsapp'
        )

        paciente.email = request.POST.get(
            'email'
        )

        # ENDEREÇO
        paciente.cep = request.POST.get('cep')

        paciente.endereco = request.POST.get(
            'endereco'
        )

        paciente.numero = request.POST.get(
            'numero'
        )

        paciente.complemento = request.POST.get(
            'complemento'
        )

        paciente.bairro = request.POST.get(
            'bairro'
        )

        paciente.cidade = request.POST.get(
            'cidade'
        )

        paciente.estado = request.POST.get(
            'estado'
        )

        # CLÍNICO
        paciente.convenio = request.POST.get(
            'convenio'
        )

        paciente.carteirinha = request.POST.get(
            'carteirinha'
        )

        paciente.alergias = request.POST.get(
            'alergias'
        )

        paciente.medicamentos = request.POST.get(
            'medicamentos'
        )

        paciente.observacoes = request.POST.get(
            'observacoes'
        )

        # RESPONSÁVEL
        paciente.responsavel = request.POST.get(
            'responsavel'
        )

        paciente.cpf_responsavel = request.POST.get(
            'cpf_responsavel'
        )

        paciente.telefone_responsavel = request.POST.get(
            'telefone_responsavel'
        )

        paciente.save()

        # =========================================
        # AUDITORIA
        # =========================================

        registrar_auditoria(

            request=request,

            modulo="Pacientes",

            acao="alteracao",

            descricao=(
                f"Paciente '{paciente.nome}' alterado."
            ),

            objeto_id=paciente.id,

            nivel="info",

        )

        return redirect(
            'pacientes'
        )

    convenios = Convenio.objects.filter(
        ativo=True
    ).order_by('nome')

    return render(

        request,

        'accounts/paciente_form.html',

        {

            'paciente': paciente,

            'convenios': convenios,

            'modo': 'editar'

        }

    )


# =========================================
# EXCLUIR PACIENTE
# =========================================

@login_required(login_url='/')
@permissao_required("pacientes", "excluir")
def excluir_paciente(request, id):

    paciente = get_object_or_404(

        Paciente,

        id=id

    )

    nome = paciente.nome

    try:

        paciente.delete()

        # =========================================
        # AUDITORIA
        # =========================================

        registrar_auditoria(

            request=request,

            modulo="Pacientes",

            acao="exclusao",

            descricao=(
                f"Paciente '{nome}' excluído."
            ),

            objeto_id=id,

            nivel="critico",

        )

        messages.success(

            request,

            "Paciente excluído com sucesso."

        )

    except ProtectedError:

        # =========================================
        # AUDITORIA
        # =========================================

        registrar_auditoria(

            request=request,

            modulo="Pacientes",

            acao="exclusao",

            descricao=(
                f"Tentativa de exclusão do paciente '{nome}' bloqueada por vínculos financeiros."
            ),

            objeto_id=id,

            nivel="erro",

        )

        messages.error(

            request,

            "Este paciente possui movimentações financeiras e não pode ser excluído."

        )

    return redirect(

        "pacientes"

    )

# =========================================
# ALTERAR STATUS PACIENTE
# =========================================

@login_required(login_url='/')
def alterar_status_paciente(request, id):

    paciente = get_object_or_404(

        Paciente,

        id=id

    )

    paciente.ativo = not paciente.ativo

    paciente.save()

    # =========================================
    # AUDITORIA
    # =========================================

    status = (
        "Ativado"
        if paciente.ativo
        else "Desativado"
    )

    registrar_auditoria(

        request=request,

        modulo="Pacientes",

        acao="alteracao",

        descricao=(
            f"Paciente '{paciente.nome}' - Status: {status}."
        ),

        objeto_id=paciente.id,

        nivel="aviso",

    )

    messages.success(

        request,

        'Status do paciente atualizado com sucesso.'

    )

    return redirect(
        'pacientes'
    )
# =========================================
# ODONTOGRAMA
# =========================================

@login_required(login_url='/')
@permissao_required("odontograma", "visualizar")
def odontograma(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

   # =====================================
    # TRATAMENTO ATIVO
    # =====================================

    tratamento = paciente.tratamentos.filter(
        status="ATIVO"
    ).first()

    # =====================================
    # HISTÓRICO DE TRATAMENTOS
    # =====================================

    historico_tratamentos = paciente.tratamentos.filter(
        status="ENCERRADO"
    ).order_by("-data_encerramento")

    # =====================================
    # ORÇAMENTO DO TRATAMENTO ATIVO
    # =====================================

    orcamento = None

    if tratamento:

        orcamento, created = Orcamento.objects.get_or_create(

            paciente=paciente,

            tratamento=tratamento,

            defaults={
                "tratamento": tratamento
            }

        )
    # =========================================
    # SALVAR EVOLUÇÃO
    # =========================================

    if request.method == 'POST':

        print(request.POST)

        procedimento = Procedimento.objects.get(
            id=request.POST.get('procedimento')
        )

        status = request.POST.get('status')

        dente = request.POST.get('dente')

        face = request.POST.get('face')

        descricao = request.POST.get('descricao')

        posicao_icone = request.POST.get(
            'posicao_icone'
        )

        print(
            'POSICAO ESCOLHIDA:',
            posicao_icone
        )

        # =====================================
        # DEFINE VALOR CONFORME CONVÊNIO
        # =====================================

        valor_unitario = procedimento.valor_particular

        if paciente.convenio:

            convenio = Convenio.objects.filter(
                nome=paciente.convenio
            ).first()

            if convenio:

                valor_unitario = (
                    procedimento.valor_particular
                    * convenio.indice
                )

        # =====================================
        # CRIA ITEM DO ORÇAMENTO
        # =====================================

        item = ItemOrcamento.objects.create(

            orcamento=orcamento,

            procedimento=procedimento,

            tipo_local='dente',

            dente=dente,

            face=face,

            posicao_icone=posicao_icone,

            valor_unitario=valor_unitario,

            status=status

        )

        # =====================================
        # REGISTRA EVOLUÇÃO CLÍNICA
        # =====================================

        EvolucaoClinica.objects.create(

            paciente=paciente,

            tratamento=tratamento,

            orcamento=orcamento,

            item_orcamento=item,

            dente=dente,

            face=face,

            posicao_icone=posicao_icone,

            procedimento=procedimento,

            status=status,

            descricao=descricao or ''

        )

        return redirect(
            'odontograma',
            id=paciente.id
        )
    
    # =========================================
    # EVOLUÇÕES CLÍNICAS
    # =========================================

    if tratamento:

        evolucoes = EvolucaoClinica.objects.filter(

            paciente=paciente,

            tratamento=tratamento

        ).order_by('-criado_em')

    else:

        evolucoes = EvolucaoClinica.objects.none()


    # =========================================
    # ITENS DO ORÇAMENTO
    # =========================================

    if orcamento:

        itens_orcamento = ItemOrcamento.objects.filter(

            orcamento=orcamento

        )

    else:

        itens_orcamento = ItemOrcamento.objects.none()
   # =========================================
    # PROCEDIMENTOS
    # =========================================

    procedimentos = Procedimento.objects.all().order_by(

        'categoria',

        'nome'

    )

    # =========================================
    # PROCEDIMENTOS GERAIS
    # =========================================

    procedimentos_gerais = Procedimento.objects.filter(

        tipo__in=['geral', 'hemiarcada']

    ).order_by(

        'categoria',

        'nome'

    )

    context = {

        'paciente': paciente,

        'tratamento': tratamento,

        'historico_tratamentos': historico_tratamentos,

        'orcamento': orcamento,

        'evolucoes': evolucoes,

        'itens_orcamento': itens_orcamento,

        'procedimentos': procedimentos,

        'procedimentos_gerais': procedimentos_gerais,
        # =========================================
        # DENTES PERMANENTES
        # =========================================

        'superiores': [
            '18','17','16','15','14','13','12','11',
            '21','22','23','24','25','26','27','28'
        ],

        'inferiores': [
            '48','47','46','45','44','43','42','41',
            '31','32','33','34','35','36','37','38'
        ],

        # =========================================
        # DENTES DECÍDUOS
        # =========================================

        'dec_superiores': [
            '55','54','53','52','51',
            '61','62','63','64','65'
        ],

        'dec_inferiores': [
            '85','84','83','82','81',
            '71','72','73','74','75'
        ],

    }
    for item in itens_orcamento:

        print(
            f'ITEM {item.id} | DENTE={item.dente} | POSICAO={item.posicao_icone}'
        )

    return render(

        request,

        'accounts/odontograma.html',

        context

    )


@login_required(login_url='/')
@perfil_required(
    "Administrador",
    "Dentista",
    "Auxiliar de Saúde Bucal",
)
def salvar_procedimento_geral(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    # =====================================
    # TRATAMENTO ATIVO
    # =====================================

    tratamento = paciente.tratamentos.filter(
        status='ATIVO'
    ).first()

    if tratamento is None:

        tratamento = Tratamento.objects.create(
            paciente=paciente,
            titulo='Tratamento Inicial'
        )

    # =====================================
    # HISTÓRICO DE TRATAMENTOS
    # =====================================

    historico_tratamentos = paciente.tratamentos.filter(
        status='ENCERRADO'
    ).order_by('-data_encerramento')

    # =====================================
    # ORÇAMENTO DO TRATAMENTO
    # =====================================

    orcamento, created = Orcamento.objects.get_or_create(

        paciente=paciente,

        tratamento=tratamento,

        defaults={
            'tratamento': tratamento
        }

    )

    if request.method == 'POST':

        procedimento = get_object_or_404(

            Procedimento,

            id=request.POST.get('procedimento')

        )

        status = request.POST.get(
            'status'
        ) or 'planejado'

        descricao = request.POST.get(
            'descricao'
        ) or ''

        posicao_icone = request.POST.get(
            'posicao_icone'
        )

        # =====================================
        # CALCULA VALOR CONFORME CONVÊNIO
        # =====================================

        valor_unitario = procedimento.valor_particular

        if paciente.convenio:

            convenio = Convenio.objects.filter(
                nome=paciente.convenio
            ).first()

            if convenio:

                valor_unitario = (
                    procedimento.valor_particular
                    * convenio.indice
                )

        # =====================================
        # CRIA ITEM NO ORÇAMENTO
        # =====================================

        item = ItemOrcamento.objects.create(

            orcamento=orcamento,

            procedimento=procedimento,

            tipo_local='geral',

            valor_unitario=valor_unitario,

            quantidade=1,

            status=status

        )

        # =====================================
        # CRIA EVOLUÇÃO CLÍNICA
        # =====================================

        EvolucaoClinica.objects.create(

            paciente=paciente,

            tratamento=tratamento,

            orcamento=orcamento,

            item_orcamento=item,

            procedimento=procedimento,

            status=status,

            descricao=descricao

        )

        messages.success(

            request,

            'Procedimento salvo com sucesso.'

        )

        return redirect(

            'odontograma',

            id=paciente.id

        )

    return redirect(

        'odontograma',

        id=paciente.id

    )   

@login_required(login_url='/')
@permissao_required("anamnese", "visualizar")
def anamnese(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )



    anamnese, created = Anamnese.objects.get_or_create(
        paciente=paciente
    )

    if request.method == 'POST':

        # =========================================
        # QUEIXA PRINCIPAL
        # =========================================

        anamnese.queixa_principal = request.POST.get(
            'queixa_principal', ''
        )
        # =========================================
        # HISTÓRIA MÉDICA
        # =========================================

        anamnese.hipertenso = 'hipertenso' in request.POST
        anamnese.diabetico = 'diabetico' in request.POST
        anamnese.cardiopatia = 'cardiopatia' in request.POST
        anamnese.asma = 'asma' in request.POST
        anamnese.bronquite = 'bronquite' in request.POST
        anamnese.anemia = 'anemia' in request.POST
        anamnese.hepatite = 'hepatite' in request.POST

        anamnese.rinite = 'rinite' in request.POST
        anamnese.sinusite = 'sinusite' in request.POST

        anamnese.problema_renal = (
            'problema_renal' in request.POST
        )

        anamnese.sangramento_excessivo = (
            'sangramento_excessivo' in request.POST
        )

        anamnese.alergico = 'alergico' in request.POST

        anamnese.alergias = request.POST.get(
            'alergias', ''
        )

        anamnese.fumante = 'fumante' in request.POST
        anamnese.gravida = 'gravida' in request.POST

        anamnese.historico_medico = request.POST.get(
            'historico_medico', ''
        )

        # NOVOS CAMPOS

        if hasattr(anamnese, 'anticoagulante'):
            anamnese.anticoagulante = (
                'anticoagulante' in request.POST
            )

        if hasattr(anamnese, 'bisfosfonato'):
            anamnese.bisfosfonato = (
                'bisfosfonato' in request.POST
            )

        if hasattr(anamnese, 'marcapasso'):
            anamnese.marcapasso = (
                'marcapasso' in request.POST
            )

        if hasattr(anamnese, 'cancer'):
            anamnese.cancer = (
                'cancer' in request.POST
            )

        if hasattr(anamnese, 'hipotireoidismo'):
            anamnese.hipotireoidismo = (
                'hipotireoidismo' in request.POST
            )

        if hasattr(anamnese, 'hipertireoidismo'):
            anamnese.hipertireoidismo = (
                'hipertireoidismo' in request.POST
            )

        # =========================================
        # MEDICAMENTOS
        # =========================================

        anamnese.usa_medicamento = (
            'usa_medicamento' in request.POST
        )

        anamnese.medicamentos = request.POST.get(
            'medicamentos', ''
        )

        anamnese.antibioticos = request.POST.get(
            'antibioticos', ''
        )

        anamnese.antiinflamatorios = request.POST.get(
            'antiinflamatorios', ''
        )

        anamnese.analgesicos = request.POST.get(
            'analgesicos', ''
        )

        # =========================================
        # CIRURGIAS
        # =========================================

        anamnese.cirurgia = (
            'cirurgia' in request.POST
        )

        anamnese.cirurgias = request.POST.get(
            'cirurgias', ''
        )

        anamnese.hospitalizado = (
            'hospitalizado' in request.POST
        )

        anamnese.hospitalizacao = request.POST.get(
            'hospitalizacao', ''
        )

        anamnese.transfusao_sangue = (
            'transfusao_sangue' in request.POST
        )

        # =========================================
        # HISTÓRIA ODONTOLÓGICA
        # =========================================

        anamnese.primeira_consulta = (
            'primeira_consulta' in request.POST
        )

        anamnese.experiencia_odontologica = request.POST.get(
            'experiencia_odontologica', ''
        )

        anamnese.abandono_tratamento = (
            'abandono_tratamento' in request.POST
        )

        anamnese.medo_dentista = (
            'medo_dentista' in request.POST
        )

        anamnese.anestesia_reacao = request.POST.get(
            'anestesia_reacao', ''
        )

        anamnese.sangramento_gengival = (
            'sangramento_gengival' in request.POST
        )

        anamnese.sensibilidade = (
            'sensibilidade' in request.POST
        )

        anamnese.dor_mastigar = (
            'dor_mastigar' in request.POST
        )

        # =========================================
        # HIGIENE ORAL
        # =========================================

        anamnese.frequencia_escovacao = request.POST.get(
            'frequencia_escovacao', ''
        )

        anamnese.usa_fio_dental = (
            'usa_fio_dental' in request.POST
        )

        anamnese.usa_enxaguante = (
            'usa_enxaguante' in request.POST
        )

        anamnese.escova_lingua = (
            'escova_lingua' in request.POST
        )

        # =========================================
        # HÁBITOS
        # =========================================

        anamnese.bruxismo = 'bruxismo' in request.POST
        anamnese.ronco = 'ronco' in request.POST

        anamnese.respiracao_bucal = (
            'respiracao_bucal' in request.POST
        )

        anamnese.roer_unhas = (
            'roer_unhas' in request.POST
        )

        anamnese.morde_objetos = (
            'morde_objetos' in request.POST
        )

        anamnese.chupeta = 'chupeta' in request.POST

        anamnese.succao_dedo = (
            'succao_dedo' in request.POST
        )

        anamnese.baba_travesseiro = (
            'baba_travesseiro' in request.POST
        )

        anamnese.dorme_boca_aberta = (
            'dorme_boca_aberta' in request.POST
        )

        anamnese.morde_labios = (
            'morde_labios' in request.POST
        )

        # =========================================
        # HÁBITOS ALIMENTARES
        # =========================================

        anamnese.belisca_refeicoes = (
            'belisca_refeicoes' in request.POST
        )

        anamnese.alimentacao_cariogenica = (
            'alimentacao_cariogenica' in request.POST
        )

        anamnese.tipo_alimentacao = request.POST.get(
            'tipo_alimentacao', ''
        )

        # =========================================
        # PERFIL COMPORTAMENTAL
        # =========================================

        anamnese.ansioso = 'ansioso' in request.POST
        anamnese.agitado = 'agitado' in request.POST
        anamnese.calmo = 'calmo' in request.POST

        anamnese.comunicativo = (
            'comunicativo' in request.POST
        )

        anamnese.retraido = (
            'retraido' in request.POST
        )

        anamnese.introvertido = (
            'introvertido' in request.POST
        )

        anamnese.extrovertido = (
            'extrovertido' in request.POST
        )

        # =========================================
        # OBSERVAÇÕES
        # =========================================

        anamnese.observacoes = request.POST.get(
            'observacoes', ''
        )

        anamnese.save()

    context = {

        'paciente': paciente,
        'anamnese': anamnese

    }

    return render(

        request,
        'accounts/anamnese.html',
        context

    )

@login_required(login_url='/')
@permissao_required("anamnese", "visualizar")
def ficha_clinica(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    evolucoes = EvolucaoClinica.objects.filter(
        paciente=paciente
    ).order_by('-criado_em')

    prontuarios = ProntuarioClinico.objects.filter(
        paciente=paciente
    )

    documentos = DocumentoClinico.objects.filter(
        paciente=paciente
    )

    anexos = AnexoPaciente.objects.filter(
        paciente=paciente
    )

    receitas = Receita.objects.filter(
        paciente=paciente
    )

    exames = Exame.objects.filter(
        paciente=paciente
    )

    solicitacoes = SolicitacaoExame.objects.filter(
        paciente=paciente
    )

    # =========================================
    # PROCEDIMENTOS DO ORÇAMENTO
    # =========================================

    itens_orcamento = (
        ItemOrcamento.objects
        .filter(
            orcamento__paciente=paciente
        )
        .select_related('procedimento')
        .order_by('-id')
    )

    # =========================================
    # RESUMO CLÍNICO
    # =========================================

    total_procedimentos = evolucoes.count()

    realizados = evolucoes.filter(
        status='realizado'
    ).count()

    planejados = evolucoes.filter(
        status='planejado'
    ).count()

    andamento = evolucoes.filter(
        status='andamento'
    ).count()

    ultima_evolucao = evolucoes.first()

    # =========================================
    # NOVO REGISTRO CLÍNICO
    # =========================================

    if request.method == 'POST':

        ProntuarioClinico.objects.create(

            paciente=paciente,

            titulo=request.POST.get(
                'titulo'
            ),

            anotacao=request.POST.get(
                'anotacao'
            )

        )

        return redirect(
            'ficha_clinica',
            id=paciente.id
        )

    context = {

        'paciente': paciente,

        'evolucoes': evolucoes,

        'prontuarios': prontuarios,

        'anexos': anexos,

        'documentos': documentos,

        'receitas': receitas,

        'exames': exames,

        'solicitacoes': solicitacoes,

        'itens_orcamento': itens_orcamento,

        'total_procedimentos': total_procedimentos,

        'realizados': realizados,

        'planejados': planejados,

        'andamento': andamento,

        'ultima_evolucao': ultima_evolucao

    }

    return render(

        request,

        'accounts/ficha_clinica.html',

        context

    )

# =========================================
# UPLOAD ANEXO
# =========================================

@login_required(login_url='/')
@perfil_required(
    "Administrador",
    "Dentista",
)
def upload_anexo(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    if request.method == 'POST':

        arquivo = request.FILES.get(
            'arquivo'
        )

        descricao = request.POST.get(
            'descricao'
        )

        if arquivo:

            AnexoPaciente.objects.create(

                paciente=paciente,

                descricao=descricao,

                arquivo=arquivo

            )

    return redirect(
        'ficha_clinica',
        id=paciente.id
    )

# =========================================
# PROCEDIMENTOS
# =========================================

@login_required(login_url='/')
@permissao_required("procedimentos")
def procedimentos(request):

    procedimentos = Procedimento.objects.all().order_by(
        'ordem',
        'nome'
    )

    # =========================================
    # ÍCONES DISPONÍVEIS
    # =========================================

    mini_path = os.path.join(

        settings.BASE_DIR,
        'static',
        'img',
        'procedimentos',
        'mini'

    )

    full_path = os.path.join(

        settings.BASE_DIR,
        'static',
        'img',
        'procedimentos',
        'full'

    )

    mini_icons = []
    full_icons = []

    # MINI

    if os.path.exists(mini_path):

        mini_icons = [

            arquivo

            for arquivo in os.listdir(mini_path)

            if arquivo.lower().endswith((

                '.png',
                '.svg',
                '.webp',
                '.jpg',
                '.jpeg'

            ))

        ]

    # FULL

    if os.path.exists(full_path):

        full_icons = [

            arquivo

            for arquivo in os.listdir(full_path)

            if arquivo.lower().endswith((

                '.png',
                '.svg',
                '.webp',
                '.jpg',
                '.jpeg'

            ))

        ]

    # REMOVE DUPLICADOS

    icones = sorted(

        list(

            set(
                mini_icons + full_icons
            )

        )

    )

    print(icones)

        # =========================================
    # FORMULÁRIO
    # =========================================

    form = ProcedimentoForm()

    # =========================================
    # SALVAR
    # =========================================

    if request.method == 'POST':

        form = ProcedimentoForm(request.POST)

        if form.is_valid():

            procedimento = form.save(commit=False)

            if not procedimento.valor_convenio:

                procedimento.valor_convenio = 0

            procedimento.save()

            return redirect('procedimentos')

        else:

            print(form.errors)

    context = {

        'form': form,
        'procedimentos': procedimentos,
        'icones': icones

    }

    return render(

        request,

        'accounts/procedimentos.html',

        context

    )

# =========================================
# EDITAR PROCEDIMENTO
# =========================================

@login_required(login_url='/')
@permissao_required("procedimentos", "editar")
def editar_procedimento(request, id):

    procedimento = get_object_or_404(
        Procedimento,
        id=id
    )

    procedimentos = Procedimento.objects.all().order_by(
        'ordem',
        'nome'
    )

    mini_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'img',
        'procedimentos',
        'mini'
    )

    full_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'img',
        'procedimentos',
        'full'
    )

    mini_icons = []
    full_icons = []

    if os.path.exists(mini_path):
        mini_icons = [
            arquivo
            for arquivo in os.listdir(mini_path)
            if arquivo.lower().endswith((
                '.png',
                '.svg',
                '.webp',
                '.jpg',
                '.jpeg'
            ))
        ]

    if os.path.exists(full_path):
        full_icons = [
            arquivo
            for arquivo in os.listdir(full_path)
            if arquivo.lower().endswith((
                '.png',
                '.svg',
                '.webp',
                '.jpg',
                '.jpeg'
            ))
        ]

    icones = sorted(list(set(mini_icons + full_icons)))

    if request.method == 'POST':

        form = ProcedimentoForm(
            request.POST,
            instance=procedimento
        )

        if form.is_valid():

            form.save()

            return redirect('procedimentos')

    else:

        form = ProcedimentoForm(
            instance=procedimento
        )

    context = {

        'form': form,
        'procedimentos': procedimentos,
        'icones': icones,
        'editando': True

    }

    return render(
        request,
        'accounts/procedimentos.html',
        context
    )

# =========================================
# EXCLUIR PROCEDIMENTO
# =========================================

@login_required(login_url='/')
@permissao_required("procedimentos", "excluir")
def excluir_procedimento(request, id):

    procedimento = get_object_or_404(

        Procedimento,

        id=id

    )

    procedimento.delete()

    return redirect('procedimentos')

# =========================================
# ORÇAMENTO
# =========================================

@login_required(login_url='/')
def orcamento(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    # =========================================
    # BUSCA O ÚLTIMO ORÇAMENTO
    # =========================================

    orcamento = (
        Orcamento.objects
        .filter(paciente=paciente)
        .order_by("-id")
        .first()
    )

    if orcamento is None:

        tratamento = paciente.tratamentos.filter(
            status="ATIVO"
        ).first()

        if tratamento is None:

            tratamento = Tratamento.objects.create(
                paciente=paciente,
                titulo="Tratamento Inicial"
            )

            # =====================================
            # ORÇAMENTO DO TRATAMENTO
            # =====================================

            orcamento, created = Orcamento.objects.get_or_create(

                paciente=paciente,

                tratamento=tratamento,

                defaults={
                    'tratamento': tratamento
                }

            )

        orcamento = Orcamento.objects.create(
            paciente=paciente,
            tratamento=tratamento
        )

    # =========================================
    # FORMULÁRIO
    # =========================================

    item_form = ItemOrcamentoForm()

    # =========================================
    # POST
    # =========================================

    if request.method == "POST":

        # =====================================
        # SALVAR FINANCEIRO
        # =====================================

        if "salvar_financeiro" in request.POST:

            print("=" * 60)
            print(request.POST)
            print("=" * 60)

            try:

                entrada = (
                    request.POST.get("entrada", "0")
                    .replace(",", ".")
                    .strip()
                )

                parcelas = request.POST.get(
                    "parcelas",
                    "1"
                )

                forma_pagamento = request.POST.get(
                    "forma_pagamento",
                    "pix"
                )

                orcamento.entrada = Decimal(
                    entrada or "0"
                )

                orcamento.parcelas = max(
                    1,
                    int(parcelas or 1)
                )

                orcamento.forma_pagamento = forma_pagamento

                orcamento.save()

                # Atualiza o objeto em memória
                orcamento.refresh_from_db()

                messages.success(
                    request,
                    "Dados financeiros salvos com sucesso."
                )

            except Exception as erro:

                messages.error(
                    request,
                    f"Erro ao salvar dados financeiros: {erro}"
                )

            return redirect(
                "orcamento",
                id=paciente.id
            )

        # =====================================
        # ADICIONAR PROCEDIMENTO
        # =====================================

        elif "adicionar_item" in request.POST:

            item_form = ItemOrcamentoForm(
                request.POST
            )

            if item_form.is_valid():

                item = item_form.save(
                    commit=False
                )

                item.orcamento = orcamento

                item.dente = request.POST.get(
                    "dente"
                )

                item.face = request.POST.get(
                    "face"
                )

                item.status = "planejado"

                item.valor_unitario = (
                    item.procedimento.valor_particular
                )

                item.save()

                messages.success(
                    request,
                    "Procedimento adicionado com sucesso."
                )

                return redirect(
                    "orcamento",
                    id=paciente.id
                )

            messages.error(
                request,
                "Verifique os dados do procedimento."
            )

        # =========================================
    # CONTEXT
    # =========================================

    context = {

        'paciente': paciente,

        'orcamento': orcamento,

        'item_form': item_form,

    }

    return render(

        request,

        'accounts/orcamento.html',

        context

    )


# =========================================
# APROVAR ORÇAMENTO
# =========================================

@login_required
@perfil_required(
    "Administrador",
    "Secretária",
)
def aprovar_orcamento(request, id):
    orcamento = get_object_or_404(
        Orcamento,
        id=id
    )

    # =========================================


    # =========================================
    # VERIFICA STATUS
    # =========================================

    if orcamento.status == 'aprovado':

        messages.warning(
            request,
            'Este orçamento já foi aprovado.'
        )

        return redirect(
            'orcamento',
            id=orcamento.paciente.id
        )

    # =========================================
    # APROVA ORÇAMENTO
    # =========================================

    orcamento.status = 'aprovado'
    orcamento.save(update_fields=['status'])

    # =========================================
    # CÁLCULOS
    # =========================================

    valor_entrada = orcamento.entrada or Decimal('0.00')

    quantidade_parcelas = max(
        1,
        int(orcamento.parcelas or 1)
    )

    saldo = orcamento.total - valor_entrada

    if saldo < 0:
        saldo = Decimal('0.00')

    valor_parcela = (
        saldo /
        quantidade_parcelas
    )

    # =========================================
    # GERA CONTAS A RECEBER
    # =========================================

    # Garante que apenas orçamentos aprovados gerem contas
    if orcamento.status != "aprovado":

        messages.warning(
            request,
            "Somente orçamentos aprovados podem gerar contas a receber."
        )

        return redirect(
            "orcamento",
            id=orcamento.paciente.id
        )

    # =========================================
    # GERA ENTRADA
    # =========================================

    if valor_entrada > 0:

        ContaReceber.objects.get_or_create(

            paciente=orcamento.paciente,

            orcamento=orcamento,

            parcela=0,

            defaults={

                "descricao": f"Entrada - Orçamento #{orcamento.id}",

                "valor": valor_entrada,

                "total_parcelas": quantidade_parcelas,

                "vencimento": timezone.now().date(),

                "status": "PENDENTE",

            }

        )

    # =========================================
    # GERA PARCELAS
    # =========================================

    for numero in range(1, quantidade_parcelas + 1):

        ContaReceber.objects.get_or_create(

            paciente=orcamento.paciente,

            orcamento=orcamento,

            parcela=numero,

            defaults={

                "descricao": f"Orçamento #{orcamento.id}",

                "valor": valor_parcela,

                "total_parcelas": quantidade_parcelas,

                "vencimento": (
                    timezone.now().date()
                    + timedelta(days=30 * numero)
                ),

                "status": "PENDENTE",

            }

        )

    # =========================================
    # SUCESSO
    # =========================================

    messages.success(

        request,

        "Orçamento aprovado e contas a receber geradas com sucesso."

    )

    return redirect(

        "orcamento",

        id=orcamento.paciente.id

    )

# =========================================
# EXCLUIR ORÇAMENTO
# =========================================

@login_required
@perfil_required(
    "Administrador",
    "Secretária",
)
def excluir_orcamento(request, id):

    orcamento = get_object_or_404(
        Orcamento,
        id=id
    )

    paciente_id = orcamento.paciente.id

    with transaction.atomic():

        # =========================================
        # REMOVE CONTAS A RECEBER
        # =========================================

        ContaReceber.objects.filter(
            orcamento=orcamento
        ).delete()

        # =========================================
        # REMOVE EVOLUÇÕES CLÍNICAS
        # =========================================

        for item in orcamento.itens.all():

            EvolucaoClinica.objects.filter(

                paciente=orcamento.paciente,

                procedimento=item.procedimento,

                dente=item.dente,

            ).delete()

        # =========================================
        # REMOVE ITENS DO ORÇAMENTO
        # =========================================

        orcamento.itens.all().delete()

        # =========================================
        # REMOVE O ORÇAMENTO
        # =========================================

        orcamento.delete()

    messages.success(
        request,
        'Orçamento excluído com sucesso.'
    )

    return redirect(
        'perfil_paciente',
        id=paciente_id
    )


# =========================================
# CONVÊNIOS
# =========================================

@login_required(login_url='/')
def convenios(request):

    convenios = (
        Convenio.objects
        .all()
        .order_by('nome')
    )

    form = ConvenioForm()

    if request.method == 'POST':

        form = ConvenioForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                'Convênio cadastrado com sucesso.'
            )

            return redirect('convenios')

    context = {

        'form': form,
        'convenios': convenios,

    }

    return render(

        request,

        'accounts/convenios.html',

        context

    )


# =========================================
# EDITAR CONVÊNIO
# =========================================

@login_required(login_url='/')
def editar_convenio(request, id):

    convenio = get_object_or_404(
        Convenio,
        id=id
    )

    if request.method == 'POST':

        form = ConvenioForm(

            request.POST,

            instance=convenio

        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                'Convênio atualizado com sucesso.'
            )

            return redirect('convenios')

    else:

        form = ConvenioForm(
            instance=convenio
        )

    convenios = (
        Convenio.objects
        .all()
        .order_by('nome')
    )

    context = {

        'form': form,
        'convenios': convenios,

    }

    return render(

        request,

        'accounts/convenios.html',

        context

    )


# =========================================
# EXCLUIR CONVÊNIO
# =========================================

@login_required(login_url='/')
def excluir_convenio(request, id):

    convenio = get_object_or_404(
        Convenio,
        id=id
    )

    convenio.delete()

    messages.success(
        request,
        'Convênio excluído com sucesso.'
    )

    return redirect(
        'convenios'
    )

# =========================================
# EXCLUIR ITEM ORÇAMENTO
# =========================================

@login_required(login_url='/')
@permissao_required("odontograma", "excluir")
def excluir_item_orcamento(request, id):

    item = get_object_or_404(
        ItemOrcamento,
        id=id
    )

    paciente = item.orcamento.paciente

    # =========================================
    # REMOVE DA EVOLUÇÃO CLÍNICA
    # =========================================

    EvolucaoClinica.objects.filter(

        paciente=paciente,

        procedimento=item.procedimento,

        dente=item.dente

    ).delete()

    # =========================================
    # REMOVE DO ORÇAMENTO
    # =========================================

    item.delete()

    messages.success(
        request,
        'Procedimento removido do orçamento.'
    )

    return redirect(
        'orcamento',
        id=paciente.id
    )


# =========================================
# EDITAR ITEM ORÇAMENTO
# =========================================

@login_required(login_url='/')
def editar_item_orcamento(request, id):

    item = get_object_or_404(

        ItemOrcamento,

        id=id

    )

    if request.method == 'POST':

        item.procedimento_id = request.POST.get(
            'procedimento'
        )

        item.dente = request.POST.get(
            'dente'
        )

        item.face = request.POST.get(
            'face'
        )

        item.status = request.POST.get(
            'status',
            'planejado'
        )

        item.quantidade = request.POST.get(
            'quantidade',
            1
        )

        procedimento = get_object_or_404(

            Procedimento,

            id=item.procedimento_id

        )

        item.valor_unitario = (
            procedimento.valor_particular
        )

        item.save()

        messages.success(

            request,

            'Procedimento atualizado com sucesso.'

        )

        return redirect(

            'orcamento',

            id=item.orcamento.paciente.id

        )

    procedimentos = (
        Procedimento.objects
        .all()
        .order_by('nome')
    )

    context = {

        'item': item,

        'procedimentos': procedimentos,

    }

    return render(

        request,

        'accounts/editar_item_orcamento.html',

        context

    )

# =========================================
# PDF ORÇAMENTO
# =========================================

@login_required(login_url='/')
def gerar_pdf_orcamento(request, id):

    orcamento = get_object_or_404(
        Orcamento,
        id=id
    )

    paciente = orcamento.paciente

    validade = (
        orcamento.criado_em.date()
        + timedelta(days=30)
    )

    config = ConfiguracaoClinica.objects.first()

    itens = (
        orcamento.itens
        .select_related('procedimento')
        .order_by('id')
    )

    evolucoes = EvolucaoClinica.objects.filter(
        paciente=paciente
    )

    parcelas_financeiras = (
        ContaReceber.objects
        .filter(orcamento=orcamento)
        .order_by('parcela', 'vencimento')
    )

    print("=" * 60)
    print("ORÇAMENTO:", orcamento.id)
    print("QTDE PARCELAS:", parcelas_financeiras.count())

    for p in parcelas_financeiras:
        print(
            p.parcela,
            p.vencimento,
            p.valor
        )

    print("=" * 60)

    # =========================================
    # ARCADAS PERMANENTES
    # =========================================

    superiores_permanentes = [
        '18','17','16','15','14','13','12','11',
        '21','22','23','24','25','26','27','28'
    ]

    inferiores_permanentes = [
        '48','47','46','45','44','43','42','41',
        '31','32','33','34','35','36','37','38'
    ]

    # =========================================
    # ARCADAS DECÍDUAS
    # =========================================

    superiores_deciduos = [
        '55','54','53','52','51',
        '61','62','63','64','65'
    ]

    inferiores_deciduos = [
        '85','84','83','82','81',
        '71','72','73','74','75'
    ]

    # =========================================
    # VERIFICA SE EXISTE DENTIÇÃO
    # =========================================

    PERMANENTES = (
        superiores_permanentes +
        inferiores_permanentes
    )

    DECIDUOS = (
        superiores_deciduos +
        inferiores_deciduos
    )

    tem_permanente = evolucoes.filter(
        dente__in=PERMANENTES
    ).exists()

    tem_deciduo = evolucoes.filter(
        dente__in=DECIDUOS
    ).exists()

    # =========================================
    # DENTES COM PROCEDIMENTO
    # =========================================

    dentes_com_procedimento = list({

        str(evolucao.dente)

        for evolucao in evolucoes

        if evolucao.dente

    })

    # =========================================
    # DADOS FINANCEIROS
    # =========================================

    entrada = orcamento.entrada or 0

    parcelas = max(
        1,
        orcamento.parcelas or 1
    )

    saldo = (
        orcamento.total -
        entrada
    )

    valor_parcela = (
        saldo /
        parcelas
    )

    # =========================================
    # CONTEXT
    # =========================================

    context = {

        'orcamento': orcamento,

        'paciente': paciente,

        'itens': itens,

        'evolucoes': evolucoes,

        'config': config,

        'validade': validade,

        'tem_permanente': tem_permanente,

        'tem_deciduo': tem_deciduo,

        'superiores_permanentes': superiores_permanentes,

        'inferiores_permanentes': inferiores_permanentes,

        'superiores_deciduos': superiores_deciduos,

        'inferiores_deciduos': inferiores_deciduos,

        'dentes_com_procedimento': dentes_com_procedimento,

        'parcelas_financeiras': parcelas_financeiras,

        # =====================================
        # FINANCEIRO
        # =====================================

        'entrada': entrada,

        'parcelas': parcelas,

        'forma_pagamento': (
            orcamento.get_forma_pagamento_display()
        ),

        'saldo': saldo,

        'valor_parcela': valor_parcela,

    }

    return render(

        request,

        'accounts/orcamento_pdf.html',

        context

    )

# =========================================
# ALTERAR STATUS PROCEDIMENTO
# =========================================

import json

@login_required(login_url='/')
@permissao_required("odontograma", "editar")
def alterar_status_procedimento(request, id):

    item = get_object_or_404(
        ItemOrcamento,
        id=id
    )

    data = json.loads(request.body)

    novo_status = data.get(
        'status',
        'planejado'
    )

    # =====================================
    # ATUALIZA ITEM DO ORÇAMENTO
    # =====================================

    item.status = novo_status
    item.save()

    # =====================================
    # ATUALIZA EVOLUÇÃO EXISTENTE
    # =====================================

    evolucao = EvolucaoClinica.objects.filter(

        paciente=item.orcamento.paciente,

        procedimento=item.procedimento,

        dente=item.dente,

        face=item.face

    ).order_by('-id').first()

    if evolucao:

        evolucao.status = novo_status
        evolucao.save()

    else:

        EvolucaoClinica.objects.create(

            paciente=item.orcamento.paciente,

            procedimento=item.procedimento,

            dente=item.dente,

            face=item.face,

            posicao_icone=item.posicao_icone,

            status=novo_status,

            descricao=''

        )

    return JsonResponse({

        'sucesso': True

    })

# =========================================
# RODAPÉ PDF
# =========================================

def adicionar_rodape(canvas, doc):

    canvas.saveState()

    pagina = canvas.getPageNumber()

    canvas.setFont(
        "Helvetica",
        9
    )

    canvas.drawRightString(
        190 * mm,
        10 * mm,
        f"Página {pagina}"
    )

    canvas.restoreState()


# =========================================
# PDF PRONTUÁRIO
# =========================================
@login_required(login_url='/')
def imprimir_prontuario(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    prontuarios = ProntuarioClinico.objects.filter(
        paciente=paciente
    ).order_by('-criado_em')

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = f'inline; filename="prontuario_{paciente.id}.pdf"'

    doc = SimpleDocTemplate(
        response,
        topMargin=30,
        bottomMargin=30,
        leftMargin=40,
        rightMargin=40
    )

    styles = getSampleStyleSheet()

    elementos = []

    # =========================================
    # CONFIGURAÇÃO CLÍNICA
    # =========================================

    config = ConfiguracaoClinica.objects.first()

    # =========================================
    # LOGO
    # =========================================

    logo_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'img',
        'logo.png'
    )

    if os.path.exists(logo_path):

        logo = Image(
            logo_path,
            width=240,
            height=90
        )

    logo.hAlign = 'CENTER'

    elementos.append(logo)

    # Pequeno espaço após a logo
    elementos.append(Spacer(1, 8))

    # =========================================
    # NOME DA CLÍNICA
    # =========================================

    estilo_clinica = ParagraphStyle(
        'Clinica',
        parent=styles['Heading2'],
        fontSize=12,
        leading=14,
        alignment=1,  # Centralizado
        spaceAfter=10
    )

    if config and config.nome_clinica:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                <b>{config.nome_clinica}</b>
                </para>
                ''',
                estilo_clinica
            )
        )

    # =========================================
    # TÍTULO
    # =========================================

    elementos.append(

        Paragraph(

            '''
            <para align="center">
            <b>PRONTUÁRIO CLÍNICO</b>
            </para>
            ''',

            styles['Title']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    elementos.append(

        HRFlowable(
            width="100%",
            thickness=1.2,
            color=colors.HexColor('#1e40af')
        )

    )

    elementos.append(
        Spacer(1, 15)
    )

    # =========================================
    # CABEÇALHO
    # =========================================

    elementos.append(Spacer(1, 15))

    elementos.append(
        Paragraph(
            f'<b>Paciente:</b> {paciente.nome}',
            styles['Normal']
        )
    )

    if paciente.cpf:

        elementos.append(
            Paragraph(
                f'<b>CPF:</b> {paciente.cpf}',
                styles['Normal']
            )
        )

    telefone_paciente = (
        paciente.telefone or 'Não informado'
    )

    whatsapp_paciente = (
        paciente.whatsapp or 'Não informado'
    )

    email_paciente = (
        paciente.email or 'Não informado'
    )

    elementos.append(
        Paragraph(
            f'<b>Telefone:</b> {telefone_paciente}',
            styles['Normal']
        )
    )

    elementos.append(
        Paragraph(
            f'<b>WhatsApp:</b> {whatsapp_paciente}',
            styles['Normal']
        )
    )

    elementos.append(
        Paragraph(
            f'<b>E-mail:</b> {email_paciente}',
            styles['Normal']
        )
    )

    agora = timezone.localtime()

    elementos.append(
        Paragraph(
            f'<b>Data de emissão:</b> {agora.strftime("%d/%m/%Y %H:%M")}',
            styles['Normal']
        )
    )

    elementos.append(Spacer(1, 15))

    elementos.append(
        HRFlowable(
            width="100%",
            thickness=1,
            color=colors.grey
        )
    )

    elementos.append(Spacer(1, 15))

    # =========================================
    # DADOS DA CLÍNICA
    # =========================================

    nome_clinica = (
        config.nome_clinica
        if config and config.nome_clinica
        else ''
    )

    telefone_clinica = (
        config.telefone
        if config and config.telefone
        else ''
    )

    whatsapp_clinica = (
        config.whatsapp
        if config and config.whatsapp
        else ''
    )

    email_clinica = (
        config.email
        if config and config.email
        else ''
    )

    # =========================================
    # REGISTROS
    # =========================================

    for item in prontuarios:

        elementos.append(
            Paragraph(
                f'<b>{item.titulo}</b>',
                styles['Heading2']
            )
        ) 

        data_local = timezone.localtime(
            item.criado_em
        )

        elementos.append(
            Paragraph(
                data_local.strftime(
                    '%d/%m/%Y %H:%M'
                ),
                styles['Italic']
            )
        )

        elementos.append(Spacer(1, 5))

        elementos.append(
            Paragraph(
                item.anotacao,
                styles['BodyText']
            )
        )

        elementos.append(Spacer(1, 10))

        elementos.append(
            HRFlowable(
                width="100%",
                thickness=0.5,
                color=colors.lightgrey
            )
        )

        elementos.append(Spacer(1, 10))
    # =========================================
    # ASSINATURA
    # =========================================

    elementos.append(Spacer(1, 15))

    elementos.append(
        Paragraph(
            '''
            <para align="center">
            _______________________________________
            <br/>
            Cirurgião-Dentista Responsável
            </para>
            ''',
            styles['Normal']
        )
    )

    # =========================================
    # RODAPÉ
    # =========================================

    elementos.append(Spacer(1, 20))

    elementos.append(
        HRFlowable(
            width="100%",
            thickness=0.8,
            color=colors.grey
        )
    )

    elementos.append(Spacer(1, 8))

    elementos.append(
        Paragraph(
            f'''
            <para align="center">
            <b>{nome_clinica}</b>
            </para>
            ''',
            styles['Normal']
        )
    )

    elementos.append(
        Paragraph(
            f'''
            <para align="center">
            Tel: {telefone_clinica}
            &nbsp;&nbsp;&nbsp;
            WhatsApp: {whatsapp_clinica}
            </para>
            ''',
            styles['BodyText']
        )
    )

    elementos.append(
        Paragraph(
            f'''
            <para align="center">
            {email_clinica}
            </para>
            ''',
            styles['BodyText']
        )
    )

    # =========================================
    # GERAR PDF
    # =========================================

    doc.build(
        elementos,
        onFirstPage=adicionar_rodape,
        onLaterPages=adicionar_rodape
    )

    return response
# =========================================

@login_required(login_url='/')
def novo_documento(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    if request.method == 'POST':

        template_id = request.POST.get(
            'template'
        )

        conteudo = request.POST.get(
            'conteudo',
            ''
        )

        titulo = request.POST.get(
            'titulo',
            ''
        )

        tipo_documento = request.POST.get(
            'tipo',
            'personalizado'
        )

        # =========================================
        # TEMPLATE SELECIONADO
        # =========================================

        if template_id:

            template = get_object_or_404(
                TemplateDocumento,
                id=template_id
            )

            # usa o tipo do template
            tipo_documento = template.tipo

            # usa o conteúdo do template
            conteudo = template.conteudo

            conteudo = conteudo.replace(
                '{{ paciente_nome }}',
                paciente.nome or ''
            )

            conteudo = conteudo.replace(
                '{{ paciente_cpf }}',
                paciente.cpf or ''
            )

            conteudo = conteudo.replace(
                '{{ data_atual }}',
                timezone.now().strftime(
                    '%d/%m/%Y'
                )
            )

            conteudo = conteudo.replace(
                '{{ data_atendimento }}',
                timezone.now().strftime(
                    '%d/%m/%Y'
                )
            )

            if paciente.nascimento:

                conteudo = conteudo.replace(
                    '{{ paciente_nascimento }}',
                    paciente.nascimento.strftime(
                        '%d/%m/%Y'
                    )
                )

            else:

                conteudo = conteudo.replace(
                    '{{ paciente_nascimento }}',
                    ''
                )

            config = ConfiguracaoClinica.objects.first()

            conteudo = conteudo.replace(
                '{{ cro_clinica }}',
                config.cro if config and config.cro else ''
            )

            # se não informar título,
            # usa o nome do template
            if not titulo:
                titulo = template.nome

        documento = DocumentoClinico.objects.create(

            paciente=paciente,

            titulo=titulo,

            tipo=tipo_documento,

            conteudo=conteudo,

            status='rascunho'

        )

        return redirect(
            'editar_documento',
            documento.id
        )

    templates = TemplateDocumento.objects.filter(
        ativo=True
    ).order_by('nome')

    return render(

        request,

        'accounts/documento_form.html',

        {

            'paciente': paciente,

            'templates': templates

        }

    )

# ========================================

@login_required(login_url='/')
def editar_documento(request, id):

    documento = get_object_or_404(

        DocumentoClinico,

        id=id

    )

    if request.method == 'POST':

        documento.titulo = request.POST.get(
            'titulo'
        )

        documento.conteudo = request.POST.get(
            'conteudo'
        )

        documento.save()

        return redirect(

            'editar_documento',

            documento.id

        )

    return render(

        request,

        'accounts/documento_form.html',

        {

            'documento': documento,

            'paciente': documento.paciente

        }

    )

# =========================================
# EXCLUIR DOCUMENTO
# =========================================

@login_required(login_url='/')
def excluir_documento(request, id):

    documento = get_object_or_404(
        DocumentoClinico,
        id=id
    )

    paciente_id = documento.paciente.id

    documento.delete()

    return redirect(
        'ficha_clinica',
        paciente_id
    ) 

@login_required(login_url='/')
def imprimir_documento(request, id):

    documento = get_object_or_404(
        DocumentoClinico,
        id=id
    )

    paciente = documento.paciente

    config = ConfiguracaoClinica.objects.first()

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = f'inline; filename="documento_{documento.id}.pdf"'

    doc = SimpleDocTemplate(
        response,
        topMargin=30,
        bottomMargin=30,
        leftMargin=40,
        rightMargin=40
    )

    styles = getSampleStyleSheet()

    elementos = []

   # =========================================
    # LOGO
    # =========================================

    logo_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'img',
        'logo.png'
    )

    if os.path.exists(logo_path):

        logo = Image(
            logo_path,
            width=240,
            height=90
        )

        logo.hAlign = 'CENTER'

        elementos.append(logo)

        elementos.append(
            Spacer(1, 8)
        )

    # =========================================
    # NOME DA CLÍNICA
    # =========================================

    if config and config.nome_clinica:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                <b>{config.nome_clinica}</b>
                </para>
                ''',
                styles['Heading2']
            )
        )

    # =========================================
    # TÍTULO
    # =========================================

    elementos.append(
        Paragraph(
            f'''
            <para align="center">
            <b>{documento.titulo.upper()}</b>
            </para>
            ''',
            styles['Title']
        )
    )

    elementos.append(Spacer(1, 10))

    elementos.append(
        HRFlowable(
            width="100%",
            thickness=1.2,
            color=colors.HexColor('#1e40af')
        )
    )

    elementos.append(Spacer(1, 15))

    # =========================================
    # PACIENTE
    # =========================================

    elementos.append(
        Paragraph(
            f'<b>Paciente:</b> {paciente.nome}',
            styles['Normal']
        )
    )

    if paciente.cpf:

        elementos.append(
            Paragraph(
                f'<b>CPF:</b> {paciente.cpf}',
                styles['Normal']
            )
        )

    elementos.append(Spacer(1, 15))

    # =========================================
    # CONTEÚDO
    # =========================================

    from reportlab.lib.styles import ParagraphStyle

    estilo_conteudo = ParagraphStyle(
        'ConteudoDocumento',
        parent=styles['BodyText'],
        fontSize=11,
        leading=20,
        alignment=4,  # Justificado
        spaceBefore=0,
        spaceAfter=12,
    )

    conteudo = documento.conteudo.replace(
        '\n',
        '<br/>'
    )

    elementos.append(
        Paragraph(
            conteudo,
            estilo_conteudo
        )
    )

    # =========================================
    # ASSINATURA
    # =========================================

    elementos.append(
        Spacer(1, 35)
    )

    elementos.append(
        Paragraph(
            '''
            <para align="center">
            ____________________________________
            <br/>
            Assinatura e Carimbo
            </para>
            ''',
            styles['Normal']
        )
    )
    # =========================================
    # DADOS DA CLÍNICA
    # =========================================

    nome_clinica = (
        config.nome_clinica
        if config and config.nome_clinica
        else ''
    )

    telefone_clinica = (
        config.telefone
        if config and config.telefone
        else ''
    )

    whatsapp_clinica = (
        config.whatsapp
        if config and config.whatsapp
        else ''
    )

    email_clinica = (
        config.email
        if config and config.email
        else ''
    )

    # =========================================
    # RODAPÉ
    # =========================================

    elementos.append(
        Spacer(1, 20)
    )

    elementos.append(
        HRFlowable(
            width="100%",
            thickness=0.8,
            color=colors.grey
        )
    )

    elementos.append(
        Spacer(1, 8)
    )

    if nome_clinica:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                <b>{nome_clinica}</b>
                </para>
                ''',
                styles['Normal']
            )
        )

    if telefone_clinica or whatsapp_clinica:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                Tel: {telefone_clinica}
                &nbsp;&nbsp;&nbsp;
                WhatsApp: {whatsapp_clinica}
                </para>
                ''',
                styles['BodyText']
            )
        )

    if email_clinica:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                {email_clinica}
                </para>
                ''',
                styles['BodyText']
            )
        )

    doc.build(elementos)

    return response

# =========================================
# VISUALIZAR DOCUMENTO
# =========================================

@login_required(login_url='/')
def visualizar_documento(request, id):

    documento = get_object_or_404(
        DocumentoClinico,
        id=id
    )

    config = ConfiguracaoClinica.objects.first()

    return render(

        request,

        'accounts/visualizar_documento.html',

        {

            'documento': documento,
            'config': config

        }

    ) 

# =========================================
# CONFIGURAÇÃO DA CLÍNICA
# =========================================

@login_required(login_url='/')
def configuracao_clinica(request):

    config, created = (
        ConfiguracaoClinica.objects.get_or_create(
            id=1
        )
    )

    if request.method == 'POST':

        config.nome_clinica = request.POST.get(
            'nome_clinica'
        )

        config.cro = request.POST.get(
            'cro'
        )

        config.cnpj = request.POST.get(
            'cnpj'
        )

        config.telefone = request.POST.get(
            'telefone'
        )

        config.whatsapp = request.POST.get(
            'whatsapp'
        )

        config.email = request.POST.get(
            'email'
        )

        config.endereco = request.POST.get(
            'endereco'
        )

        config.numero = request.POST.get(
            'numero'
        )

        config.bairro = request.POST.get(
            'bairro'
        )

        config.cidade = request.POST.get(
            'cidade'
        )

        config.estado = request.POST.get(
            'estado'
        )

        config.cep = request.POST.get(
            'cep'
        )

        config.observacoes_orcamento = (
            request.POST.get(
                'observacoes_orcamento'
            )
        )

        if request.FILES.get('logo'):

            config.logo = request.FILES.get(
                'logo'
            )

        config.save()

        return redirect(
            'configuracao_clinica'
        )

    return render(

        request,

        'accounts/configuracao_clinica.html',

        {
            'config': config
        }

    ) 

# =========================================
# PDF ANAMNESE
# =========================================

@login_required(login_url='/')
def imprimir_anamnese(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    anamnese = get_object_or_404(
        Anamnese,
        paciente=paciente
    )

    resumo_clinico = gerar_resumo_clinico(
        anamnese
    )

    print(resumo_clinico)

    config = ConfiguracaoClinica.objects.first()    

    context = {

        'paciente': paciente,
        'anamnese': anamnese,
        'config': config,
        'resumo_clinico': resumo_clinico
    }

    return render(

        request,

        'accounts/anamnese_pdf.html',

        context

    ) 

# =========================================
# RESUMO CLÍNICO AUTOMÁTICO
# =========================================

def gerar_resumo_clinico(anamnese):

    resumo = []

    # =========================================
    # CONDIÇÕES SISTÊMICAS
    # =========================================

    if anamnese.hipertenso:
        resumo.append('Hipertenso')

    if anamnese.diabetico:
        resumo.append('Diabético')

    if anamnese.cardiopatia:
        resumo.append('Cardiopata')

    if anamnese.asma:
        resumo.append('Asma')

    if anamnese.bronquite:
        resumo.append('Bronquite')

    if anamnese.anemia:
        resumo.append('Anemia')

    if anamnese.hepatite:
        resumo.append('Hepatite')

    if anamnese.rinite:
        resumo.append('Rinite')

    if anamnese.sinusite:
        resumo.append('Sinusite')

    if anamnese.problema_renal:
        resumo.append('Problema renal')

    if anamnese.sangramento_excessivo:
        resumo.append('Sangramento excessivo')

    if hasattr(anamnese, 'anticoagulante') and anamnese.anticoagulante:
        resumo.append('Uso de anticoagulantes')

    if hasattr(anamnese, 'bisfosfonato') and anamnese.bisfosfonato:
        resumo.append('Uso de bisfosfonatos')

    if hasattr(anamnese, 'marcapasso') and anamnese.marcapasso:
        resumo.append('Portador de marcapasso')

    if hasattr(anamnese, 'cancer') and anamnese.cancer:
        resumo.append('Histórico de câncer')

    if hasattr(anamnese, 'hipotireoidismo') and anamnese.hipotireoidismo:
        resumo.append('Hipotireoidismo')

    if hasattr(anamnese, 'hipertireoidismo') and anamnese.hipertireoidismo:
        resumo.append('Hipertireoidismo')

    # =========================================
    # ALERGIAS E MEDICAMENTOS
    # =========================================

    if anamnese.alergias:
        resumo.append(
            f'Alergias: {anamnese.alergias}'
        )

    if anamnese.medicamentos:
        resumo.append(
            f'Medicamentos: {anamnese.medicamentos}'
        )

    # =========================================
    # CIRURGIAS
    # =========================================

    if anamnese.cirurgia:

        if anamnese.cirurgias:

            resumo.append(
                f'Cirurgias prévias: {anamnese.cirurgias}'
            )

        else:

            resumo.append(
                'Histórico cirúrgico positivo'
            )

    if anamnese.hospitalizado:

        if anamnese.hospitalizacao:

            resumo.append(
                f'Hospitalização: {anamnese.hospitalizacao}'
            )

        else:

            resumo.append(
                'Hospitalização prévia'
            )

    if anamnese.transfusao_sangue:
        resumo.append(
            'Histórico de transfusão sanguínea'
        )

    # =========================================
    # HISTÓRIA ODONTOLÓGICA
    # =========================================

    if anamnese.bruxismo:
        resumo.append('Bruxismo')

    if anamnese.sensibilidade:
        resumo.append('Sensibilidade dentária')

    if anamnese.sangramento_gengival:
        resumo.append('Sangramento gengival')

    if anamnese.dor_mastigar:
        resumo.append('Dor à mastigação')

    if anamnese.medo_dentista:
        resumo.append('Ansiedade odontológica')

    if anamnese.abandono_tratamento:
        resumo.append('Histórico de abandono de tratamento')

    # =========================================
    # HÁBITOS
    # =========================================

    if anamnese.fumante:
        resumo.append('Fumante')

    if anamnese.ronco:
        resumo.append('Ronco')

    if anamnese.respiracao_bucal:
        resumo.append('Respiração bucal')

    # =========================================
    # RESULTADO
    # =========================================

    if not resumo:

        return (
            'Nenhuma condição clínica relevante informada.'
        )

    return ' • '.join(resumo)

# =========================================
# MEDICAMENTOS
# =========================================

@login_required(login_url='/')
def medicamentos(request):

    medicamentos = Medicamento.objects.all().order_by(
        'nome'
    )

    return render(

        request,

        'accounts/medicamentos.html',

        {
            'medicamentos': medicamentos
        }

    )


@login_required(login_url='/')
def novo_medicamento(request):

    if request.method == 'POST':

        Medicamento.objects.create(

            nome=request.POST.get(
                'nome'
            ),

            concentracao=request.POST.get(
                'concentracao'
            ),

            categoria=request.POST.get(
                'categoria'
            )

        )

        return redirect(
            'medicamentos'
        )

    return render(

        request,

        'accounts/medicamento_form.html'

    )

# =========================================
# RECEITAS
# =========================================

@login_required(login_url='/')
@perfil_required(
    "Administrador",
    "Dentista",
)
def receitas(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    receitas = Receita.objects.filter(
        paciente=paciente
    )

    return render(

        request,

        'accounts/receitas.html',

        {

            'paciente': paciente,
            'receitas': receitas

        }

    )


# =========================================
# NOVA RECEITA
# =========================================

@login_required(login_url='/')
def nova_receita(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    if request.method == 'POST':

        Receita.objects.create(

            paciente=paciente,

            medicamento=request.POST.get(
                'medicamento'
            ),

            quantidade=request.POST.get(
                'quantidade'
            ),

            posologia=request.POST.get(
                'posologia'
            ),

            observacoes=request.POST.get(
                'observacoes'
            ),

            tipo_receita=request.POST.get(
                'tipo_receita',
                'simples'
            )

        )

        return redirect(
            'receitas',
            id=paciente.id
        )

    return render(

        request,

        'accounts/receita_form.html',

        {

            'paciente': paciente,

            'modelos_receita': ModeloReceita.objects.filter(
                ativo=True
            ).order_by('nome')

        }

    )


# =========================================
# EDITAR RECEITA
# =========================================

@login_required(login_url='/')
def editar_receita(request, id):

    receita = get_object_or_404(
        Receita,
        id=id
    )

    if request.method == 'POST':

        receita.medicamento = request.POST.get(
            'medicamento'
        )

        receita.quantidade = request.POST.get(
            'quantidade'
        )

        receita.posologia = request.POST.get(
            'posologia'
        )

        receita.observacoes = request.POST.get(
            'observacoes'
        )

        receita.tipo_receita = request.POST.get(
            'tipo_receita',
            'simples'
        )

        receita.save()

        return redirect(
            'receitas',
            id=receita.paciente.id
        )

    return render(

        request,

        'accounts/receita_form.html',

        {

            'receita': receita,

            'paciente': receita.paciente

        }

    )


# =========================================
# EXCLUIR RECEITA
# =========================================

@login_required(login_url='/')
def excluir_receita(request, id):

    receita = get_object_or_404(
        Receita,
        id=id
    )

    paciente_id = receita.paciente.id

    receita.delete()

    return redirect(
        'receitas',
        id=paciente_id
    )

# =========================================
# PDF RECEITA
# =========================================

@login_required(login_url='/')
def imprimir_receita(request, id):

    receita = get_object_or_404(
        Receita,
        id=id
    )

    paciente = receita.paciente

    config = ConfiguracaoClinica.objects.first()

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = f'inline; filename="receita_{receita.id}.pdf"'

    doc = SimpleDocTemplate(
        response,
        topMargin=30,
        bottomMargin=30,
        leftMargin=40,
        rightMargin=40
    )

    styles = getSampleStyleSheet()

    elementos = []

    # =========================================
    # LOGO
    # =========================================

    logo_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'img',
        'logo.png'
    )

    if os.path.exists(logo_path):

        logo = Image(
            logo_path,
            width=240,
            height=90
        )

        logo.hAlign = 'CENTER'

        elementos.append(logo)

        elementos.append(
            Spacer(1, 8)
        )



    # =========================================
    # TÍTULO
    # =========================================

    titulo = (
        'RECEITA DE CONTROLE ESPECIAL'
        if receita.tipo_receita == 'controle'
        else 'RECEITUÁRIO'
    )

    elementos.append(

        Paragraph(

            f'''
            <para align="center">
            <b>{titulo}</b>
            </para>
            ''',

            styles['Title']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    elementos.append(

        HRFlowable(
            width="100%",
            thickness=1.2,
            color=colors.HexColor('#1e40af')
        )

    )

    elementos.append(
        Spacer(1, 15)
    )

    # =========================================
    # PACIENTE
    # =========================================

    elementos.append(

        Paragraph(

            f'<b>Paciente:</b> {paciente.nome}',

            styles['Normal']

        )

    )

    if paciente.cpf:

        elementos.append(

            Paragraph(

                f'<b>CPF:</b> {paciente.cpf}',

                styles['Normal']

            )

        )

    elementos.append(
        Spacer(1, 15)
    )

    # =========================================
    # MEDICAMENTO
    # =========================================

    elementos.append(

        Paragraph(

            '<b>Medicamento:</b>',

            styles['Heading3']

        )

    )

    elementos.append(

        Paragraph(

            str(receita.medicamento),

            styles['BodyText']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    # =========================================
    # QUANTIDADE
    # =========================================

    elementos.append(

        Paragraph(

            '<b>Quantidade:</b>',

            styles['Heading3']

        )

    )

    elementos.append(

        Paragraph(

            receita.quantidade,

            styles['BodyText']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    # =========================================
    # POSOLOGIA
    # =========================================

    elementos.append(

        Paragraph(

            '<b>Posologia:</b>',

            styles['Heading3']

        )

    )

    elementos.append(

        Paragraph(

            receita.posologia,

            styles['BodyText']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    # =========================================
    # OBSERVAÇÕES
    # =========================================

    if receita.observacoes:

        elementos.append(

            Paragraph(

                '<b>Observações:</b>',

                styles['Heading3']

            )

        )

        elementos.append(

            Paragraph(

                receita.observacoes,

                styles['BodyText']

            )

        )

        elementos.append(
            Spacer(1, 10)
        )

    

    # =========================================
    # ASSINATURA
    # =========================================

    elementos.append(
        Spacer(1, 25)
    )

    elementos.append(
        Paragraph(
            '''
            <para align="center">
            _______________________________________
            <br/>
            Cirurgião-Dentista Responsável
            </para>
            ''',
            styles['Normal']
        )
    )

    # =========================================
    # RODAPÉ
    # =========================================

    elementos.append(
        Spacer(1, 20)
    )

    elementos.append(
        HRFlowable(
            width="100%",
            thickness=0.8,
            color=colors.grey
        )
    )

    elementos.append(
        Spacer(1, 8)
    )

    if config and config.nome_clinica:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                <b>{config.nome_clinica}</b>
                </para>
                ''',
                styles['Normal']
            )
        )

    if config and config.telefone:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                Tel: {config.telefone}
                &nbsp;&nbsp;&nbsp;
                WhatsApp: {config.whatsapp}
                </para>
                ''',
                styles['BodyText']
            )
        )

    if config and config.endereco:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                {config.endereco}
                </para>
                ''',
                styles['BodyText']
            )
        )

    if config and config.email:

        elementos.append(
            Paragraph(
                f'''
                <para align="center">
                {config.email}
                </para>
                ''',
                styles['BodyText']
            )
        )

    doc.build(elementos)

    return response

# =========================================
# MODELO DE RECEITA AJAX
# =========================================

@login_required(login_url='/')
def buscar_modelo_receita(request, id):

    modelo = get_object_or_404(
        ModeloReceita,
        id=id
    )

    return JsonResponse({

        'medicamento': modelo.medicamento,

        'quantidade': modelo.quantidade,

        'posologia': modelo.posologia,

    })

# =========================================
# MODELOS DE RECEITA
# =========================================

@login_required(login_url='/')
def modelos_receita(request):

    modelos = ModeloReceita.objects.all().order_by(
        'nome'
    )

    if request.method == 'POST':

        ModeloReceita.objects.create(

            nome=request.POST.get(
                'nome'
            ),

            medicamento=request.POST.get(
                'medicamento'
            ),

            quantidade=request.POST.get(
                'quantidade'
            ),

            posologia=request.POST.get(
                'posologia'
            ),

            ativo='ativo' in request.POST

        )

        return redirect(
            'modelos_receita'
        )

    return render(

        request,

        'accounts/modelos_receita.html',

        {

            'modelos': modelos

        }

    )

# =========================================
# EXCLUIR MODELO RECEITA
# =========================================

@login_required(login_url='/')
def excluir_modelo_receita(request, id):

    modelo = get_object_or_404(
        ModeloReceita,
        id=id
    )

    modelo.delete()

    return redirect(
        'modelos_receita'
    )

# =========================================
# EDITAR MODELO RECEITA
# =========================================

@login_required(login_url='/')
def editar_modelo_receita(request, id):

    modelo = get_object_or_404(
        ModeloReceita,
        id=id
    )

    if request.method == 'POST':

        modelo.nome = request.POST.get(
            'nome'
        )

        modelo.medicamento = request.POST.get(
            'medicamento'
        )

        modelo.quantidade = request.POST.get(
            'quantidade'
        )

        modelo.posologia = request.POST.get(
            'posologia'
        )

        modelo.ativo = (
            'ativo' in request.POST
        )

        modelo.save()

        return redirect(
            'modelos_receita'
        )

    return render(

        request,

        'accounts/modelo_receita_form.html',

        {

            'modelo': modelo

        }

    )

# =========================================
# EXAMES
# =========================================

@login_required(login_url='/')
def exames(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    exames = Exame.objects.filter(
        paciente=paciente
    ).order_by(
        '-criado_em'
    )

    return render(

        request,

        'accounts/exames.html',

        {

            'paciente': paciente,
            'exames': exames

        }

    )


# =========================================
# NOVO EXAME
# =========================================

@login_required(login_url='/')
def novo_exame(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    if request.method == 'POST':

        Exame.objects.create(

            paciente=paciente,

            nome=request.POST.get(
                'nome'
            ),

            data_exame=request.POST.get(
                'data_exame'
            ),

            observacoes=request.POST.get(
                'observacoes'
            ),

            arquivo=request.FILES.get(
                'arquivo'
            )

        )

        return redirect(
            'exames',
            paciente.id
        )

    return render(

        request,

        'accounts/novo_exame.html',

        {

            'paciente': paciente

        }

    )


# =========================================
# EXCLUIR EXAME
# =========================================

@login_required(login_url='/')
def excluir_exame(request, id):

    exame = get_object_or_404(
        Exame,
        id=id
    )

    paciente_id = exame.paciente.id

    if exame.arquivo:

        exame.arquivo.delete(
            save=False
        )

    exame.delete()

    return redirect(
        'exames',
        paciente_id
    )

# =========================================
# SOLICITAÇÕES DE EXAMES
# =========================================

@login_required(login_url='/')
@perfil_required(
    "Administrador",
    "Dentista",
)
def solicitacoes_exames(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    solicitacoes = SolicitacaoExame.objects.filter(
        paciente=paciente
    )

    return render(

        request,

        'accounts/solicitacoes_exames.html',

        {

            'paciente': paciente,
            'solicitacoes': solicitacoes

        }

    )


# =========================================
# NOVA SOLICITAÇÃO DE EXAME
# =========================================

@login_required(login_url='/')
@perfil_required(
    "Administrador",
    "Dentista",
)
def nova_solicitacao_exame(request, id):

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    if request.method == 'POST':

        SolicitacaoExame.objects.create(

            paciente=paciente,

            titulo=request.POST.get(
                'titulo'
            ),

            exames_solicitados=request.POST.get(
                'exames_solicitados'
            ),

            observacoes=request.POST.get(
                'observacoes'
            )

        )

        return redirect(
            'solicitacoes_exames',
            paciente.id
        )

    return render(

        request,

        'accounts/nova_solicitacao_exame.html',

        {

            'paciente': paciente

        }

    )


# =========================================
# EXCLUIR SOLICITAÇÃO
# =========================================

@login_required(login_url='/')
def excluir_solicitacao_exame(request, id):

    solicitacao = get_object_or_404(
        SolicitacaoExame,
        id=id
    )

    paciente_id = solicitacao.paciente.id

    solicitacao.delete()

    return redirect(
        'solicitacoes_exames',
        paciente_id
    )


# =========================================
# PDF SOLICITAÇÃO DE EXAME
# =========================================

@login_required(login_url='/')
def imprimir_solicitacao_exame(request, id):

    solicitacao = get_object_or_404(
        SolicitacaoExame,
        id=id
    )

    paciente = solicitacao.paciente

    config = ConfiguracaoClinica.objects.first()

    response = HttpResponse(
        content_type='application/pdf'
    )

    response[
        'Content-Disposition'
    ] = (
        f'inline; '
        f'filename="solicitacao_exame_{solicitacao.id}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        topMargin=30,
        bottomMargin=30,
        leftMargin=40,
        rightMargin=40
    )

    styles = getSampleStyleSheet()

    elementos = []

    # =========================================
    # LOGO
    # =========================================

    logo_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'img',
        'logo.png'
    )

    if os.path.exists(logo_path):

        logo = Image(
            logo_path,
            width=240,
            height=90
        )

        logo.hAlign = 'CENTER'

        elementos.append(logo)

        elementos.append(
            Spacer(1, 8)
        )

    # =========================================
    # TÍTULO
    # =========================================

    elementos.append(

        Paragraph(

            '''
            <para align="center">
            <b>SOLICITAÇÃO DE EXAMES</b>
            </para>
            ''',

            styles['Title']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    elementos.append(

        HRFlowable(
            width="100%",
            thickness=1.2,
            color=colors.HexColor('#1e40af')
        )

    )

    elementos.append(
        Spacer(1, 15)
    )

    # =========================================
    # PACIENTE
    # =========================================

    elementos.append(

        Paragraph(

            f'<b>Paciente:</b> {paciente.nome}',

            styles['Normal']

        )

    )

    if paciente.cpf:

        elementos.append(

            Paragraph(

                f'<b>CPF:</b> {paciente.cpf}',

                styles['Normal']

            )

        )

    elementos.append(
        Spacer(1, 15)
    )

    # =========================================
    # TÍTULO DA SOLICITAÇÃO
    # =========================================

    elementos.append(

        Paragraph(

            '<b>Título:</b>',

            styles['Heading3']

        )

    )

    elementos.append(

        Paragraph(

            solicitacao.titulo,

            styles['BodyText']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    # =========================================
    # EXAMES SOLICITADOS
    # =========================================

    elementos.append(

        Paragraph(

            '<b>Exames Solicitados:</b>',

            styles['Heading3']

        )

    )

    elementos.append(

        Paragraph(

            solicitacao.exames_solicitados.replace(
                '\n',
                '<br/>'
            ),

            styles['BodyText']

        )

    )

    elementos.append(
        Spacer(1, 10)
    )

    # =========================================
    # OBSERVAÇÕES
    # =========================================

    if solicitacao.observacoes:

        elementos.append(

            Paragraph(

                '<b>Observações:</b>',

                styles['Heading3']

            )

        )

        elementos.append(

            Paragraph(

                solicitacao.observacoes.replace(
                    '\n',
                    '<br/>'
                ),

                styles['BodyText']

            )

        )

        elementos.append(
            Spacer(1, 10)
        )

    # =========================================
    # DATA
    # =========================================

    elementos.append(

        Paragraph(

            (
                f'Franca/SP, '
                f'{solicitacao.criado_em.strftime("%d/%m/%Y")}'
            ),

            styles['Normal']

        )

    )

    # =========================================
    # ASSINATURA
    # =========================================

    elementos.append(
        Spacer(1, 35)
    )

    elementos.append(

        Paragraph(

            '''
            <para align="center">
            _______________________________________
            <br/>
            Cirurgião-Dentista Responsável
            </para>
            ''',

            styles['Normal']

        )

    )

    # =========================================
    # RODAPÉ
    # =========================================

    elementos.append(
        Spacer(1, 20)
    )

    elementos.append(

        HRFlowable(
            width="100%",
            thickness=0.8,
            color=colors.grey
        )

    )

    elementos.append(
        Spacer(1, 8)
    )

    if config and config.nome_clinica:

        elementos.append(

            Paragraph(

                f'''
                <para align="center">
                <b>{config.nome_clinica}</b>
                </para>
                ''',

                styles['Normal']

            )

        )

    if config and config.telefone:

        elementos.append(

            Paragraph(

                f'''
                <para align="center">
                Tel: {config.telefone}
                &nbsp;&nbsp;&nbsp;
                WhatsApp: {config.whatsapp}
                </para>
                ''',

                styles['BodyText']

            )

        )

    if config and config.endereco:

        elementos.append(

            Paragraph(

                f'''
                <para align="center">
                {config.endereco}
                </para>
                ''',

                styles['BodyText']

            )

        )

    if config and config.email:

        elementos.append(

            Paragraph(

                f'''
                <para align="center">
                {config.email}
                </para>
                ''',

                styles['BodyText']

            )

        )

    doc.build(elementos)

    return response

# =========================================
# PERFIS
# =========================================

def usuario_tem_perfil(user, *tipos):
    """
    Verifica se o usuário pertence a um ou mais perfis.
    Exemplo:
        usuario_tem_perfil(request.user, PerfilUsuario.ADMIN)
        usuario_tem_perfil(request.user,
                           PerfilUsuario.ADMIN,
                           PerfilUsuario.GESTOR)
    """

    if not hasattr(user, "perfil"):
        return False

    return user.perfil.tipo_usuario in tipos

# =========================================
# USUÁRIOS
# =========================================

@login_required
@permissao_required("usuarios")
def usuarios(request):

    usuarios = User.objects.select_related(
        "perfil"
    ).order_by(
        "first_name",
        "username"
    )

    context = {
        "usuarios": usuarios
    }

    return render(
        request,
        "accounts/usuarios.html",
        context
    )

# =========================================
# PERFIL DO USUÁRIO
# =========================================

@login_required
@perfil_required(
    "Administrador",
    "Gestor",
)
def perfil_usuario(request, id):

    usuario = get_object_or_404(
        User,
        id=id
    )

    perfil = usuario.perfil

    context = {

        "usuario": usuario,

        "perfil": perfil,

    }

    return render(
        request,
        "accounts/perfil_usuario.html",
        context
    )

# =========================================
# NOVO USUÁRIO
# =========================================

@login_required
def novo_usuario(request):

    if request.method == "POST":

        # =========================================
        # DADOS DO USUÁRIO
        # =========================================

        nome = request.POST.get("nome")
        username = request.POST.get("username")
        email = request.POST.get("email")
        senha = request.POST.get("senha")

        if User.objects.filter(username=username).exists():

            messages.error(

                request,

                "Já existe um usuário com esse login."

            )

            return redirect("novo_usuario")

        usuario = User.objects.create_user(

            username=username,

            email=email,

            password=senha,

            first_name=nome,

        )

        # =========================================
        # PERFIL DO USUÁRIO
        # =========================================

        perfil = PerfilUsuario.objects.create(

            usuario=usuario,

            perfil_acesso_id=request.POST.get("perfil_acesso") or None,

            tipo_usuario=request.POST.get("tipo_usuario"),

            foto=request.FILES.get("foto"),

            cpf=request.POST.get("cpf"),
            rg=request.POST.get("rg"),
            data_nascimento=request.POST.get("data_nascimento") or None,
            sexo=request.POST.get("sexo"),

            telefone=request.POST.get("telefone"),
            celular=request.POST.get("celular"),

            cep=request.POST.get("cep"),
            logradouro=request.POST.get("logradouro"),
            numero=request.POST.get("numero"),
            complemento=request.POST.get("complemento"),
            bairro=request.POST.get("bairro"),
            cidade=request.POST.get("cidade"),
            uf=request.POST.get("uf"),

            cro=request.POST.get("cro"),
            cro_uf=request.POST.get("cro_uf"),
            especialidade=request.POST.get("especialidade"),

            assinatura=request.FILES.get("assinatura"),

        )

        # =========================================
        # CRIAR PROFISSIONAL AUTOMATICAMENTE
        # =========================================

        if perfil.tipo_usuario in [

            PerfilUsuario.DENTISTA,

            PerfilUsuario.ACD,

        ]:

            Profissional.objects.get_or_create(

                usuario=usuario,

                defaults={

                    "nome": nome,

                    "email": email,

                    "telefone": request.POST.get("celular"),

                    "especialidade": request.POST.get("especialidade"),

                    "ativo": True,

                }

            )

        # =========================================
        # MENSAGEM
        # =========================================

        messages.success(

            request,

            "Usuário criado com sucesso."

        )

        return redirect("usuarios")

    # =========================================
    # PERFIS DE ACESSO
    # =========================================

    perfis = Perfil.objects.filter(

        ativo=True

    ).order_by(

        "nome"

    )

    return render(

        request,

        "accounts/usuario_form.html",

        {

            "perfis": perfis,

        }

    )

# =========================================
# EDITAR USUÁRIO
# =========================================

@login_required
def editar_usuario(request, id):

    usuario = get_object_or_404(User, id=id)

    perfil = usuario.perfil

    if request.method == "POST":

        # =========================================
        # DADOS DO USUÁRIO
        # =========================================

        usuario.first_name = request.POST.get("nome")

        usuario.username = request.POST.get("username")

        usuario.email = request.POST.get("email")

        usuario.save()

        # =========================================
        # DADOS DO PERFIL
        # =========================================

        perfil.perfil_acesso_id = request.POST.get(
            "perfil_acesso"
        ) or None

        # =========================================
        # COMPATIBILIDADE COM O CAMPO ANTIGO
        # =========================================

        if perfil.perfil_acesso:

            mapa = {

                "Administrador": PerfilUsuario.ADMIN,
                "Gestor": PerfilUsuario.GESTOR,
                "Dentista": PerfilUsuario.DENTISTA,
                "Secretária": PerfilUsuario.SECRETARIA,
                "Auxiliar de Saúde Bucal": PerfilUsuario.ACD,
                "Contabilidade": PerfilUsuario.CONTABILIDADE,
                "Marketing": PerfilUsuario.MARKETING,
                "Auditoria": PerfilUsuario.AUDITORIA,

            }

            perfil.tipo_usuario = mapa.get(

                perfil.perfil_acesso.nome,

                perfil.tipo_usuario

            )

        # =========================================
        # DADOS COMPLEMENTARES
        # =========================================

        perfil.cro = request.POST.get("cro")

        perfil.cro_uf = request.POST.get("cro_uf")

        perfil.especialidade = request.POST.get(
            "especialidade"
        )

        perfil.telefone = request.POST.get(
            "telefone"
        )

        perfil.celular = request.POST.get(
            "celular"
        )

        perfil.cpf = request.POST.get(
            "cpf"
        )

        perfil.rg = request.POST.get(
            "rg"
        )

        perfil.data_nascimento = request.POST.get(
            "data_nascimento"
        ) or None

        perfil.sexo = request.POST.get(
            "sexo"
        )

        perfil.cep = request.POST.get(
            "cep"
        )

        perfil.logradouro = request.POST.get(
            "logradouro"
        )

        perfil.numero = request.POST.get(
            "numero"
        )

        perfil.complemento = request.POST.get(
            "complemento"
        )

        perfil.bairro = request.POST.get(
            "bairro"
        )

        perfil.cidade = request.POST.get(
            "cidade"
        )

        perfil.uf = request.POST.get(
            "uf"
        )

        if request.FILES.get("foto"):

            perfil.foto = request.FILES["foto"]

        if request.FILES.get("assinatura"):

            perfil.assinatura = request.FILES["assinatura"]

        perfil.save()

        messages.success(

            request,

            "Usuário atualizado com sucesso."

        )

        return redirect("usuarios")

    # =========================================
    # PERFIS DE ACESSO
    # =========================================

    perfis = Perfil.objects.filter(

        ativo=True

    ).order_by(

        "nome"

    )

    context = {

        "usuario": usuario,

        "perfil": perfil,

        "perfis": perfis,

    }

    return render(

        request,

        "accounts/usuario_form.html",

        context

    )
# ==============================================================

@login_required
def alterar_status_usuario(request, id):

    usuario = get_object_or_404(User, id=id)

    if usuario.id == request.user.id:

        messages.warning(
            request,
            'Você não pode desativar seu próprio usuário.'
        )

        return redirect('usuarios')

    perfil = usuario.perfil

    perfil.ativo = not perfil.ativo

    perfil.save()

    # =========================================
    # SINCRONIZAR PROFISSIONAL
    # =========================================

    try:

        profissional = Profissional.objects.get(
            usuario=usuario
        )

        profissional.ativo = perfil.ativo

        profissional.save()

    except Profissional.DoesNotExist:

        pass

    messages.success(
        request,
        'Status atualizado com sucesso.'
    )

    return redirect('usuarios')


    # =========================================
# CRIAR PERFIS PADRÃO
# =========================================

@login_required
def criar_perfis(request):

    perfis = [

        ("Administrador", "Acesso total ao sistema"),
        ("Gestor", "Gestão da clínica"),
        ("Dentista", "Atendimento odontológico"),
        ("Secretária", "Atendimento e agenda"),
        ("Auxiliar de Saúde Bucal", "Apoio clínico"),
        ("Contabilidade", "Financeiro"),
        ("Marketing", "Marketing"),
        ("Auditoria", "Auditoria e consultas"),

    ]

    for nome, descricao in perfis:

        Perfil.objects.get_or_create(

            nome=nome,

            defaults={

                "descricao": descricao,
                "ativo": True,

            }

        )

    messages.success(

        request,

        "Perfis criados com sucesso."

    )

    return redirect("perfis")

# =========================================
# CRIAR MÓDULOS DO SISTEMA
# =========================================

@login_required
def criar_modulos(request):

    modulos = [

        ("Dashboard", "Painel principal", "bi-speedometer2", 1),
        ("Agenda", "Agenda da clínica", "bi-calendar3", 2),
        ("Pacientes", "Cadastro de pacientes", "bi-people", 3),
        ("Anamnese", "Anamnese", "bi-clipboard2-pulse", 4),
        ("Odontograma", "Odontograma", "bi-grid-3x3", 5),
        ("Orçamentos", "Orçamentos", "bi-file-earmark-text", 6),
        ("Financeiro", "Financeiro", "bi-cash-stack", 7),
        ("Compras", "Compras", "bi-cart3", 8),
        ("Estoque", "Controle de estoque", "bi-box-seam", 9),
        ("Relatórios", "Relatórios", "bi-bar-chart", 10),
        ("Auditoria", "Auditoria", "bi-search", 11),
        ("Usuários", "Usuários", "bi-person-gear", 12),
        ("Configurações", "Configurações", "bi-gear", 13),

    ]

    for nome, descricao, icone, ordem in modulos:

        Modulo.objects.get_or_create(

            nome=nome,

            defaults={

                "descricao": descricao,

                "icone": icone,

                "ordem": ordem,

                "ativo": True,

            }

        )

    messages.success(

        request,

        "Módulos criados com sucesso."

    )

    return redirect("perfis")

# =========================================
# MEU PERFIL
# =========================================

@login_required
def meu_perfil(request):

    usuario = request.user
    perfil = usuario.perfil

    if request.method == 'POST':

        # =====================================
        # USUÁRIO
        # =====================================

        usuario.first_name = request.POST.get('nome')
        usuario.email = request.POST.get('email')

        usuario.save()

        # =====================================
        # DADOS PESSOAIS
        # =====================================

        perfil.telefone = request.POST.get('telefone')
        perfil.celular = request.POST.get('celular')

        perfil.cpf = request.POST.get('cpf')
        perfil.rg = request.POST.get('rg')

        perfil.sexo = request.POST.get('sexo')

        data_nascimento = request.POST.get('data_nascimento')

        if data_nascimento:
            perfil.data_nascimento = data_nascimento

        # =====================================
        # ENDEREÇO
        # =====================================

        perfil.cep = request.POST.get('cep')
        perfil.logradouro = request.POST.get('logradouro')
        perfil.numero = request.POST.get('numero')
        perfil.complemento = request.POST.get('complemento')

        perfil.bairro = request.POST.get('bairro')
        perfil.cidade = request.POST.get('cidade')
        perfil.uf = request.POST.get('uf')

        # =====================================
        # DADOS PROFISSIONAIS
        # =====================================

        perfil.cro = request.POST.get('cro')
        perfil.cro_uf = request.POST.get('cro_uf')
        perfil.especialidade = request.POST.get('especialidade')

        # =====================================
        # FOTO
        # =====================================

        if request.FILES.get('foto'):

            perfil.foto = request.FILES.get('foto')

        # =====================================
        # ASSINATURA
        # =====================================

        if request.FILES.get('assinatura'):

            perfil.assinatura = request.FILES.get('assinatura')

        perfil.save()

        # =========================================
        # SINCRONIZAR PROFISSIONAL
        # =========================================

        if perfil.tipo_usuario in [

            PerfilUsuario.DENTISTA,
            PerfilUsuario.ACD,

        ]:

            profissional, criado = Profissional.objects.get_or_create(

                usuario=usuario,

                defaults={

                    'nome': usuario.first_name,
                    'email': usuario.email,
                    'telefone': perfil.celular or perfil.telefone,
                    'especialidade': perfil.especialidade,
                    'ativo': perfil.ativo,

                }

            )

            profissional.nome = usuario.first_name
            profissional.email = usuario.email
            profissional.telefone = (
                perfil.celular or perfil.telefone
            )
            profissional.especialidade = (
                perfil.especialidade
            )
            profissional.ativo = perfil.ativo

            profissional.save()

            messages.success(
                request,
                'Perfil atualizado com sucesso.'
            )

            return redirect('meu_perfil')

    context = {

        'usuario': usuario,
        'perfil': perfil

    }

    return render(

        request,

        'accounts/meu_perfil.html',

        context

    )

# =========================================
# ALTERAR SENHA
# =========================================

@login_required
def alterar_senha(request):

    if request.method == 'POST':

        senha_atual = request.POST.get('senha_atual')
        nova_senha = request.POST.get('nova_senha')
        confirmar_senha = request.POST.get('confirmar_senha')

        if not request.user.check_password(senha_atual):

            messages.error(
                request,
                'Senha atual incorreta.'
            )

            return redirect('alterar_senha')

        if nova_senha != confirmar_senha:

            messages.error(
                request,
                'As senhas não conferem.'
            )

            return redirect('alterar_senha')

        request.user.set_password(nova_senha)
        request.user.save()

        update_session_auth_hash(
            request,
            request.user
        )

        messages.success(
            request,
            'Senha alterada com sucesso.'
        )

        return redirect('meu_perfil')

    return render(
        request,
        'accounts/alterar_senha.html'
    )

# =========================================
# ATIVAR USUÁRIO
# =========================================

@login_required
def ativar_usuario(request, id):

    usuario = get_object_or_404(User, id=id)

    usuario.perfil.ativo = True

    usuario.perfil.save()

    return redirect('usuarios')


# =========================================
# DESATIVAR USUÁRIO
# =========================================

@login_required
def desativar_usuario(request, id):

    usuario = get_object_or_404(User, id=id)

    usuario.perfil.ativo = False

    usuario.perfil.save()

    return redirect('usuarios')


# =========================================
# PERFIS
# =========================================

@login_required
def perfis(request):

    perfis = (

    Perfil.objects

    .annotate(

        total_usuarios=Count("usuarios")

    )

    .order_by(

        "tipo",

        "nome"

    )

)

    context = {

        "perfis": perfis,

        "total_perfis": perfis.count(),

        "perfis_ativos": perfis.filter(
            ativo=True
        ).count(),

        "perfis_inativos": perfis.filter(
            ativo=False
        ).count(),

        "total_usuarios": User.objects.count(),

    }

    return render(

        request,

        "accounts/perfis.html",

        context

    )

# =========================================
# NOVO PERFIL
# =========================================

@login_required
def novo_perfil(request):

    if request.method == "POST":

        form = PerfilForm(request.POST)

        if form.is_valid():

            perfil = form.save()

            # Aplica automaticamente as permissões padrão
            aplicar_permissoes_padrao(perfil)

            messages.success(
                request,
                "Perfil cadastrado com sucesso."
            )

            return redirect("perfis")

    else:

        form = PerfilForm()

    return render(

        request,

        "accounts/perfil_form.html",

        {

            "form": form,

            "titulo": "Novo Perfil"

        }

    )

# =========================================
# EDITAR PERFIL
# =========================================

@login_required
def editar_perfil(request, perfil_id):

    perfil = get_object_or_404(

        Perfil,

        id=perfil_id

    )

    if request.method == "POST":

        form = PerfilForm(

            request.POST,

            instance=perfil

        )

        if form.is_valid():

            form.save()

            messages.success(

                request,

                "Perfil atualizado com sucesso."

            )

            return redirect("perfis")

    else:

        form = PerfilForm(

            instance=perfil

        )

    return render(

        request,

        "accounts/perfil_form.html",

        {

            "form": form,
            "perfil": perfil,
            "titulo": "Editar Perfil",

        }

    )

# =========================================
# DETALHES DO PERFIL
# =========================================

@login_required
def perfil_detalhes(request, perfil_id):

    perfil = get_object_or_404(

        Perfil,

        id=perfil_id

    )

    context = {

        "perfil": perfil,

    }

    return render(

        request,

        "accounts/perfil_detalhes.html",

        context

    )

# =========================================
# PERMISSÕES DO PERFIL
# =========================================

@login_required
def permissoes_perfil(request, perfil_id):

    perfil = get_object_or_404(
        Perfil,
        id=perfil_id
    )

    modulos = (
        Modulo.objects
        .filter(ativo=True)
        .order_by(
            "grupo",
            "ordem",
            "nome"
        )
    )

        # =====================================
    # SALVAR PERMISSÕES
    # =====================================

    if request.method == "POST":

        for modulo in modulos:

            permissao, created = Permissao.objects.get_or_create(
                perfil=perfil,
                modulo=modulo
            )

            permissao.visualizar = (
                f"visualizar_{modulo.id}" in request.POST
            )

            permissao.inserir = (
                f"inserir_{modulo.id}" in request.POST
            )

            permissao.editar = (
                f"editar_{modulo.id}" in request.POST
            )

            permissao.excluir = (
                f"excluir_{modulo.id}" in request.POST
            )

            permissao.exportar = (
                f"exportar_{modulo.id}" in request.POST
            )

            permissao.aprovar = (
                f"aprovar_{modulo.id}" in request.POST
            )

            permissao.save()

        messages.success(
            request,
            "Permissões atualizadas com sucesso."
        )

        return redirect(
            "permissoes_perfil",
            perfil.id
        )

    # =====================================
    # CARREGAR PERMISSÕES
    # =====================================

    permissoes = {
        p.modulo_id: p
        for p in Permissao.objects.filter(perfil=perfil)
    }

    grupos = {}

    for modulo in modulos:

        modulo.permissao = permissoes.get(modulo.id)

        grupos.setdefault(
            modulo.grupo,
            []
        ).append(modulo)

    context = {
        "perfil": perfil,
        "grupos": grupos,
    }

    return render(
        request,
        "accounts/permissoes_perfil.html",
        context
    )


# =========================================
# APLICAR PERMISSÕES PADRÃO
# =========================================

@login_required
def aplicar_permissoes_padrao_view(request, perfil_id):

    perfil = get_object_or_404(
        Perfil,
        id=perfil_id
    )

    aplicar_permissoes_padrao(perfil)

    messages.success(
        request,
        f'Permissões padrão aplicadas ao perfil "{perfil.nome}".'
    )

    return redirect(
        "permissoes_perfil",
        perfil.id
    )

# =========================================
# FORNECEDORES
# =========================================

@login_required
@permissao_required("fornecedores", "visualizar")
def fornecedores(request):

    fornecedores = Fornecedor.objects.order_by(
        'nome'
    )

    context = {
        'fornecedores': fornecedores
    }

    return render(
        request,
        'accounts/fornecedores.html',
        context
    )


# =========================================
# VALIDAÇÃO DO CNPJ
# =========================================

import re

def validar_cnpj(cnpj):

    cnpj = re.sub(r'\D', '', cnpj)

    if len(cnpj) != 14:
        return False

    if cnpj == cnpj[0] * 14:
        return False

    peso1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    soma = sum(int(cnpj[i]) * peso1[i] for i in range(12))

    dig1 = 0 if soma % 11 < 2 else 11 - (soma % 11)

    if dig1 != int(cnpj[12]):
        return False

    peso2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]
    soma = sum(int(cnpj[i]) * peso2[i] for i in range(13))

    dig2 = 0 if soma % 11 < 2 else 11 - (soma % 11)

    return dig2 == int(cnpj[13])

# =========================================
# NOVO FORNECEDOR
# =========================================
@login_required
@permissao_required("fornecedores", "inserir")
def novo_fornecedor(request):
    if request.method == 'POST':

        nome = request.POST.get(
            'nome'
        )

        if not nome:

            messages.error(
                request,
                'Informe o nome do fornecedor.'
            )

            return render(
                request,
                'accounts/fornecedor_form.html'
            )

        cnpj = request.POST.get(
            'cnpj'
        )

        if cnpj:

            if not validar_cnpj(cnpj):

                messages.error(
                    request,
                    'CNPJ inválido.'
                )

                return render(
                    request,
                    'accounts/fornecedor_form.html'
                )

            if Fornecedor.objects.filter(
                cnpj=cnpj
            ).exists():

                messages.error(
                    request,
                    'Já existe um fornecedor com este CNPJ.'
                )

                return render(
                    request,
                    'accounts/fornecedor_form.html'
                )

        Fornecedor.objects.create(

            nome=nome,

            razao_social=request.POST.get(
                'razao_social'
            ),

            cnpj=cnpj,

            contato=request.POST.get(
                'contato'
            ),

            telefone=request.POST.get(
                'telefone'
            ),

            celular=request.POST.get(
                'celular'
            ),

            email=request.POST.get(
                'email'
            ),

            cep=request.POST.get(
                'cep'
            ),

            logradouro=request.POST.get(
                'logradouro'
            ),

            numero=request.POST.get(
                'numero'
            ),

            complemento=request.POST.get(
                'complemento'
            ),

            bairro=request.POST.get(
                'bairro'
            ),

            cidade=request.POST.get(
                'cidade'
            ),

            uf=request.POST.get(
                'uf'
            ),

            ativo=True

        )

        messages.success(
            request,
            'Fornecedor cadastrado com sucesso.'
        )

        return redirect(
            'fornecedores'
        )

    return render(
        request,
        'accounts/fornecedor_form.html'
    )

# =========================================
# EDITAR FORNECEDOR
# =========================================

@login_required
@permissao_required("fornecedores", "editar")
def editar_fornecedor(request):

    fornecedor = get_object_or_404(
        Fornecedor,
        id=fornecedor_id
    )

    if request.method == 'POST':

        cnpj = request.POST.get(
            'cnpj'
        )

        if cnpj:

            cnpj_limpo = re.sub(
                r'\D',
                '',
                cnpj
            )

            if not validar_cnpj(cnpj_limpo):

                messages.error(
                    request,
                    'CNPJ inválido.'
                )

                return render(
                    request,
                    'accounts/fornecedor_form.html',
                    {
                        'fornecedor': fornecedor
                    }
                )

            if Fornecedor.objects.filter(
                cnpj=cnpj
            ).exclude(
                id=fornecedor.id
            ).exists():

                messages.error(
                    request,
                    'Já existe outro fornecedor com este CNPJ.'
                )

                return render(
                    request,
                    'accounts/fornecedor_form.html',
                    {
                        'fornecedor': fornecedor
                    }
                )

        fornecedor.nome = request.POST.get(
            'nome'
        )

        fornecedor.razao_social = request.POST.get(
            'razao_social'
        )

        fornecedor.cnpj = cnpj

        fornecedor.nome = request.POST.get(
            'nome'
        )

        fornecedor.razao_social = request.POST.get(
            'razao_social'
        )

        fornecedor.cnpj = request.POST.get(
            'cnpj'
        )

        fornecedor.contato = request.POST.get(
            'contato'
        )

        fornecedor.telefone = request.POST.get(
            'telefone'
        )

        fornecedor.celular = request.POST.get(
            'celular'
        )

        fornecedor.email = request.POST.get(
            'email'
        )

        fornecedor.cep = request.POST.get(
            'cep'
        )

        fornecedor.logradouro = request.POST.get(
            'logradouro'
        )

        fornecedor.numero = request.POST.get(
            'numero'
        )

        fornecedor.complemento = request.POST.get(
            'complemento'
        )

        fornecedor.bairro = request.POST.get(
            'bairro'
        )

        fornecedor.cidade = request.POST.get(
            'cidade'
        )

        fornecedor.uf = request.POST.get(
            'uf'
        )

        fornecedor.save()

        messages.success(
            request,
            'Fornecedor atualizado com sucesso.'
        )

        return redirect(
            'fornecedores'
        )

    return render(
        request,
        'accounts/fornecedor_form.html',
        {
            'fornecedor': fornecedor
        }
    )

# =========================================
# ALTERAR STATUS DO FORNECEDOR
# =========================================

@login_required
@perfil_required(
    "Administrador",
    "Secretária",
)
def alterar_status_fornecedor(
    request,
    fornecedor_id
):

    fornecedor = get_object_or_404(
        Fornecedor,
        id=fornecedor_id
    )

    fornecedor.ativo = not fornecedor.ativo

    fornecedor.save()

    messages.success(
        request,
        'Status do fornecedor atualizado.'
    )

    return redirect(
        'fornecedores'
    )

# =========================================
# EXCLUIR FORNECEDOR
# =========================================

@login_required
@permissao_required("fornecedores", "excluir")
def excluir_fornecedor(request, fornecedor_id):

    fornecedor = get_object_or_404(
        Fornecedor,
        id=fornecedor_id
    )

    fornecedor.delete()

    messages.success(
        request,
        "Fornecedor excluído com sucesso."
    )

    return redirect(
        "fornecedores"
    )
# =========================================
# PRODUTOS
# =========================================

from django.db.models import F

@login_required
@permissao_required("produtos", "visualizar")
def produtos(request):

    produtos = Produto.objects.order_by(
        'nome'
    )

    produtos_baixos = Produto.objects.filter(
        ativo=True,
        estoque__lte=F('estoque_minimo')
    )

    context = {

        'produtos': produtos,

        'produtos_baixos': produtos_baixos,

        'total_produtos_baixos': produtos_baixos.count()

    }

    return render(

        request,

        'accounts/produtos.html',

        context

    )

# =========================================
# NOVO PRODUTO
# =========================================

@login_required
@permissao_required("produtos", "inserir")
def novo_produto(request):

    if request.method == 'POST':

        nome = request.POST.get(
            'nome'
        )

        if not nome:

            messages.error(
                request,
                'Informe o nome do produto.'
            )

            return render(
                request,
                'accounts/produto_form.html'
            )

        Produto.objects.create(

            nome=nome,

            descricao=request.POST.get(
                'descricao'
            ),

            codigo=request.POST.get(
                'codigo'
            ),

            estoque=request.POST.get(
                'estoque'
            ) or 0,

            estoque_minimo=request.POST.get(
                'estoque_minimo'
            ) or 0,

            valor_compra=request.POST.get(
                'valor_compra'
            ) or 0,

            valor_venda=request.POST.get(
                'valor_venda'
            ) or 0,

            ativo=True

        )

        messages.success(
            request,
            'Produto cadastrado com sucesso.'
        )

        return redirect(
            'produtos'
        )

    return render(
        request,
        'accounts/produto_form.html'
    )

# =========================================
# EDITAR PRODUTO
# =========================================

@login_required
@permissao_required("produtos", "editar")
def editar_produto(request, produto_id):

    produto = get_object_or_404(
        Produto,
        id=produto_id
    )

    if request.method == 'POST':

        produto.nome = request.POST.get(
            'nome'
        )

        produto.descricao = request.POST.get(
            'descricao'
        )

        produto.codigo = request.POST.get(
            'codigo'
        )

        produto.estoque = request.POST.get(
            'estoque'
        ) or 0

        produto.estoque_minimo = request.POST.get(
            'estoque_minimo'
        ) or 0

        produto.valor_compra = request.POST.get(
            'valor_compra'
        ) or 0

        produto.valor_venda = request.POST.get(
            'valor_venda'
        ) or 0

        produto.save()

        messages.success(
            request,
            'Produto atualizado com sucesso.'
        )

        return redirect(
            'produtos'
        )

    context = {
        'produto': produto
    }

    return render(
        request,
        'accounts/produto_form.html',
        context
    )
# =========================================
# STATUS PRODUTO
# =========================================

@login_required
@perfil_required(
    "Administrador",
    "Secretária",
)
def alterar_status_produto(
    request,
    produto_id
):

    produto = get_object_or_404(
        Produto,
        id=produto_id
    )

    produto.ativo = not produto.ativo

    produto.save()

    return redirect(
        'produtos'
    )

# =========================================
# EXCLUIR PRODUTO
# =========================================

@login_required
@permissao_required("produtos", "excluir")
def excluir_produto(request, produto_id):

    produto = get_object_or_404(
        Produto,
        id=produto_id
    )

    produto.delete()

    messages.success(
        request,
        'Produto excluído com sucesso.'
    )

    return redirect(
        'produtos'
    )
# =========================================
# COMPRAS
# =========================================

@login_required
@permissao_required("compras", "visualizar")
def compras(request):

    compras = Compra.objects.select_related(
        'fornecedor'
    ).order_by('-id')

    context = {
        'compras': compras
    }

    return render(
        request,
        'accounts/compras.html',
        context
    )

# =========================================
# NOVA COMPRA
# =========================================

@login_required
@permissao_required("compras", "inserir")
def nova_compra(request):

    fornecedores = Fornecedor.objects.filter(
        ativo=True
    ).order_by('nome')

    produtos = Produto.objects.filter(
        ativo=True
    ).order_by('nome')

    if request.method == 'POST':

        fornecedor_id = request.POST.get(
            'fornecedor'
        )

        data_compra = request.POST.get(
            'data_compra'
        )

        vencimento = request.POST.get(
            'vencimento'
        )

        numero_nf = request.POST.get(
            'numero_nf'
        )

        observacoes = request.POST.get(
            'observacoes'
        )

        arquivo_nf = request.FILES.get(
            'arquivo_nf'
        )

        compra = Compra.objects.create(

            fornecedor_id=fornecedor_id,

            data_compra=data_compra,

            numero_nf=numero_nf,

            observacoes=observacoes,

            arquivo_nf=arquivo_nf,

            valor_total=0

        )

        produtos_ids = request.POST.getlist(
            'produto[]'
        )

        quantidades = request.POST.getlist(
            'quantidade[]'
        )

        valores = request.POST.getlist(
            'valor_unitario[]'
        )

        total_compra = Decimal('0.00')

        for i in range(
            len(produtos_ids)
        ):

            produto = Produto.objects.get(
                id=produtos_ids[i]
            )

            quantidade = int(
                quantidades[i]
            )

            valor_unitario = Decimal(
                valores[i]
            )

            subtotal = (
                quantidade *
                valor_unitario
            )

            ItemCompra.objects.create(

                compra=compra,

                produto=produto,

                quantidade=quantidade,

                valor_unitario=valor_unitario,

                subtotal=subtotal

            )

            # Atualiza estoque
            produto.estoque += quantidade

            produto.save()

            total_compra += subtotal

        compra.valor_total = total_compra

        compra.save()

        # =========================================
        # GERA CONTA A PAGAR AUTOMÁTICA
        # =========================================

        ContaPagar.objects.create(

            fornecedor=compra.fornecedor,

            compra=compra,

            descricao=(
                f'Compra NF '
                f'{compra.numero_nf or compra.id}'
            ),

            valor=compra.valor_total,

            vencimento=vencimento,

            observacao=(
                compra.observacoes
            ),

            status='PENDENTE'

        )

        messages.success(

            request,

            'Compra cadastrada com sucesso.'

        )

        return redirect(
            'compras'
        )

    context = {

        'fornecedores': fornecedores,

        'produtos': produtos

    }

    return render(

        request,

        'accounts/compra_form.html',

        context

    )


# =========================================
# VISUALIZAR COMPRA
# =========================================

@login_required
@perfil_required(
    "Administrador",
    "Secretária",
)
def visualizar_compra(request, compra_id):

    compra = get_object_or_404(
        Compra,
        id=compra_id
    )

    itens = ItemCompra.objects.filter(
        compra=compra
    )

    context = {

        'compra': compra,
        'itens': itens

    }

    return render(

        request,

        'accounts/compra_visualizar.html',

        context

    )

# =========================================
# EDITAR COMPRA
# =========================================

@login_required
@permissao_required("compras", "editar")
def editar_compra(request):

    compra = get_object_or_404(
        Compra,
        id=compra_id
    )

    item = ItemCompra.objects.filter(
        compra=compra
    ).first()

    fornecedores = Fornecedor.objects.filter(
        ativo=True
    ).order_by('nome')

    produtos = Produto.objects.filter(
        ativo=True
    ).order_by('nome')

    if request.method == 'POST':

        fornecedor_id = request.POST.get(
            'fornecedor'
        )

        data_compra = request.POST.get(
            'data_compra'
        )

        numero_nf = request.POST.get(
            'numero_nf'
        )

        observacoes = request.POST.get(
            'observacoes'
        )

        produto_id = request.POST.get(
            'produto'
        )

        nova_quantidade = int(
            request.POST.get(
                'quantidade',
                0
            )
        )

        novo_valor = Decimal(
            request.POST.get(
                'valor_unitario',
                0
            )
        )

        if item:

            produto_antigo = item.produto

            produto_antigo.estoque -= (
                item.quantidade
            )

            if produto_antigo.estoque < 0:

                produto_antigo.estoque = 0

            produto_antigo.save()

        produto = Produto.objects.get(
            id=produto_id
        )

        subtotal = (
            nova_quantidade *
            novo_valor
        )

        produto.estoque += (
            nova_quantidade
        )

        produto.save()

        compra.fornecedor_id = (
            fornecedor_id
        )

        compra.data_compra = (
            data_compra
        )

        compra.numero_nf = (
            numero_nf
        )

        compra.observacoes = (
            observacoes
        )

        compra.valor_total = (
            subtotal
        )

        compra.save()

        if item:

            item.produto = produto

            item.quantidade = (
                nova_quantidade
            )

            item.valor_unitario = (
                novo_valor
            )

            item.subtotal = (
                subtotal
            )

            item.save()

        messages.success(

            request,

            'Compra atualizada com sucesso.'

        )

        return redirect(
            'compras'
        )

    context = {

        'compra': compra,

        'item': item,

        'fornecedores': fornecedores,

        'produtos': produtos

    }

    return render(

        request,

        'accounts/compra_editar.html',

        context

    )

# =========================================
# EXCLUIR COMPRA
# =========================================

@login_required
@permissao_required("compras", "excluir")
def excluir_compra(request):

    compra = get_object_or_404(
        Compra,
        id=compra_id
    )

    itens = ItemCompra.objects.filter(
        compra=compra
    )

    for item in itens:

        produto = item.produto

        produto.estoque -= item.quantidade

        if produto.estoque < 0:

            produto.estoque = 0

        produto.save()

    compra.delete()

    messages.success(
        request,
        'Compra excluída com sucesso.'
    )

    return redirect(
        'compras'
    )

# =========================================
# ESTOQUE
# =========================================

from django.db.models import F
from decimal import Decimal

@login_required
@permissao_required("estoque", "visualizar")
def estoque(request):

    produtos = Produto.objects.filter(
        ativo=True
    ).order_by('nome')

    produtos_baixos = produtos.filter(
        estoque__lte=F('estoque_minimo')
    )

    valor_total_estoque = Decimal('0.00')

    for produto in produtos:

        produto.valor_estoque = (
            produto.estoque *
            produto.valor_compra
        )

        valor_total_estoque += (
            produto.valor_estoque
        )

    context = {

        'produtos': produtos,

        'produtos_baixos': produtos_baixos,

        'total_produtos': produtos.count(),

        'total_produtos_baixos': produtos_baixos.count(),

        'valor_total_estoque': valor_total_estoque

    }

    return render(

        request,

        'accounts/estoque.html',

        context

    )

# =========================================
# MOVIMENTAÇÕES DE ESTOQUE
# =========================================

@login_required
@permissao_required("movimentacoes_estoque", "visualizar")
def movimentacoes_estoque(request):

    movimentacoes = (
        MovimentacaoEstoque.objects
        .select_related(
            'produto',
            'usuario'
        )
        .order_by(
            '-criado_em'
        )
    )

    produtos = Produto.objects.filter(
        ativo=True
    ).order_by(
        'nome'
    )

    produto_id = request.GET.get(
        'produto'
    )

    tipo = request.GET.get(
        'tipo'
    )

    data_inicial = request.GET.get(
        'data_inicial'
    )

    data_final = request.GET.get(
        'data_final'
    )

    if produto_id:

        movimentacoes = movimentacoes.filter(
            produto_id=produto_id
        )

    if tipo:

        movimentacoes = movimentacoes.filter(
            tipo=tipo
        )

    if data_inicial:

        movimentacoes = movimentacoes.filter(
            criado_em__date__gte=data_inicial
        )

    if data_final:

        movimentacoes = movimentacoes.filter(
            criado_em__date__lte=data_final
        )

    context = {

        'movimentacoes': movimentacoes,

        'produtos': produtos,

        'produto_id': produto_id,

        'tipo': tipo,

        'data_inicial': data_inicial,

        'data_final': data_final

    }

    return render(

        request,

        'accounts/movimentacoes_estoque.html',

        context

    )

# =========================================
# FUNÇÃO AUXILIAR
# =========================================

def registrar_movimentacao(
    produto,
    tipo,
    quantidade,
    usuario,
    observacao=''
):

    estoque_anterior = produto.estoque

    if tipo == 'SAIDA':

        if quantidade > produto.estoque:

            raise ValueError(
                'Estoque insuficiente.'
            )

        produto.estoque -= quantidade

    elif tipo in [

        'COMPRA',
        'ENTRADA'

    ]:

        produto.estoque += quantidade

    elif tipo == 'AJUSTE':

        produto.estoque = quantidade

    produto.save()

    MovimentacaoEstoque.objects.create(

        produto=produto,

        tipo=tipo,

        quantidade=quantidade,

        estoque_anterior=estoque_anterior,

        estoque_atual=produto.estoque,

        observacao=observacao,

        usuario=usuario

    )


# =========================================
# NOVA MOVIMENTAÇÃO DE ESTOQUE
# =========================================

@login_required
@permissao_required("movimentacoes_estoque", "inserir")
def nova_movimentacao_estoque(request):

    produtos = Produto.objects.filter(
        ativo=True
    ).order_by(
        'nome'
    )

    if request.method == 'POST':

        produto = get_object_or_404(
            Produto,
            id=request.POST.get('produto')
        )

        tipo = request.POST.get(
            'tipo'
        )

        quantidade = int(
            request.POST.get(
                'quantidade'
            ) or 0
        )

        observacao = request.POST.get(
            'observacao'
        )

        try:

            registrar_movimentacao(
                produto=produto,
                tipo=tipo,
                quantidade=quantidade,
                usuario=request.user,
                observacao=observacao
            )

            messages.success(
                request,
                'Movimentação registrada com sucesso.'
            )

            return redirect(
                'movimentacoes_estoque'
            )

        except ValueError as erro:

            messages.error(
                request,
                str(erro)
            )

    context = {

        'produtos': produtos

    }

    return render(

        request,

        'accounts/movimentacao_form.html',

        context

    )


# =========================================
# PRODUTOS CRÍTICOS
# =========================================

@login_required
@permissao_required("produtos_criticos", "visualizar")
def produtos_criticos(request):

    produtos = Produto.objects.filter(
        ativo=True,
        estoque__lte=F('estoque_minimo')
    ).order_by(
        'nome'
    )

    context = {

        'produtos': produtos,

        'total_produtos': produtos.count()

    }

    return render(

        request,

        'accounts/produtos_criticos.html',

        context

    )

# =========================================
# LOTES
# =========================================

from datetime import date, timedelta

@login_required
@permissao_required("lotes", "visualizar")
def lotes(request):

    lotes = LoteProduto.objects.select_related(
        'produto'
    ).order_by(
        'validade'
    )

    hoje = date.today()

    for lote in lotes:

        dias = (
            lote.validade - hoje
        ).days

        if dias < 0:

            lote.status_validade = 'VENCIDO'

        elif dias <= 30:

            lote.status_validade = '30_DIAS'

        elif dias <= 60:

            lote.status_validade = '60_DIAS'

        else:

            lote.status_validade = 'VALIDO'

    context = {

        'lotes': lotes,

        'total_lotes': lotes.count()

    }

    return render(

        request,

        'accounts/lotes.html',

        context

    )

# =========================================
# NOVO LOTE
# =========================================

@login_required
@permissao_required("lotes", "inserir")
def novo_lote(request):

    produtos = Produto.objects.filter(
        ativo=True
    ).order_by(
        'nome'
    )

    if request.method == 'POST':

        LoteProduto.objects.create(

            produto_id=request.POST.get(
                'produto'
            ),

            lote=request.POST.get(
                'lote'
            ),

            quantidade=request.POST.get(
                'quantidade'
            ) or 0,

            validade=request.POST.get(
                'validade'
            )

        )

        messages.success(
            request,
            'Lote cadastrado com sucesso.'
        )

        return redirect(
            'lotes'
        )

    return render(

        request,

        'accounts/lote_form.html',

        {
            'produtos': produtos
        }

    )

# =========================================
# CONTAS A PAGAR
# =========================================

@login_required
@permissao_required("contas_pagar", "visualizar")
def contas_pagar(request):

    hoje = timezone.now().date()

    # Atualiza contas vencidas automaticamente
    ContaPagar.objects.filter(
        status='PENDENTE',
        vencimento__lt=hoje
    ).update(
        status='VENCIDO'
    )

    contas = ContaPagar.objects.select_related(
        'fornecedor'
    ).order_by(
        'vencimento'
    )

    # Filtro por status
    status = request.GET.get(
        'status'
    )

    if status:

        contas = contas.filter(
            status=status
        )

    contas_pendentes = ContaPagar.objects.filter(
        status='PENDENTE'
    )

    contas_pagas = ContaPagar.objects.filter(
        status='PAGO'
    )

    contas_vencidas = ContaPagar.objects.filter(
        status='VENCIDO'
    )

    total_pendente = sum(
        conta.valor
        for conta in contas_pendentes
    )

    total_pago = sum(
        conta.valor
        for conta in contas_pagas
    )

    total_vencido = sum(
        conta.valor
        for conta in contas_vencidas
    )

    context = {

        'contas': contas,

        'status': status,

        'total_pendente': total_pendente,

        'total_pago': total_pago,

        'total_vencido': total_vencido,

        'quantidade_pendentes': contas_pendentes.count(),

        'quantidade_pagas': contas_pagas.count(),

        'quantidade_vencidas': contas_vencidas.count(),

    }

    return render(

        request,

        'accounts/contas_pagar.html',

        context

    )

# =========================================
# NOVA CONTA A PAGAR
# =========================================

@login_required
@permissao_required("contas_pagar", "inserir")
def nova_conta_pagar(request):

    fornecedores = Fornecedor.objects.filter(
        ativo=True
    ).order_by(
        'nome'
    )

    if request.method == 'POST':

        ContaPagar.objects.create(

            fornecedor_id=request.POST.get(
                'fornecedor'
            ),

            descricao=request.POST.get(
                'descricao'
            ),

            valor=request.POST.get(
                'valor'
            ),

            vencimento=request.POST.get(
                'vencimento'
            ),

            status='PENDENTE'

        )

        messages.success(
            request,
            'Conta cadastrada com sucesso.'
        )

        return redirect(
            'contas_pagar'
        )

    return render(

        request,

        'accounts/nova_conta_pagar.html',

        {

            'fornecedores': fornecedores

        }

    )

# =========================================
# PAGAR CONTA
# =========================================

@login_required
@permissao_required("contas_pagar", "financeiro")
def pagar_conta(request, conta_id):

    conta = get_object_or_404(
        ContaPagar,
        id=conta_id
    )

    if conta.status != 'PAGO':

        conta.status = 'PAGO'

        conta.data_pagamento = timezone.now().date()

        conta.save()

        # Lança saída no Caixa
        Caixa.objects.create(

            data=timezone.now().date(),

            descricao=conta.descricao,

            tipo='SAIDA',

            valor=conta.valor,

            conta_pagar=conta

        )

        messages.success(

            request,

            'Conta paga com sucesso.'

        )

    return redirect(
        'contas_pagar'
    )

# =========================================
# CONTAS A RECEBER
# =========================================

@login_required(login_url='/')
@permissao_required("contas_receber")
def contas_receber(request):

    hoje = timezone.now().date()

    # Atualiza automaticamente contas vencidas
    ContaReceber.objects.filter(
        status='PENDENTE',
        vencimento__lt=hoje
    ).update(
        status='VENCIDO'
    )

    # Não exibe contas canceladas
    contas = ContaReceber.objects.select_related(
        'paciente',
        'orcamento'
    ).exclude(
        status='CANCELADO'
    ).order_by(
        'vencimento'
    )

    contas_pendentes = contas.filter(
        status='PENDENTE'
    )

    contas_recebidas = contas.filter(
        status='RECEBIDO'
    )

    contas_vencidas = contas.filter(
        status='VENCIDO'
    )

    context = {

        'contas': contas,

        'total_pendente': sum(
            conta.valor
            for conta in contas_pendentes
        ),

        'total_recebido': sum(
            conta.valor
            for conta in contas_recebidas
        ),

        'total_vencido': sum(
            conta.valor
            for conta in contas_vencidas
        ),

    }

    return render(

        request,

        'accounts/contas_receber.html',

        context

    )

# =========================================
# RECEBER CONTA
# =========================================

@login_required(login_url='/')
@permissao_required("contas_receber", "financeiro")
def receber_conta(request, conta_id):

    conta = get_object_or_404(

        ContaReceber,

        id=conta_id

    )

    if conta.status != 'RECEBIDO':

        conta.status = 'RECEBIDO'

        conta.data_recebimento = timezone.now().date()

        conta.save()

        Caixa.objects.create(

            data=timezone.now().date(),

            descricao=conta.descricao,

            tipo='ENTRADA',

            valor=conta.valor,

            conta_receber=conta

        )

        messages.success(

            request,

            'Conta recebida com sucesso.'

        )

    return redirect(

        'contas_receber'

    )

# =========================================
# CAIXA
# =========================================

@login_required
@permissao_required("caixa", "visualizar")

def caixa(request):

    movimentacoes = Caixa.objects.all()

    total_entradas = sum(
        item.valor
        for item in movimentacoes
        if item.tipo == 'ENTRADA'
    )

    total_saidas = sum(
        item.valor
        for item in movimentacoes
        if item.tipo == 'SAIDA'
    )

    saldo = total_entradas - total_saidas

    context = {

        'movimentacoes': movimentacoes,

        'total_entradas': total_entradas,

        'total_saidas': total_saidas,

        'saldo': saldo,

    }

    return render(

        request,

        'accounts/caixa.html',

        context

    )

# =========================================
# CENTRAL ORÇAMENTO
# =========================================

def central_orcamentos(request):

    status = request.GET.get('status')

    orcamentos = Orcamento.objects.all()

    if status:
        orcamentos = orcamentos.filter(status=status)

    total_orcamentos = Orcamento.objects.count()

    valor_total = 0
    valor_aprovado = 0

    for orcamento in Orcamento.objects.all():

        valor_total += orcamento.total

        if orcamento.status == 'aprovado':
            valor_aprovado += orcamento.total

    percentual_conversao = 0

    if valor_total > 0:

        percentual_conversao = round(
            (valor_aprovado / valor_total) * 100,
            1
        )

    rascunhos = Orcamento.objects.filter(
        status='rascunho'
    ).count()

    aprovados = Orcamento.objects.filter(
        status='aprovado'
    ).count()

    finalizados = Orcamento.objects.filter(
        status='finalizado'
    ).count()

    cancelados = Orcamento.objects.filter(
        status='cancelado'
    ).count()

    context = {

        'orcamentos': orcamentos,

        'status': status,

        'total_orcamentos': total_orcamentos,

        'valor_total': valor_total,

        'valor_aprovado': valor_aprovado,

        'percentual_conversao': percentual_conversao,

        'rascunhos': rascunhos,

        'aprovados': aprovados,

        'finalizados': finalizados,

        'cancelados': cancelados,

    }

    return render(
        request,
        'accounts/central_orcamentos.html',
        context
    )

# =========================================
# ENCERRAR TRATAMENTO
# =========================================

@login_required(login_url="/")
@permissao_required("odontograma", "editar")
def encerrar_tratamento(request, tratamento_id):

    if request.method != "POST":

        return redirect(
            "odontograma",
            id=get_object_or_404(
                Tratamento,
                id=tratamento_id
            ).paciente.id
        )

    tratamento = get_object_or_404(
        Tratamento,
        id=tratamento_id
    )

    # Evita encerrar duas vezes
    if tratamento.status == "ENCERRADO":

        messages.warning(
            request,
            "Este tratamento já está encerrado."
        )

        return redirect(
            "odontograma",
            id=tratamento.paciente.id
        )

    # =========================================
    # ENCERRA O TRATAMENTO
    # =========================================

    tratamento.status = "ENCERRADO"
    tratamento.data_encerramento = timezone.now().date()

    tratamento.save()

    # =========================================
    # FINALIZA O ORÇAMENTO DO TRATAMENTO
    # =========================================

    orcamento = Orcamento.objects.filter(
        paciente=tratamento.paciente,
        tratamento=tratamento,
        status="aprovado"
    ).order_by("-id").first()

    if orcamento:

        orcamento.status = "finalizado"
        orcamento.save()

    messages.success(
        request,
        "Tratamento encerrado e orçamento finalizado com sucesso."
    )

    return redirect(
        "odontograma",
        id=tratamento.paciente.id
    )


# =========================================
# CANCELAR TRATAMENTO
# =========================================

@login_required(login_url="/")
def cancelar_tratamento(request, tratamento_id):

    if request.method != "POST":

        return redirect(
            "odontograma",
            id=get_object_or_404(
                Tratamento,
                id=tratamento_id
            ).paciente.id
        )

    tratamento = get_object_or_404(
        Tratamento,
        id=tratamento_id
    )

    # Evita cancelar duas vezes
    if tratamento.status == "CANCELADO":

        messages.warning(
            request,
            "Este tratamento já está cancelado."
        )

        return redirect(
            "odontograma",
            id=tratamento.paciente.id
        )

    # =========================================
    # CANCELA O TRATAMENTO
    # =========================================

    tratamento.status = "CANCELADO"
    tratamento.data_encerramento = timezone.now().date()

    tratamento.save()

    # =========================================
    # CANCELA O ORÇAMENTO DO TRATAMENTO
    # =========================================

    orcamento = Orcamento.objects.filter(
        paciente=tratamento.paciente,
        tratamento=tratamento,
        status="aprovado"
    ).order_by("-id").first()

    if orcamento:

        # Cancela o orçamento
        orcamento.status = "cancelado"
        orcamento.save()

        # =========================================
        # CANCELA AS CONTAS A RECEBER
        # =========================================

        ContaReceber.objects.filter(
            orcamento=orcamento,
            status__in=[
                "PENDENTE",
                "VENCIDO"
            ]
        ).update(
            status="CANCELADO"
        )

    messages.success(
        request,
        "Tratamento, orçamento e contas a receber cancelados com sucesso."
    )

    return redirect(
        "odontograma",
        id=tratamento.paciente.id
    )


   
# =========================================
# NOVO TRATAMENTO
# =========================================

@login_required(login_url="/")
@permissao_required("odontograma", "inserir")
def novo_tratamento(request, paciente_id):

    paciente = get_object_or_404(
        Paciente,
        id=paciente_id
    )

    if request.method != "POST":

        return redirect(
            "odontograma",
            id=paciente.id
        )

    # Verifica se já existe tratamento ativo
    tratamento_ativo = paciente.tratamentos.filter(
        status="ATIVO"
    ).first()

    if tratamento_ativo:

        messages.warning(
            request,
            "Já existe um tratamento ativo para este paciente."
        )

        return redirect(
            "odontograma",
            id=paciente.id
        )

    titulo = request.POST.get(
        "titulo",
        ""
    ).strip()

    observacoes = request.POST.get(
        "observacoes",
        ""
    ).strip()

    if not titulo:

        titulo = "Novo Tratamento"

    tratamento = Tratamento.objects.create(

        paciente=paciente,

        titulo=titulo,

        observacoes=observacoes,

        status="ATIVO"

    )

    Orcamento.objects.get_or_create(

        paciente=paciente,

        tratamento=tratamento,

        defaults={
            "tratamento": tratamento
        }

    )

    messages.success(
        request,
        "Novo tratamento criado com sucesso."
    )

    return redirect(
        "odontograma",
        id=paciente.id
    )

# =========================================
# TRATAMENTOS DO PACIENTE
# =========================================

@login_required(login_url='/')
@perfil_required(
    "Administrador",
    "Dentista",
    "Auxiliar de Saúde Bucal",
    "Secretária",
)
def tratamentos_paciente(request, id):

    print("ID recebido:", id)

    print("Pacientes cadastrados:")

    for p in Paciente.objects.all():

        print(
            p.id,
            p.nome
        )

    paciente = get_object_or_404(
        Paciente,
        id=id
    )

    tratamentos = Tratamento.objects.filter(
        paciente=paciente
    ).order_by(
        '-data_inicio',
        '-id'
    )

    context = {

        'paciente': paciente,

        'tratamentos': tratamentos,

    }

    return render(
        request,
        'accounts/tratamentos.html',
        context
    )

# =========================================
# CENTRAL RELATÓRIOS
# =========================================
@login_required
def central_relatorios(request):

    return render(
        request,
        'accounts/relatorios/central_relatorios.html'
    )

@login_required
def relatorio_pacientes(request):

    return render(
        request,
        'accounts/relatorios/pacientes.html'
    )


@login_required
def relatorio_tratamentos(request):

    return render(
        request,
        'accounts/relatorios/tratamentos.html'
    )


@login_required
def relatorio_orcamentos(request):

    return render(
        request,
        'accounts/relatorios/orcamentos.html'
    )


@login_required
def relatorio_agenda(request):

    return render(
        request,
        'accounts/relatorios/agenda.html'
    )


@login_required
def relatorio_receber(request):

    return render(
        request,
        'accounts/relatorios/receber.html'
    )


@login_required
def relatorio_pagar(request):

    return render(
        request,
        'accounts/relatorios/pagar.html'
    )


@login_required
def relatorio_caixa(request):

    return render(
        request,
        'accounts/relatorios/caixa.html'
    )


@login_required
def relatorio_indicadores(request):

    return render(
        request,
        'accounts/relatorios/indicadores.html'
    )


@login_required
def relatorio_producao(request):

    return render(
        request,
        'accounts/relatorios/producao.html'
    )


@login_required
def relatorio_faturamento(request):

    return render(
        request,
        'accounts/relatorios/faturamento.html'
    )

# =========================================
# RELATÓRIO DE PACIENTES
# =========================================

@login_required
def relatorio_pacientes(request):

    pacientes = Paciente.objects.all()

    # Nome
    q = request.GET.get('q')

    if q:
        pacientes = pacientes.filter(
            nome__icontains=q
        )

    # Convênio
    convenio = request.GET.get("convenio")

    if convenio:

        try:

            convenio_obj = Convenio.objects.get(id=convenio)

            pacientes = pacientes.filter(
                convenio=convenio_obj.nome
            )

        except Convenio.DoesNotExist:
            pass

    # =========================================
    # STATUS
    # =========================================

    status = request.GET.get('status')

    if status != "":

        if status == "1":

            pacientes = pacientes.filter(
                ativo=True
            )

        elif status == "0":

            pacientes = pacientes.filter(
                ativo=False
            )

    # =========================================
    # ORDENAÇÃO
    # =========================================

    pacientes = pacientes.order_by('nome')

    # =========================================
    # PAGINAÇÃO
    # =========================================

    paginator = Paginator(
        pacientes,
        20
    )

    page = request.GET.get('page')

    pacientes = paginator.get_page(page)

    # =========================================
    # DADOS AUXILIARES
    # =========================================

    convenios = Convenio.objects.all()

    total_pacientes = Paciente.objects.count()

    pacientes_ativos = Paciente.objects.filter(
        ativo=True
    ).count()

    pacientes_inativos = Paciente.objects.filter(
        ativo=False
    ).count()

    total_convenios = Convenio.objects.count()

    # =========================================
    # CONTEXTO
    # =========================================

    context = {

        'pacientes': pacientes,

        'convenios': convenios,

        'total_pacientes': total_pacientes,

        'pacientes_ativos': pacientes_ativos,

        'pacientes_inativos': pacientes_inativos,

        'total_convenios': total_convenios,

    }

    return render(
        request,
        'accounts/relatorios/pacientes.html',
        context
    )

# =========================================
# RELATÓRIO DE ANIVERSARIANTES
# =========================================

@login_required
def relatorio_aniversariantes(request):

    pacientes = Paciente.objects.all()

    # =========================================
    # MÊS
    # =========================================

    mes = request.GET.get("mes", "").strip()

    if mes:
        pacientes = pacientes.filter(
            nascimento__month=int(mes)
        )

    # =========================================
    # NOME
    # =========================================

    q = (request.GET.get("q") or "").strip()

    if q.lower() == "none":
        q = ""

    if q:
        pacientes = pacientes.filter(
            nome__icontains=q
        )

    # =========================================
    # STATUS
    # =========================================

    status = request.GET.get("status", "").strip()

    if status == "1":
        pacientes = pacientes.filter(
            ativo=True
        )

    elif status == "0":
        pacientes = pacientes.filter(
            ativo=False
        )

    # =========================================
    # ORDENAÇÃO
    # =========================================

    pacientes = pacientes.order_by(
        "nascimento__month",
        "nascimento__day",
        "nome"
    )

    # =========================================
    # PAGINAÇÃO
    # =========================================

    paginator = Paginator(
        pacientes,
        20
    )

    page = request.GET.get("page")

    pacientes = paginator.get_page(page)

    # =========================================
    # CONTEXTO
    # =========================================

    context = {
        "pacientes": pacientes,
        "mes": mes,
        "q": q,
        "status": status,
        "total_aniversariantes": pacientes.paginator.count,
    }

    return render(
        request,
        "accounts/relatorios/aniversariantes.html",
        context
    )

# =========================================
# PDF RELATÓRIO DE PACIENTES
# =========================================

@login_required(login_url='/')
def gerar_pdf_relatorio_pacientes(request):

    pacientes = Paciente.objects.all()

    # =========================================
    # FILTRO POR NOME
    # =========================================

    q = request.GET.get('q', '').strip()

    if q:

        pacientes = pacientes.filter(
            nome__icontains=q
        )

    # =========================================
    # FILTRO POR CONVÊNIO
    # =========================================

    convenio = request.GET.get('convenio', '').strip()

    if convenio:

        try:

            convenio_obj = Convenio.objects.get(
                id=convenio
            )

            pacientes = pacientes.filter(
                convenio=convenio_obj.nome
            )

        except Convenio.DoesNotExist:

            pass

    # =========================================
    # FILTRO POR STATUS
    # =========================================

    status = request.GET.get('status', '').strip()

    if status == "1":

        pacientes = pacientes.filter(
            ativo=True
        )

    elif status == "0":

        pacientes = pacientes.filter(
            ativo=False
        )

    # =========================================
    # ORDENAÇÃO
    # =========================================

    pacientes = pacientes.order_by('nome')

    # =========================================
    # CONFIGURAÇÕES DA CLÍNICA
    # =========================================

    config = ConfiguracaoClinica.objects.first()

    # =========================================
    # CONTEXT
    # =========================================

    context = {

        'pacientes': pacientes,

        'config': config,

        'q': q,

        'convenio': convenio,

        'status': status,

        'now': timezone.now(),

        'total_pacientes': pacientes.count(),

        'titulo_relatorio': 'RELATÓRIO DE PACIENTES'

    }

    # =========================================
    # TEMPLATE
    # =========================================

    return render(
        request,
        'accounts/pdf/pacientes_pdf.html',
        context
    )

    # =========================================
# PDF RELATÓRIO DE ANIVERSARIANTES
# =========================================

@login_required(login_url='/')
def gerar_pdf_relatorio_aniversariantes(request):

    pacientes = Paciente.objects.all()

    # Nome
    q = request.GET.get('q', '').strip()

    if q:
        pacientes = pacientes.filter(
            nome__icontains=q
        )

    # Mês
    mes = request.GET.get('mes', '').strip()

    if mes:
        pacientes = pacientes.filter(
            nascimento__month=mes
        )

    # Status
    status = request.GET.get('status', '').strip()

    if status == "1":

        pacientes = pacientes.filter(
            ativo=True
        )

    elif status == "0":

        pacientes = pacientes.filter(
            ativo=False
        )

    # Ordenação
    pacientes = pacientes.order_by(
        'nascimento__month',
        'nascimento__day',
        'nome'
    )

    config = ConfiguracaoClinica.objects.first()

    context = {

        'pacientes': pacientes,

        'config': config,

        'q': q,

        'mes': mes,

        'status': status,

        'now': timezone.now(),

        'total_pacientes': pacientes.count(),

        'titulo_relatorio': 'RELATÓRIO DE ANIVERSARIANTES',

    }

    return render(
        request,
        'accounts/pdf/aniversariantes_pdf.html',
        context
        )

# =========================================
# DADOS - ORÇAMENTOS
# =========================================

def obter_dados_orcamentos(request):

    # =========================================
    # FILTROS
    # =========================================

    paciente = request.GET.get("paciente", "")
    status = request.GET.get("status", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_final = request.GET.get("data_final", "")

    # =========================================
    # QUERY
    # =========================================

    orcamentos = (
        Orcamento.objects
        .select_related(
            "paciente",
            "tratamento",
        )
        .order_by("-criado_em")
    )

    orcamentos = filtrar_escopo(
        request.user,
        orcamentos
    )

    # =========================================
    # ESCOPO DO USUÁRIO
    # =========================================

    perfil = getattr(request.user, "perfil", None)

    if perfil:

        # Dentista vê apenas seus pacientes
        if perfil.tipo_usuario == PerfilUsuario.DENTISTA:

            orcamentos = orcamentos.filter(
                paciente__dentista=request.user
            )

        # Auxiliar de Saúde Bucal
        elif perfil.tipo_usuario == PerfilUsuario.ACD:

            orcamentos = orcamentos.filter(
                paciente__dentista=request.user
            )

        # Administrador, Gestor, Secretária,
        # Contabilidade e Auditoria veem todos.
    # =========================================
    # FILTRO PACIENTE
    # =========================================

    if paciente:
        orcamentos = orcamentos.filter(
            paciente__nome__icontains=paciente
        )

    # =========================================
    # FILTRO STATUS
    # =========================================

    if status:
        orcamentos = orcamentos.filter(
            status=status
        )

    # =========================================
    # FILTRO DATA
    # =========================================

    if data_inicio:
        orcamentos = orcamentos.filter(
            criado_em__date__gte=data_inicio
        )

    if data_final:
        orcamentos = orcamentos.filter(
            criado_em__date__lte=data_final
        )

    # =========================================
    # INDICADORES
    # =========================================

    total_orcamentos = orcamentos.count()

    orcamentos_aprovados = orcamentos.filter(
        status="aprovado"
    ).count()

    orcamentos_execucao = orcamentos.filter(
        status="em_execucao"
    ).count()

    orcamentos_finalizados = orcamentos.filter(
        status="finalizado"
    ).count()

    # =========================================
    # CONTEXT
    # =========================================

    return {
        "orcamentos": orcamentos,
        "paciente": paciente,
        "status": status,
        "data_inicio": data_inicio,
        "data_final": data_final,
        "total_orcamentos": total_orcamentos,
        "orcamentos_aprovados": orcamentos_aprovados,
        "orcamentos_execucao": orcamentos_execucao,
        "orcamentos_finalizados": orcamentos_finalizados,
    }

# =========================================
# RELATÓRIO DE ORÇAMENTOS
# =========================================

@login_required
def relatorio_orcamentos(request):

    context = obter_dados_orcamentos(request)

    context.update({

        "config": ConfiguracaoClinica.objects.first(),

        "titulo_relatorio": "RELATÓRIO DE ORÇAMENTOS",

        "data_emissao": timezone.localtime(),

    })

    return render(

        request,

        "accounts/relatorios/orcamentos.html",

        context,

    )

# =========================================
# PDF - RELATÓRIO DE ORÇAMENTOS
# =========================================

@login_required
def relatorio_orcamentos_pdf(request):

    context = obter_dados_orcamentos(request)

    config = ConfiguracaoClinica.objects.first()

    context.update({

        "config": config,

        "titulo_relatorio": "RELATÓRIO DE ORÇAMENTOS",

        "now": timezone.now(),

    })

    return render(

        request,

        "accounts/pdf/relatorio_orcamentos_pdf.html",

        context,

    )

# =========================================
# EXCEL - RELATÓRIO DE ORÇAMENTOS
# =========================================

@login_required
def excel_orcamentos(request):

    context = obter_dados_orcamentos(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Orçamentos"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:H1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE ORÇAMENTOS"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:H2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Total de Orçamentos"
    ws["B4"] = context["total_orcamentos"]
    ws["B4"].number_format = "#,##0"

    ws["A5"] = "Aprovados"
    ws["B5"] = context["orcamentos_aprovados"]
    ws["B5"].number_format = "#,##0"

    ws["A6"] = "Em Execução"
    ws["B6"] = context["orcamentos_execucao"]
    ws["B6"].number_format = "#,##0"

    ws["A7"] = "Finalizados"
    ws["B7"] = context["orcamentos_finalizados"]
    ws["B7"].number_format = "#,##0"

    for linha in range(4, 8):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Paciente",

        "Tratamento",

        "Data",

        "Status",

        "Valor",

        "Entrada",

        "Desconto",

        "Forma Pagamento",

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 10

    for coluna, texto in enumerate(

        cabecalho,

        start=1

    ):

        cell = ws.cell(

            row=linha_inicio,

            column=coluna

        )

        cell.value = texto

        cell.font = Font(

            bold=True,

            color="FFFFFF"

        )

        cell.fill = fill

        cell.alignment = Alignment(

            horizontal="center"

        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for orcamento in context["orcamentos"]:

        # Paciente
        ws.cell(
            linha,
            1
        ).value = orcamento.paciente.nome

        # Tratamento
        ws.cell(
            linha,
            2
        ).value = (
            orcamento.tratamento.titulo
            if orcamento.tratamento
            else "-"
        )

        # Data
        celula = ws.cell(
            linha,
            3
        )

        if orcamento.criado_em:

            celula.value = orcamento.criado_em.replace(tzinfo=None)

            celula.number_format = "DD/MM/YYYY"

        # Status
        ws.cell(
            linha,
            4
        ).value = orcamento.status

        # Valor
        celula = ws.cell(
            linha,
            5
        )

        celula.value = float(
            orcamento.total or 0
        )

        celula.number_format = 'R$ #,##0.00'

        # Entrada
        celula = ws.cell(
            linha,
            6
        )

        celula.value = float(
            orcamento.entrada or 0
        )

        celula.number_format = 'R$ #,##0.00'

        # Desconto
        celula = ws.cell(
            linha,
            7
        )

        celula.value = float(
            orcamento.desconto or 0
        )

        celula.number_format = 'R$ #,##0.00'

        # Forma de Pagamento
        ws.cell(
            linha,
            8
        ).value = (
            orcamento.get_forma_pagamento_display()
            if hasattr(
                orcamento,
                "get_forma_pagamento_display"
            )
            else orcamento.forma_pagamento
        )

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 35,

        "B": 35,

        "C": 18,

        "D": 18,

        "E": 18,

        "F": 18,

        "G": 18,

        "H": 22,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Orcamentos_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response

# =========================================
# DADOS - ATENDIMENTOS
# =========================================

def obter_dados_atendimentos(request):

    # =========================================
    # FILTROS
    # =========================================

    paciente = request.GET.get("paciente", "")

    profissional = request.GET.get("profissional", "")

    procedimento = request.GET.get("procedimento", "")

    status = request.GET.get("status", "")

    data_inicio = request.GET.get("data_inicio", "")

    data_final = request.GET.get("data_final", "")

    # =========================================
    # CONSULTA
    # =========================================

    atendimentos = (

        Agendamento.objects

        .select_related(

            "paciente",

            "profissional",

            "procedimento",

        )

        .all()

        .order_by(

            "-data",

            "-hora_inicio",

        )

    )

    # =========================================
    # FILTRO PACIENTE
    # =========================================

    if paciente:

        atendimentos = atendimentos.filter(

            paciente__nome__icontains=paciente

        )

    # =========================================
    # FILTRO PROFISSIONAL
    # =========================================

    if profissional:

        atendimentos = atendimentos.filter(

            profissional__nome__icontains=profissional

        )

    # =========================================
    # FILTRO PROCEDIMENTO
    # =========================================

    if procedimento:

        atendimentos = atendimentos.filter(

            procedimento__nome__icontains=procedimento

        )

    # =========================================
    # FILTRO STATUS
    # =========================================

    if status:

        atendimentos = atendimentos.filter(

            status=status

        )

    # =========================================
    # FILTRO DATA
    # =========================================

    if data_inicio:

        atendimentos = atendimentos.filter(

            data__gte=data_inicio

        )

    if data_final:

        atendimentos = atendimentos.filter(

            data__lte=data_final

        )

    # =========================================
    # INDICADORES
    # =========================================

    total_atendimentos = atendimentos.count()

    total_confirmados = atendimentos.filter(
        status="confirmado"
    ).count()

    total_atendimento = atendimentos.filter(
        status="atendimento"
    ).count()

    total_finalizados = atendimentos.filter(
        status="finalizado"
    ).count()

    total_faltou = atendimentos.filter(
        status="faltou"
    ).count()

    total_cancelados = atendimentos.filter(
        status="cancelado"
    ).count()

    # =========================================
    # CONTEXT
    # =========================================

    return {

        "atendimentos": atendimentos,

        "paciente": paciente,

        "profissional": profissional,

        "procedimento": procedimento,

        "status": status,

        "data_inicio": data_inicio,

        "data_final": data_final,

        "total_atendimentos": total_atendimentos,

        "total_confirmados": total_confirmados,

        "total_atendimento": total_atendimento,

        "total_finalizados": total_finalizados,

        "total_faltou": total_faltou,

        "total_cancelados": total_cancelados,

    }

# =========================================
# RELATÓRIO DE ATENDIMENTOS
# =========================================

@login_required
def relatorio_atendimentos(request):

    context = obter_dados_atendimentos(request)

    context.update({

        "config": ConfiguracaoClinica.objects.first(),

        "titulo_relatorio": "RELATÓRIO DE ATENDIMENTOS",

        "data_emissao": timezone.localtime(),

    })

    return render(

        request,

        "accounts/relatorios/atendimentos.html",

        context,

    )

# =========================================
# PDF - RELATÓRIO DE ATENDIMENTOS
# =========================================

@login_required
def relatorio_atendimentos_pdf(request):

    context = obter_dados_atendimentos(request)

    config = ConfiguracaoClinica.objects.first()

    context.update({

        "config": config,

        "titulo_relatorio": "RELATÓRIO DE ATENDIMENTOS",

        "now": timezone.now(),

    })

    return render(

        request,

        "accounts/pdf/relatorio_atendimentos_pdf.html",

        context,

    )

# =========================================
# EXCEL - RELATÓRIO DE ATENDIMENTOS
# =========================================

@login_required
def excel_atendimentos(request):

    context = obter_dados_atendimentos(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Atendimentos"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:H1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE ATENDIMENTOS"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:H2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Total de Atendimentos"
    ws["B4"] = context["total_atendimentos"]
    ws["B4"].number_format = "#,##0"

    ws["A5"] = "Confirmados"
    ws["B5"] = context["total_confirmados"]
    ws["B5"].number_format = "#,##0"

    ws["A6"] = "Em Atendimento"
    ws["B6"] = context["total_atendimento"]
    ws["B6"].number_format = "#,##0"

    ws["A7"] = "Finalizados"
    ws["B7"] = context["total_finalizados"]
    ws["B7"].number_format = "#,##0"

    ws["A8"] = "Faltou"
    ws["B8"] = context["total_faltou"]
    ws["B8"].number_format = "#,##0"

    ws["A9"] = "Cancelados"
    ws["B9"] = context["total_cancelados"]
    ws["B9"].number_format = "#,##0"

    for linha in range(4, 10):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Paciente",

        "Profissional",

        "Procedimento",

        "Data",

        "Hora",

        "Status",

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 12

    for coluna, texto in enumerate(
        cabecalho,
        start=1
    ):

        cell = ws.cell(
            row=linha_inicio,
            column=coluna
        )

        cell.value = texto

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for atendimento in context["atendimentos"]:

        # Paciente
        ws.cell(
            linha,
            1
        ).value = atendimento.paciente.nome

        # Profissional
        if atendimento.profissional:

            nome_profissional = getattr(
                atendimento.profissional,
                "nome",
                str(atendimento.profissional)
            )

        else:

            nome_profissional = "-"

        ws.cell(
            linha,
            2
        ).value = nome_profissional

        # Procedimento
        if atendimento.procedimento:

            nome_procedimento = atendimento.procedimento.nome

        else:

            nome_procedimento = "-"

        ws.cell(
            linha,
            3
        ).value = nome_procedimento

        # Data
        celula = ws.cell(
            linha,
            4
        )

        if atendimento.data:

            celula.value = atendimento.data

            celula.number_format = "DD/MM/YYYY"

        # Hora
        celula = ws.cell(
            linha,
            5
        )

        if atendimento.hora_inicio:

            celula.value = atendimento.hora_inicio

            celula.number_format = "HH:MM"

        # Status
        ws.cell(
            linha,
            6
        ).value = atendimento.status

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 35,

        "B": 30,

        "C": 35,

        "D": 15,

        "E": 15,

        "F": 18,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Atendimentos_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response

# =========================================
# RELATÓRIO DE CONTAS A RECEBER
# =========================================

@login_required
def relatorio_contas_receber(request):

    contas = ContaReceber.objects.select_related(
        "paciente",
        "orcamento"
    ).all().order_by(
        "-vencimento"
    )

    # =========================================
    # FILTROS
    # =========================================

    paciente = request.GET.get("paciente", "")
    status = request.GET.get("status", "")
    forma_pagamento = request.GET.get("forma_pagamento", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_final = request.GET.get("data_final", "")

    if paciente:

        contas = contas.filter(
            paciente__nome__icontains=paciente
        )

    if status:

        contas = contas.filter(
            status=status
        )

    # O filtro por forma de pagamento será implementado
    # posteriormente utilizando o relacionamento com Orcamento.

    if data_inicio:

        contas = contas.filter(
            vencimento__gte=data_inicio
        )

    if data_final:

        contas = contas.filter(
            vencimento__lte=data_final
        )

    config = ConfiguracaoClinica.objects.first()

    context = {

        "config": config,

        "contas": contas,

        "paciente": paciente,
        "status": status,
        "forma_pagamento": forma_pagamento,
        "data_inicio": data_inicio,
        "data_final": data_final,

        # =========================================
        # INDICADORES
        # =========================================

        "total_contas": contas.count(),

        "contas_pendentes": contas.filter(
            status="PENDENTE"
        ).count(),

        "contas_recebidas": contas.filter(
            status="RECEBIDO"
        ).count(),

        "contas_canceladas": contas.filter(
            status="CANCELADO"
        ).count(),

        "contas_vencidas": contas.filter(
            status="PENDENTE",
            vencimento__lt=timezone.localdate()
        ).count(),

        "valor_total": sum(
            conta.valor for conta in contas
        ),

        "valor_recebido": sum(
            conta.valor
            for conta in contas.filter(
                status="RECEBIDO"
            )
        ),

        "titulo_relatorio": "RELATÓRIO DE CONTAS A RECEBER",

    }

    return render(

        request,

        "accounts/relatorios/contas_receber.html",

        context

    )

    # =========================================
# PDF - RELATÓRIO DE CONTAS A RECEBER
# =========================================

@login_required
def relatorio_contas_receber_pdf(request):

    contas = ContaReceber.objects.select_related(
        "paciente",
        "orcamento"
    ).all().order_by(
        "-vencimento"
    )

    # ============================
    # FILTROS
    # ============================

    paciente = request.GET.get("paciente", "")
    status = request.GET.get("status", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_final = request.GET.get("data_final", "")

    if paciente:

        contas = contas.filter(
            paciente__nome__icontains=paciente
        )

    if status:

        contas = contas.filter(
            status=status
        )

    if data_inicio:

        contas = contas.filter(
            vencimento__gte=data_inicio
        )

    if data_final:

        contas = contas.filter(
            vencimento__lte=data_final
        )

    config = ConfiguracaoClinica.objects.first()

    context = {

        "config": config,

        "contas": contas,

        "total_contas": contas.count(),

        "contas_pendentes": contas.filter(
            status="PENDENTE"
        ).count(),

        "contas_recebidas": contas.filter(
            status="RECEBIDO"
        ).count(),

        "contas_canceladas": contas.filter(
            status="CANCELADO"
        ).count(),

        "contas_vencidas": contas.filter(
            status="PENDENTE",
            vencimento__lt=timezone.localdate()
        ).count(),

        "valor_total": sum(
            conta.valor for conta in contas
        ),

        "valor_recebido": sum(
            conta.valor
            for conta in contas.filter(
                status="RECEBIDO"
            )
        ),

        "titulo_relatorio": "RELATÓRIO DE CONTAS A RECEBER",

        "now": timezone.now(),

    }

    return render(

        request,

        "accounts/pdf/relatorio_contas_receber_pdf.html",

        context

    )

# =========================================
# RELATÓRIO DE CONTAS A PAGAR
# =========================================

@login_required
def relatorio_contas_pagar(request):

    contas = ContaPagar.objects.select_related(
        "fornecedor"
    ).all().order_by(
        "-vencimento"
    )

    # =========================================
    # FILTROS
    # =========================================

    fornecedor = request.GET.get("fornecedor", "")
    status = request.GET.get("status", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_final = request.GET.get("data_final", "")

    if fornecedor:

        contas = contas.filter(
            fornecedor__nome__icontains=fornecedor
        )

    if status:

        contas = contas.filter(
            status=status
        )

    if data_inicio:

        contas = contas.filter(
            vencimento__gte=data_inicio
        )

    if data_final:

        contas = contas.filter(
            vencimento__lte=data_final
        )

    config = ConfiguracaoClinica.objects.first()

    context = {

        "config": config,

        "contas": contas,

        "fornecedor": fornecedor,
        "status": status,
        "data_inicio": data_inicio,
        "data_final": data_final,

        # =========================================
        # INDICADORES
        # =========================================

        "total_contas": contas.count(),

        "contas_pendentes": contas.filter(
            status="PENDENTE"
        ).count(),

        "contas_pagas": contas.filter(
            status="PAGO"
        ).count(),

        "contas_vencidas": contas.filter(
            status="PENDENTE",
            vencimento__lt=timezone.localdate()
        ).count(),

        "valor_total": sum(
            conta.valor for conta in contas
        ),

        "valor_pago": sum(
            conta.valor
            for conta in contas.filter(
                status="PAGO"
            )
        ),

        "titulo_relatorio": "RELATÓRIO DE CONTAS A PAGAR",

    }

    return render(

        request,

        "accounts/relatorios/contas_pagar.html",

        context

    )

    # =========================================
# PDF - RELATÓRIO DE CONTAS A PAGAR
# =========================================

@login_required
def relatorio_contas_pagar_pdf(request):

    contas = ContaPagar.objects.select_related(
        "fornecedor",
        "compra"
    ).all().order_by(
        "-vencimento"
    )

    # =========================================
    # FILTROS
    # =========================================

    fornecedor = request.GET.get("fornecedor", "")
    status = request.GET.get("status", "")
    categoria = request.GET.get("categoria", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_final = request.GET.get("data_final", "")

    if fornecedor:

        contas = contas.filter(
            fornecedor__nome__icontains=fornecedor
        )

    if status:

        contas = contas.filter(
            status=status.upper()
        )

    if categoria:

        contas = contas.filter(
            descricao__icontains=categoria
        )

    if data_inicio:

        contas = contas.filter(
            vencimento__gte=data_inicio
        )

    if data_final:

        contas = contas.filter(
            vencimento__lte=data_final
        )

    config = ConfiguracaoClinica.objects.first()

    context = {

        "config": config,

        "contas": contas,

        "total_contas": contas.count(),

        "contas_pendentes": contas.filter(
            status="PENDENTE"
        ).count(),

        "contas_pagas": contas.filter(
            status="PAGO"
        ).count(),

        "contas_vencidas": contas.filter(
            status="VENCIDO"
        ).count(),

        "valor_total": sum(
            conta.valor for conta in contas
        ),

        "valor_pago": sum(
            conta.valor
            for conta in contas.filter(
                status="PAGO"
            )
        ),

        "titulo_relatorio": "RELATÓRIO DE CONTAS A PAGAR",

    }

    return render(
        request,
        "accounts/pdf/relatorio_contas_pagar_pdf.html",
        context
    )

# =========================================
# RELATÓRIO DE CAIXA
# =========================================

@login_required
def relatorio_caixa(request):

    movimentos = Caixa.objects.select_related(
        "conta_receber",
        "conta_pagar"
    ).all().order_by(
        "-data",
        "-id"
    )

    # =========================================
    # FILTROS
    # =========================================

    descricao = request.GET.get("descricao", "")
    tipo = request.GET.get("tipo", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_final = request.GET.get("data_final", "")

    if descricao:

        movimentos = movimentos.filter(
            descricao__icontains=descricao
        )

    if tipo:

        movimentos = movimentos.filter(
            tipo=tipo
        )

    if data_inicio:

        movimentos = movimentos.filter(
            data__gte=data_inicio
        )

    if data_final:

        movimentos = movimentos.filter(
            data__lte=data_final
        )

    config = ConfiguracaoClinica.objects.first()

    total_entradas = sum(
        m.valor
        for m in movimentos.filter(
            tipo="ENTRADA"
        )
    )

    total_saidas = sum(
        m.valor
        for m in movimentos.filter(
            tipo="SAIDA"
        )
    )

    context = {

        "config": config,

        "movimentos": movimentos,

        "descricao": descricao,
        "tipo": tipo,
        "data_inicio": data_inicio,
        "data_final": data_final,

        "total_movimentos": movimentos.count(),

        "total_entradas": movimentos.filter(
            tipo="ENTRADA"
        ).count(),

        "total_saidas": movimentos.filter(
            tipo="SAIDA"
        ).count(),

        "valor_entradas": total_entradas,

        "valor_saidas": total_saidas,

        "saldo": total_entradas - total_saidas,

        "titulo_relatorio": "RELATÓRIO DE CAIXA",

    }

    return render(

        request,

        "accounts/relatorios/caixa.html",

        context

    )

    # =========================================
# PDF - RELATÓRIO DE CAIXA
# =========================================

@login_required
def relatorio_caixa_pdf(request):

    movimentos = Caixa.objects.select_related(
        "conta_receber",
        "conta_pagar"
    ).all().order_by(
        "-data",
        "-id"
    )

    # =========================================
    # FILTROS
    # =========================================

    descricao = request.GET.get("descricao", "")
    tipo = request.GET.get("tipo", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_final = request.GET.get("data_final", "")

    if descricao:

        movimentos = movimentos.filter(
            descricao__icontains=descricao
        )

    if tipo:

        movimentos = movimentos.filter(
            tipo=tipo
        )

    if data_inicio:

        movimentos = movimentos.filter(
            data__gte=data_inicio
        )

    if data_final:

        movimentos = movimentos.filter(
            data__lte=data_final
        )

    config = ConfiguracaoClinica.objects.first()

    valor_entradas = sum(
        movimento.valor
        for movimento in movimentos.filter(
            tipo="ENTRADA"
        )
    )

    valor_saidas = sum(
        movimento.valor
        for movimento in movimentos.filter(
            tipo="SAIDA"
        )
    )

    context = {

        "config": config,

        "movimentos": movimentos,

        "total_movimentos": movimentos.count(),

        "total_entradas": movimentos.filter(
            tipo="ENTRADA"
        ).count(),

        "total_saidas": movimentos.filter(
            tipo="SAIDA"
        ).count(),

        "valor_entradas": valor_entradas,

        "valor_saidas": valor_saidas,

        "saldo": valor_entradas - valor_saidas,

        "titulo_relatorio": "RELATÓRIO DE CAIXA",

    }

    return render(

        request,

        "accounts/pdf/relatorio_caixa_pdf.html",

        context

    )

# =========================================
# CENTRAL DE PRODUÇÃO
# =========================================

@login_required
def relatorio_producao(request):

    config = ConfiguracaoClinica.objects.first()

    context = {

        "config": config,

        "titulo_relatorio": "Central de Produção",

    }

    return render(

        request,

        "accounts/relatorios/producao.html",

        context

    )



# =========================================
# DADOS - PRODUÇÃO MENSAL
# =========================================

def obter_dados_producao_mensal(request):

    ano_atual = timezone.now().year

    # =========================================
    # FILTROS
    # =========================================

    ano = request.GET.get(
        "ano",
        str(ano_atual)
    )

    mes = request.GET.get(
        "mes",
        ""
    )

    profissional_id = request.GET.get(
        "profissional",
        ""
    )

    # =========================================
    # LISTA DE ANOS
    # =========================================

    anos = list(
        range(
            ano_atual - 5,
            ano_atual + 1
        )
    )

    # =========================================
    # LISTA DE MESES
    # =========================================

    meses_lista = [

        (1, "Janeiro"),
        (2, "Fevereiro"),
        (3, "Março"),
        (4, "Abril"),
        (5, "Maio"),
        (6, "Junho"),
        (7, "Julho"),
        (8, "Agosto"),
        (9, "Setembro"),
        (10, "Outubro"),
        (11, "Novembro"),
        (12, "Dezembro"),

    ]

    # =========================================
    # PROFISSIONAIS
    # =========================================

    profissionais = (
        User.objects
        .filter(
            is_active=True
        )
        .order_by(
            "first_name",
            "username"
        )
    )

    # =========================================
    # PROCEDIMENTOS REALIZADOS
    # =========================================

    procedimentos = (
        ItemOrcamento.objects
        .filter(
            status="realizado"
        )
        .select_related(
            "procedimento",
            "orcamento",
            "orcamento__paciente",
            "orcamento__paciente__dentista",
        )
    )

    # =========================================
    # FILTRO POR ANO
    # =========================================

    procedimentos = procedimentos.filter(
        orcamento__criado_em__year=int(ano)
    )

    # =========================================
    # FILTRO POR MÊS
    # =========================================

    if mes:

        procedimentos = procedimentos.filter(
            orcamento__criado_em__month=int(mes)
        )

    # =========================================
    # FILTRO POR PROFISSIONAL
    # =========================================

    if profissional_id:

        procedimentos = procedimentos.filter(
            orcamento__paciente__dentista_id=profissional_id
        )

    # =========================================
    # INDICADORES
    # =========================================

    total_procedimentos = procedimentos.count()

    total_producao = sum(
        (item.total for item in procedimentos),
        Decimal("0.00")
    )

    total_pacientes = (
        procedimentos
        .values("orcamento__paciente")
        .distinct()
        .count()
    )

    ticket_medio = (
        total_producao / total_procedimentos
        if total_procedimentos
        else Decimal("0.00")
    )

        # =========================================
    # EVOLUÇÃO MENSAL
    # =========================================

    meses = []

    if mes:
        meses_consulta = [int(mes)]
    else:
        meses_consulta = range(1, 13)

    for numero_mes in meses_consulta:

        dados_mes = procedimentos.filter(
            orcamento__criado_em__month=numero_mes
        )

        quantidade_procedimentos = dados_mes.count()

        quantidade_pacientes = (
            dados_mes
            .values("orcamento__paciente")
            .distinct()
            .count()
        )

        producao_mes = sum(
            (item.total for item in dados_mes),
            Decimal("0.00")
        )

        ticket_mes = (
            producao_mes / quantidade_procedimentos
            if quantidade_procedimentos
            else Decimal("0.00")
        )

        nome_mes = dict(meses_lista)[numero_mes]

        meses.append({

            "numero": numero_mes,

            "nome": nome_mes,

            "procedimentos": quantidade_procedimentos,

            "pacientes": quantidade_pacientes,

            "producao": producao_mes,

            "ticket_medio": ticket_mes,

        })

    # =========================================
    # CONTEXT
    # =========================================

    return {

        "anos": anos,

        "ano_selecionado": int(ano),

        "meses_lista": meses_lista,

        "mes_selecionado": mes,

        "profissionais": profissionais,

        "profissional_id": profissional_id,

        "total_producao": total_producao,

        "total_procedimentos": total_procedimentos,

        "total_pacientes": total_pacientes,

        "ticket_medio": ticket_medio,

        "melhor_mes": "-",

        "meses": meses,

    }

# =========================================
# DADOS - PRODUÇÃO POR PROFISSIONAL
# =========================================

def obter_dados_producao_profissional(request):

    profissional_id = request.GET.get("profissional")
    data_inicio = request.GET.get("data_inicio")
    data_final = request.GET.get("data_final")

    # =========================================
    # TRATA PARÂMETROS
    # =========================================

    if profissional_id in ("", "None", None):
        profissional_id = None

    if data_inicio in ("", "None", None):
        data_inicio = None

    if data_final in ("", "None", None):
        data_final = None

    # =========================================
    # PROFISSIONAIS
    # =========================================

    profissionais = User.objects.filter(
        is_active=True
    ).order_by(
        "first_name",
        "username"
    )

    # =========================================
    # NOME DO PROFISSIONAL
    # =========================================

    nome_profissional = "Todos"

    if profissional_id:

        profissional = User.objects.filter(
            id=profissional_id
        ).first()

        if profissional:

            nome_profissional = (
                profissional.get_full_name()
                or profissional.username
            )

    # =========================================
    # PERÍODO
    # =========================================

    if data_inicio and data_final:

        periodo = f"{data_inicio} até {data_final}"

    elif data_inicio:

        periodo = f"A partir de {data_inicio}"

    elif data_final:

        periodo = f"Até {data_final}"

    else:

        periodo = "Todos"

    # =========================================
    # PROCEDIMENTOS
    # =========================================

    procedimentos = (
        ItemOrcamento.objects
        .filter(status="realizado")
        .select_related(
            "procedimento",
            "orcamento",
            "orcamento__paciente",
            "orcamento__paciente__dentista",
        )
    )

    # =========================================
    # FILTRO PROFISSIONAL
    # =========================================

    if profissional_id:

        procedimentos = procedimentos.filter(
            orcamento__paciente__dentista_id=profissional_id
        )

    # =========================================
    # FILTRO DATA
    # =========================================

    if data_inicio:

        procedimentos = procedimentos.filter(
            orcamento__criado_em__date__gte=data_inicio
        )

    if data_final:

        procedimentos = procedimentos.filter(
            orcamento__criado_em__date__lte=data_final
        )

    # =========================================
    # INDICADORES
    # =========================================

    total_procedimentos = procedimentos.count()

    total_producao = sum(
        (item.total for item in procedimentos),
        Decimal("0.00")
    )

    ticket_medio = (
        total_producao / total_procedimentos
        if total_procedimentos
        else Decimal("0.00")
    )

    total_pacientes = (
        procedimentos
        .values("orcamento__paciente")
        .distinct()
        .count()
    )

    return {

        "profissionais": profissionais,

        "procedimentos": procedimentos,

        "total_producao": total_producao,

        "total_procedimentos": total_procedimentos,

        "ticket_medio": ticket_medio,

        "total_pacientes": total_pacientes,

        "profissional_id": profissional_id,

        "data_inicio": data_inicio,

        "data_final": data_final,

        "nome_profissional": nome_profissional,

        "periodo": periodo,

    }

# =========================================
# RELATÓRIO DE PRODUÇÃO POR PROFISSIONAL
# =========================================

@login_required
def relatorio_producao_profissional(request):

    context = obter_dados_producao_profissional(request)

    return render(

        request,

        "accounts/relatorios/producao_profissional.html",

        context

    )


# =========================================
# PDF - PRODUÇÃO POR PROFISSIONAL
# =========================================

@login_required
def pdf_producao_profissional(request):

    context = obter_dados_producao_profissional(request)

    # =========================================
    # CONFIGURAÇÃO DA CLÍNICA
    # =========================================

    config = ConfiguracaoClinica.objects.first()

    # =========================================
    # DADOS DO CABEÇALHO
    # =========================================

    context.update({

        "config": config,

        "titulo_relatorio": "Relatório de Produção por Profissional",

    })

    return render(

        request,

        "accounts/pdf/producao_profissional_pdf.html",

        context,

    )

    # =========================================
# EXCEL - PRODUÇÃO POR PROFISSIONAL
# =========================================

@login_required
def excel_producao_profissional(request):

    context = obter_dados_producao_profissional(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Produção por Profissional"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:E1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE PRODUÇÃO POR PROFISSIONAL"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:E2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Profissional"
    ws["B4"] = context["nome_profissional"]

    ws["A5"] = "Período"
    ws["B5"] = context["periodo"]

    ws["A6"] = "Produção Total"
    ws["B6"] = float(
        context["total_producao"]
    )
    ws["B6"].number_format = 'R$ #,##0.00'

    ws["A7"] = "Procedimentos"
    ws["B7"] = context["total_procedimentos"]
    ws["B7"].number_format = "#,##0"

    ws["A8"] = "Pacientes"
    ws["B8"] = context["total_pacientes"]
    ws["B8"].number_format = "#,##0"

    ws["A9"] = "Ticket Médio"
    ws["B9"] = float(
        context["ticket_medio"]
    )
    ws["B9"].number_format = 'R$ #,##0.00'

    for linha in range(4, 10):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Profissional",

        "Paciente",

        "Procedimento",

        "Valor",

        "Data",

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 12

    for coluna, texto in enumerate(

        cabecalho,

        start=1

    ):

        cell = ws.cell(

            row=linha_inicio,

            column=coluna

        )

        cell.value = texto

        cell.font = Font(

            bold=True,

            color="FFFFFF"

        )

        cell.fill = fill

        cell.alignment = Alignment(

            horizontal="center"

        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for item in context["procedimentos"]:

        # Profissional
        if item.orcamento.paciente.dentista:

            profissional = (
                item.orcamento.paciente.dentista.get_full_name()
                or item.orcamento.paciente.dentista.username
            )

        else:

            profissional = "-"

        ws.cell(
            linha,
            1
        ).value = profissional

        # Paciente
        ws.cell(
            linha,
            2
        ).value = item.orcamento.paciente.nome

        # Procedimento
        ws.cell(
            linha,
            3
        ).value = item.procedimento.nome

        # Valor
        celula = ws.cell(
            linha,
            4
        )

        celula.value = float(item.total)

        celula.number_format = 'R$ #,##0.00'

        celula.alignment = Alignment(
            horizontal="right"
        )

        # Data
        celula = ws.cell(
            linha,
            5
        )

        if item.orcamento.criado_em:

            celula.value = item.orcamento.criado_em.date()

            celula.number_format = "DD/MM/YYYY"

            celula.alignment = Alignment(
                horizontal="center"
            )

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 30,

        "B": 35,

        "C": 40,

        "D": 18,

        "E": 15,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[coluna].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Producao_Profissional_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response


# =========================================
# DADOS - PRODUÇÃO POR PROCEDIMENTO
# =========================================

def obter_dados_producao_procedimento(request):

    procedimento_id = request.GET.get("procedimento")
    data_inicio = request.GET.get("data_inicio")
    data_final = request.GET.get("data_final")

    # =========================================
    # TRATA PARÂMETROS
    # =========================================

    if procedimento_id in ("", "None", None):
        procedimento_id = None

    if data_inicio in ("", "None", None):
        data_inicio = None

    if data_final in ("", "None", None):
        data_final = None

    # =========================================
    # LISTA DE PROCEDIMENTOS
    # =========================================

    procedimentos = Procedimento.objects.filter(
        ativo=True
    ).order_by("nome")

    # =========================================
    # ITENS REALIZADOS
    # =========================================

    itens = (
        ItemOrcamento.objects
        .filter(status="realizado")
        .select_related(
            "procedimento",
            "orcamento",
            "orcamento__paciente",
            "orcamento__paciente__dentista",
        )
    )

    # =========================================
    # FILTRO POR PROCEDIMENTO
    # =========================================

    if procedimento_id:

        itens = itens.filter(
            procedimento_id=procedimento_id
        )

    # =========================================
    # FILTRO POR DATA
    # =========================================

    if data_inicio:

        itens = itens.filter(
            orcamento__criado_em__date__gte=data_inicio
        )

    if data_final:

        itens = itens.filter(
            orcamento__criado_em__date__lte=data_final
        )

    # =========================================
    # INDICADORES
    # =========================================

    total_procedimentos = itens.count()

    total_producao = sum(
        (item.total for item in itens),
        Decimal("0.00")
    )

    total_pacientes = (
        itens
        .values("orcamento__paciente")
        .distinct()
        .count()
    )

    media_paciente = (
        total_producao / total_pacientes
        if total_pacientes
        else Decimal("0.00")
    )

    # =========================================
    # RANKING POR PROCEDIMENTO
    # =========================================

    ranking = []

    if not procedimento_id:

        agrupados = (
            itens
            .values(
                "procedimento_id",
                "procedimento__nome",
            )
            .annotate(
                quantidade=Sum("quantidade"),
                pacientes=Count(
                    "orcamento__paciente",
                    distinct=True,
                ),
            )
            .order_by(
                "-quantidade",
                "procedimento__nome",
            )
        )

        for grupo in agrupados:

            total = sum(
                item.total
                for item in itens.filter(
                    procedimento_id=grupo["procedimento_id"]
                )
            )

            grupo["total"] = total

            ranking.append(grupo)

    # =========================================
    # PRODUÇÃO POR PROFISSIONAL
    # =========================================

    ranking_profissionais = []

    if procedimento_id:

        profissionais = (
            itens
            .values(
                "orcamento__paciente__dentista_id",
                "orcamento__paciente__dentista__first_name",
                "orcamento__paciente__dentista__last_name",
            )
            .annotate(
                quantidade=Sum("quantidade"),
                pacientes=Count(
                    "orcamento__paciente",
                    distinct=True,
                ),
            )
            .order_by("-quantidade")
        )

        for profissional in profissionais:

            total = sum(
                item.total
                for item in itens.filter(
                    orcamento__paciente__dentista_id=
                    profissional["orcamento__paciente__dentista_id"]
                )
            )

            profissional["total"] = total

            ranking_profissionais.append(profissional)

    # =========================================
    # CONTEXT
    # =========================================

    dados_excel = (
        ranking_profissionais
        if procedimento_id
        else ranking
    )

    modo_excel = (
        "profissionais"
        if procedimento_id
        else "procedimentos"
    )

    return {

        "procedimentos": procedimentos,

        "itens": itens,

        "ranking": ranking,

        "ranking_profissionais": ranking_profissionais,

        "dados_excel": dados_excel,

        "modo_excel": modo_excel,

        "total_producao": total_producao,

        "total_procedimentos": total_procedimentos,

        "total_pacientes": total_pacientes,

        "media_paciente": media_paciente,

        "procedimento_id": procedimento_id,

        "data_inicio": data_inicio,

        "data_final": data_final,

    }

# =========================================
# PRODUÇÃO POR PROCEDIMENTO
# =========================================

@login_required
def relatorio_producao_procedimento(request):

    context = obter_dados_producao_procedimento(request)

    return render(
        request,
        "accounts/relatorios/producao_procedimento.html",
        context,
    )

    # =========================================
# PDF - PRODUÇÃO POR PROCEDIMENTO
# =========================================

@login_required
def pdf_producao_procedimento(request):

    context = obter_dados_producao_procedimento(request)

    config = ConfiguracaoClinica.objects.first()

    context.update({

        "config": config,

        "titulo_relatorio": "Relatório de Produção por Procedimento",

    })

    return render(

        request,

        "accounts/pdf/producao_procedimento_pdf.html",

        context,

    )

# =========================================
# EXCEL - PRODUÇÃO POR PROCEDIMENTO
# =========================================

@login_required
def excel_producao_procedimento(request):

    context = obter_dados_producao_procedimento(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Produção por Procedimento"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:D1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE PRODUÇÃO POR PROCEDIMENTO"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:D2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Produção Total"

    ws["B4"] = float(
        context["total_producao"]
    )

    ws["B4"].number_format = 'R$ #,##0.00'

    ws["A5"] = "Procedimentos"

    ws["B5"] = context["total_procedimentos"]

    ws["B5"].number_format = "#,##0"

    ws["A6"] = "Pacientes"

    ws["B6"] = context["total_pacientes"]

    ws["B6"].number_format = "#,##0"

    ws["A7"] = "Média por Paciente"

    ws["B7"] = float(
        context["media_paciente"]
    )

    ws["B7"].number_format = 'R$ #,##0.00'

    for linha in range(4, 8):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # DEFINIÇÃO DO CABEÇALHO
    # =========================================

    if context["modo_excel"] == "procedimentos":

        cabecalho = [

            "Procedimento",

            "Quantidade",

            "Pacientes",

            "Produção",

        ]

    else:

        cabecalho = [

            "Profissional",

            "Quantidade",

            "Pacientes",

            "Produção",

        ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 10

    for coluna, texto in enumerate(
        cabecalho,
        start=1
    ):

        cell = ws.cell(
            row=linha_inicio,
            column=coluna
        )

        cell.value = texto

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for item in context["dados_excel"]:

        if context["modo_excel"] == "procedimentos":

            # Procedimento
            ws.cell(
                linha,
                1
            ).value = item["procedimento__nome"]

        else:

            nome = (
                f'{item["orcamento__paciente__dentista__first_name"] or ""} '
                f'{item["orcamento__paciente__dentista__last_name"] or ""}'
            ).strip()

            ws.cell(
                linha,
                1
            ).value = nome

        # Quantidade
        celula = ws.cell(
            linha,
            2
        )

        celula.value = item["quantidade"]

        celula.number_format = "#,##0"

        # Pacientes
        celula = ws.cell(
            linha,
            3
        )

        celula.value = item["pacientes"]

        celula.number_format = "#,##0"

        # Produção
        celula = ws.cell(
            linha,
            4
        )

        celula.value = float(item["total"])

        celula.number_format = 'R$ #,##0.00'

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 45,

        "B": 18,

        "C": 18,

        "D": 22,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[coluna].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Producao_Procedimento_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response



# =========================================
# RELATÓRIO DE PRODUÇÃO MENSAL
# =========================================

@login_required
def relatorio_producao_mensal(request):

    context = obter_dados_producao_mensal(request)

    return render(

        request,

        "accounts/relatorios/producao_mensal.html",

        context,

    )

# =========================================
# PDF - PRODUÇÃO MENSAL
# =========================================

@login_required
def pdf_producao_mensal(request):

    context = obter_dados_producao_mensal(request)

    config = ConfiguracaoClinica.objects.first()

    context.update({

        "config": config,

        "titulo_relatorio": "Relatório de Produção Mensal",

    })

    return render(

        request,

        "accounts/pdf/producao_mensal_pdf.html",

        context,

    )


# =========================================
# DADOS - PRODUÇÃO POR CONVÊNIO
# =========================================

def obter_dados_producao_convenio(request):

    from decimal import Decimal
    from django.db.models import Count

    # =========================================
    # FILTROS
    # =========================================

    convenio = request.GET.get(
        "convenio",
        ""
    )

    data_inicio = request.GET.get(
        "data_inicio",
        ""
    )

    data_final = request.GET.get(
        "data_final",
        ""
    )

    # =========================================
    # LISTA DE CONVÊNIOS
    # =========================================

    convenios = (
        Paciente.objects
        .exclude(convenio__isnull=True)
        .exclude(convenio="")
        .values_list(
            "convenio",
            flat=True
        )
        .distinct()
        .order_by("convenio")
    )

    # =========================================
    # PROCEDIMENTOS REALIZADOS
    # =========================================

    procedimentos = (
        ItemOrcamento.objects
        .filter(
            status="realizado"
        )
        .select_related(
            "orcamento",
            "orcamento__paciente"
        )
    )

    # =========================================
    # FILTRO CONVÊNIO
    # =========================================

    if convenio:

        procedimentos = procedimentos.filter(
            orcamento__paciente__convenio=convenio
        )

    # =========================================
    # FILTRO DATA INICIAL
    # =========================================

    if data_inicio:

        procedimentos = procedimentos.filter(
            orcamento__criado_em__date__gte=data_inicio
        )

    # =========================================
    # FILTRO DATA FINAL
    # =========================================

    if data_final:

        procedimentos = procedimentos.filter(
            orcamento__criado_em__date__lte=data_final
        )

    # =========================================
    # INDICADORES
    # =========================================

    total_procedimentos = procedimentos.count()

    total_producao = sum(

        (item.total for item in procedimentos),

        Decimal("0.00")

    )

    total_pacientes = (

        procedimentos

        .values(
            "orcamento__paciente"
        )

        .distinct()

        .count()

    )

    total_convenios = (

        procedimentos

        .values(
            "orcamento__paciente__convenio"
        )

        .exclude(
            orcamento__paciente__convenio=""
        )

        .distinct()

        .count()

    )

    ticket_medio = (

        total_producao / total_procedimentos

        if total_procedimentos

        else Decimal("0.00")

    )

    # =========================================
    # TABELA
    # =========================================

    resumo = []

    for nome_convenio in convenios:

        itens = procedimentos.filter(
            orcamento__paciente__convenio=nome_convenio
        )

        producao = sum(

            (item.total for item in itens),

            Decimal("0.00")

        )

        qtd_proc = itens.count()

        qtd_pacientes = (

            itens

            .values(
                "orcamento__paciente"
            )

            .distinct()

            .count()

        )

        ticket = (

            producao / qtd_proc

            if qtd_proc

            else Decimal("0.00")

        )

        resumo.append({

            "nome": nome_convenio,

            "procedimentos": qtd_proc,

            "pacientes": qtd_pacientes,

            "producao": producao,

            "ticket_medio": ticket,

        })

    # =========================================
    # CONTEXT
    # =========================================

    return {

        "convenios": convenios,

        "convenio_id": convenio,

        "data_inicio": data_inicio,

        "data_final": data_final,

        "total_producao": total_producao,

        "total_procedimentos": total_procedimentos,

        "total_pacientes": total_pacientes,

        "total_convenios": total_convenios,

        "ticket_medio": ticket_medio,

        "convenios_resumo": resumo,

    }


# =========================================
# RELATÓRIO PRODUÇÃO POR CONVÊNIO
# =========================================

@login_required
def relatorio_producao_convenio(request):

    context = obter_dados_producao_convenio(request)

    return render(

        request,

        "accounts/relatorios/producao_convenio.html",

        context,

    )


# =========================================
# PDF - PRODUÇÃO POR CONVÊNIO
# =========================================

@login_required
def pdf_producao_convenio(request):

    context = obter_dados_producao_convenio(request)

    config = ConfiguracaoClinica.objects.first()

    context.update({

        "config": config,

        "titulo_relatorio": "Relatório de Produção por Convênio",

    })

    return render(

        request,

        "accounts/pdf/producao_convenio_pdf.html",

        context,

    )

    # =========================================
# EXCEL - PRODUÇÃO POR CONVÊNIO
# =========================================

@login_required
def excel_producao_convenio(request):

    context = obter_dados_producao_convenio(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Produção por Convênio"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:E1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE PRODUÇÃO POR CONVÊNIO"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:E2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Convênio"
    ws["B4"] = context["convenio_id"] or "Todos"

    ws["A5"] = "Produção Total"
    ws["B5"] = float(
        context["total_producao"]
    )
    ws["B5"].number_format = 'R$ #,##0.00'

    ws["A6"] = "Procedimentos"
    ws["B6"] = context["total_procedimentos"]
    ws["B6"].number_format = "#,##0"

    ws["A7"] = "Pacientes"
    ws["B7"] = context["total_pacientes"]
    ws["B7"].number_format = "#,##0"

    ws["A8"] = "Convênios"
    ws["B8"] = context["total_convenios"]
    ws["B8"].number_format = "#,##0"

    ws["A9"] = "Ticket Médio"
    ws["B9"] = float(
        context["ticket_medio"]
    )
    ws["B9"].number_format = 'R$ #,##0.00'

    for linha in range(4, 10):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Convênio",

        "Procedimentos",

        "Pacientes",

        "Produção",

        "Ticket Médio",

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 12

    for coluna, texto in enumerate(

        cabecalho,

        start=1

    ):

        cell = ws.cell(

            row=linha_inicio,

            column=coluna

        )

        cell.value = texto

        cell.font = Font(

            bold=True,

            color="FFFFFF"

        )

        cell.fill = fill

        cell.alignment = Alignment(

            horizontal="center"

        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for convenio in context["convenios_resumo"]:

        # Convênio
        ws.cell(
            linha,
            1
        ).value = convenio["nome"]

        # Procedimentos
        celula = ws.cell(
            linha,
            2
        )

        celula.value = convenio["procedimentos"]

        celula.number_format = "#,##0"

        celula.alignment = Alignment(
            horizontal="center"
        )

        # Pacientes
        celula = ws.cell(
            linha,
            3
        )

        celula.value = convenio["pacientes"]

        celula.number_format = "#,##0"

        celula.alignment = Alignment(
            horizontal="center"
        )

        # Produção
        celula = ws.cell(
            linha,
            4
        )

        celula.value = float(
            convenio["producao"]
        )

        celula.number_format = 'R$ #,##0.00'

        celula.alignment = Alignment(
            horizontal="right"
        )

        # Ticket Médio
        celula = ws.cell(
            linha,
            5
        )

        celula.value = float(
            convenio["ticket_medio"]
        )

        celula.number_format = 'R$ #,##0.00'

        celula.alignment = Alignment(
            horizontal="right"
        )

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 35,

        "B": 18,

        "C": 18,

        "D": 20,

        "E": 20,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Producao_Convenio_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response

# =========================================
# DADOS - FATURAMENTO
# =========================================

def obter_dados_faturamento(request):

    # =========================================
    # FILTROS
    # =========================================

    data_inicio = request.GET.get(
        "data_inicio",
        ""
    )

    data_final = request.GET.get(
        "data_final",
        ""
    )

    forma_pagamento = request.GET.get(
        "forma_pagamento",
        ""
    )

    # =========================================
    # ORÇAMENTOS
    # =========================================

    orcamentos = Orcamento.objects.all()

    if data_inicio:

        orcamentos = orcamentos.filter(
            criado_em__date__gte=data_inicio
        )

    if data_final:

        orcamentos = orcamentos.filter(
            criado_em__date__lte=data_final
        )

    if forma_pagamento:

        orcamentos = orcamentos.filter(
            forma_pagamento=forma_pagamento
        )

    # =========================================
    # QUANTIDADE
    # =========================================

    quantidade_orcamentos = orcamentos.count()

    # =========================================
    # TOTAIS
    # =========================================

    totais = orcamentos.aggregate(

        total_descontos=Sum("desconto"),

        total_entrada=Sum("entrada"),

        total_parcelas=Sum("parcelas"),

    )

    total_descontos = totais["total_descontos"] or Decimal("0.00")

    total_entrada = totais["total_entrada"] or Decimal("0.00")

    total_parcelas = totais["total_parcelas"] or 0


    faturamento_bruto = Decimal("0.00")

    maior_orcamento = Decimal("0.00")

    menor_orcamento = None


    for orcamento in orcamentos:

        valor = orcamento.total

        faturamento_bruto += valor

        if valor > maior_orcamento:

            maior_orcamento = valor

        if menor_orcamento is None or valor < menor_orcamento:

            menor_orcamento = valor


    if menor_orcamento is None:

        menor_orcamento = Decimal("0.00")

    # =========================================
    # FATURAMENTO LÍQUIDO
    # =========================================

    faturamento_liquido = (
        faturamento_bruto - total_descontos
    )

    # =========================================
    # TICKET MÉDIO
    # =========================================

    if quantidade_orcamentos:

        ticket_medio = (
            faturamento_bruto /
            quantidade_orcamentos
        )

        media_desconto = (
            total_descontos /
            quantidade_orcamentos
        )

    else:

        ticket_medio = Decimal("0.00")

        media_desconto = Decimal("0.00")

    # =========================================
    # PERCENTUAL DESCONTO
    # =========================================

    percentual_desconto = Decimal("0.00")

    if faturamento_bruto > 0:

        percentual_desconto = (
            total_descontos * 100
        ) / faturamento_bruto

    # =========================================
    # FATURAMENTO POR FORMA PAGAMENTO
    # =========================================

    formas = [

        ("dinheiro", "Dinheiro"),

        ("pix", "PIX"),

        ("debito", "Cartão Débito"),

        ("credito", "Cartão Crédito"),

        ("boleto", "Boleto"),

    ]

    faturamento = []

    forma_mais_utilizada = ""

    maior_quantidade = 0

    for codigo, descricao in formas:

        lista = orcamentos.filter(
            forma_pagamento=codigo
        )

        quantidade = lista.count()

        total = sum(
            (orcamento.total for orcamento in lista),
            Decimal("0.00")
        )

        percentual = Decimal("0.00")

        if faturamento_bruto > 0:

            percentual = (
                total * 100
            ) / faturamento_bruto

        faturamento.append({

            "forma": descricao,

            "quantidade": quantidade,

            "total": total,

            "percentual": percentual,

        })

        if quantidade > maior_quantidade:

            maior_quantidade = quantidade

            forma_mais_utilizada = descricao

    # =========================================
    # CONTEXT
    # =========================================

    return {

        "data_inicio": data_inicio,

        "data_final": data_final,

        "forma_pagamento": forma_pagamento,

        "quantidade_orcamentos": quantidade_orcamentos,

        "faturamento_bruto": faturamento_bruto,

        "faturamento_liquido": faturamento_liquido,

        "total_descontos": total_descontos,

        "total_entrada": total_entrada,

        "total_parcelas": total_parcelas,

        "ticket_medio": ticket_medio,

        "media_desconto": media_desconto,

        "percentual_desconto": percentual_desconto,

        "maior_orcamento": maior_orcamento,

        "menor_orcamento": menor_orcamento,

        "forma_mais_utilizada": forma_mais_utilizada,

        "faturamento": faturamento,

    }

# =========================================
# RELATÓRIO FATURAMENTO
# =========================================

@login_required
def relatorio_faturamento(request):

    context = obter_dados_faturamento(request)

    return render(

        request,

        "accounts/relatorios/faturamento.html",

        context,

    )


# =========================================
# PDF - FATURAMENTO
# =========================================

@login_required
def pdf_faturamento(request):

    context = obter_dados_faturamento(request)

    config = ConfiguracaoClinica.objects.first()

    context.update({

        "config": config,

        "titulo_relatorio": "Relatório de Faturamento",

    })

    return render(

        request,

        "accounts/pdf/faturamento_pdf.html",

        context,

    )

# =========================================
# EXCEL - FATURAMENTO
# =========================================

@login_required
def excel_faturamento(request):

    context = obter_dados_faturamento(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Faturamento"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:D1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE FATURAMENTO"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:D2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO FINANCEIRO
    # =========================================

    ws["A4"] = "Faturamento Bruto"
    ws["B4"] = float(context["faturamento_bruto"])
    ws["B4"].number_format = 'R$ #,##0.00'

    ws["A5"] = "Faturamento Líquido"
    ws["B5"] = float(context["faturamento_liquido"])
    ws["B5"].number_format = 'R$ #,##0.00'

    ws["A6"] = "Descontos"
    ws["B6"] = float(context["total_descontos"])
    ws["B6"].number_format = 'R$ #,##0.00'

    ws["A7"] = "Entradas"
    ws["B7"] = float(context["total_entrada"])
    ws["B7"].number_format = 'R$ #,##0.00'

    ws["A8"] = "Ticket Médio"
    ws["B8"] = float(context["ticket_medio"])
    ws["B8"].number_format = 'R$ #,##0.00'

    ws["A9"] = "Qtd. Orçamentos"
    ws["B9"] = context["quantidade_orcamentos"]
    ws["B9"].number_format = "#,##0"

    for linha in range(4, 10):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Forma de Pagamento",

        "Quantidade",

        "Total",

        "% do Faturamento",

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 12

    for coluna, texto in enumerate(
        cabecalho,
        start=1
    ):

        cell = ws.cell(
            row=linha_inicio,
            column=coluna
        )

        cell.value = texto

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for item in context["faturamento"]:

        # Forma de pagamento
        ws.cell(
            linha,
            1
        ).value = item["forma"]

        # Quantidade
        celula = ws.cell(
            linha,
            2
        )

        celula.value = item["quantidade"]

        celula.number_format = "#,##0"

        # Total
        celula = ws.cell(
            linha,
            3
        )

        celula.value = float(
            item["total"]
        )

        celula.number_format = 'R$ #,##0.00'

        # Percentual
        celula = ws.cell(
            linha,
            4
        )

        celula.value = float(
            item["percentual"]
        )

        celula.number_format = '0.00'

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 35,

        "B": 18,

        "C": 22,

        "D": 18,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Faturamento_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response


# =========================================
# DADOS - AUDITORIA
# =========================================

def obter_dados_auditoria(request):

    # =========================================
    # FILTROS
    # =========================================

    data_inicio = request.GET.get(
        "data_inicio",
        ""
    )

    data_final = request.GET.get(
        "data_final",
        ""
    )

    usuario = request.GET.get(
        "usuario",
        ""
    )

    modulo = request.GET.get(
        "modulo",
        ""
    )

    acao = request.GET.get(
        "acao",
        ""
    )

    nivel = request.GET.get(
        "nivel",
        ""
    )

    # =========================================
    # CONSULTA
    # =========================================

    auditorias = (
        Auditoria.objects
        .select_related("usuario")
        .order_by("-data_hora")
    )

    # =========================================
    # FILTRO DATA INICIAL
    # =========================================

    if data_inicio:

        auditorias = auditorias.filter(
            data_hora__date__gte=data_inicio
        )

    # =========================================
    # FILTRO DATA FINAL
    # =========================================

    if data_final:

        auditorias = auditorias.filter(
            data_hora__date__lte=data_final
        )

    # =========================================
    # FILTRO USUÁRIO
    # =========================================

    if usuario:

        auditorias = auditorias.filter(
            usuario_id=usuario
        )

    # =========================================
    # FILTRO MÓDULO
    # =========================================

    if modulo:

        auditorias = auditorias.filter(
            modulo=modulo
        )

    # =========================================
    # FILTRO AÇÃO
    # =========================================

    if acao:

        auditorias = auditorias.filter(
            acao=acao
        )

    # =========================================
    # FILTRO NÍVEL
    # =========================================

    if nivel:

        auditorias = auditorias.filter(
            nivel=nivel
        )

    # =========================================
    # USUÁRIOS
    # =========================================

    usuarios = User.objects.filter(
        is_active=True
    ).order_by(
        "first_name",
        "username"
    )

    # =========================================
    # INDICADORES
    # =========================================

    total_eventos = auditorias.count()

    total_logins = auditorias.filter(
        acao="login"
    ).count()

    total_logouts = auditorias.filter(
        acao="logout"
    ).count()

    total_erros = auditorias.filter(
        nivel="erro"
    ).count()

    total_criticos = auditorias.filter(
        nivel="critico"
    ).count()

    # =========================================
    # LISTAS DOS FILTROS
    # =========================================

    modulos = (
        Auditoria.objects
        .values_list(
            "modulo",
            flat=True
        )
        .distinct()
        .order_by("modulo")
    )

    acoes = (
        Auditoria.objects
        .values_list(
            "acao",
            flat=True
        )
        .distinct()
        .order_by("acao")
    )

    niveis = (
        Auditoria.objects
        .values_list(
            "nivel",
            flat=True
        )
        .distinct()
        .order_by("nivel")
    )

    # =========================================
    # CONTEXT
    # =========================================

    context = {

        "auditorias": auditorias,

        "usuarios": usuarios,

        "modulos": modulos,

        "acoes": acoes,

        "niveis": niveis,

        "data_inicio": data_inicio,

        "data_final": data_final,

        "usuario": usuario,

        "modulo": modulo,

        "acao": acao,

        "nivel": nivel,

        "total_eventos": total_eventos,

        "total_logins": total_logins,

        "total_logouts": total_logouts,

        "total_erros": total_erros,

        "total_criticos": total_criticos,

    }

    return context


# =========================================
# RELATÓRIO - AUDITORIA
# =========================================

@login_required
def relatorio_auditoria(request):

    context = obter_dados_auditoria(request)

    return render(

        request,

        "accounts/relatorios/auditoria.html",

        context,

    )

# =========================================
# PDF - AUDITORIA
# =========================================

@login_required
def pdf_auditoria(request):

    context = obter_dados_auditoria(request)

    # =========================================
    # NOME DO USUÁRIO
    # =========================================

    usuario_id = request.GET.get("usuario")

    if usuario_id:

        try:

            usuario = User.objects.get(id=usuario_id)

            context["usuario_nome"] = (
                usuario.get_full_name() or usuario.username
            )

        except User.DoesNotExist:

            context["usuario_nome"] = "Todos"

    else:

        context["usuario_nome"] = "Todos"

    # =========================================
    # AÇÃO (NOME AMIGÁVEL)
    # =========================================

    acoes = {
        "login": "Login",
        "logout": "Logout",
        "cadastro": "Cadastro",
        "alteracao": "Alteração",
        "exclusao": "Exclusão",
        "impressao": "Impressão",
        "exportacao": "Exportação",
    }

    context["acao"] = acoes.get(
        request.GET.get("acao"),
        "Todas"
    )

    # =========================================
    # NÍVEL (NOME AMIGÁVEL)
    # =========================================

    niveis = {
        "info": "Informação",
        "aviso": "Aviso",
        "erro": "Erro",
        "critico": "Crítico",
    }

    context["nivel"] = niveis.get(
        request.GET.get("nivel"),
        "Todos"
    )

    # =========================================
    # MÓDULO
    # =========================================

    context["modulo"] = (
        request.GET.get("modulo")
        or "Todos"
    )

    return render(

        request,

        "accounts/pdf/auditoria_pdf.html",

        context,

    )
# =========================================
# EXCEL - AUDITORIA
# =========================================

@login_required
def excel_auditoria(request):

    context = obter_dados_auditoria(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Auditoria"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:G1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE AUDITORIA DO SISTEMA"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:G2")

    data = ws["A2"]

    data.value = (

        "Gerado em: "

        + timezone.localtime().strftime("%d/%m/%Y %H:%M")

    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Total de Eventos"
    ws["B4"] = context["total_eventos"]
    ws["B4"].number_format = "#,##0"

    ws["A5"] = "Logins"
    ws["B5"] = context["total_logins"]
    ws["B5"].number_format = "#,##0"

    ws["A6"] = "Logouts"
    ws["B6"] = context["total_logouts"]
    ws["B6"].number_format = "#,##0"

    ws["A7"] = "Erros"
    ws["B7"] = context["total_erros"]
    ws["B7"].number_format = "#,##0"

    ws["A8"] = "Críticos"
    ws["B8"] = context["total_criticos"]
    ws["B8"].number_format = "#,##0"

    for linha in range(4, 9):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Data/Hora",

        "Usuário",

        "Módulo",

        "Ação",

        "Descrição",

        "IP",

        "Nível",

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 11

    for coluna, texto in enumerate(

        cabecalho,

        start=1

    ):

        cell = ws.cell(

            row=linha_inicio,

            column=coluna

        )

        cell.value = texto

        cell.font = Font(

            bold=True,

            color="FFFFFF"

        )

        cell.fill = fill

        cell.alignment = Alignment(

            horizontal="center"

        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for auditoria in context["auditorias"]:

       # =========================================
        # DATA / HORA
        # =========================================

        celula = ws.cell(
            linha,
            1
        )

        # O Excel não aceita datetime com timezone.
        # Converte para o horário local e remove o tzinfo.

        data_hora = timezone.localtime(
            auditoria.data_hora
        ).replace(
            tzinfo=None
        )

        celula.value = data_hora

        celula.number_format = "DD/MM/YYYY HH:MM"

        celula.alignment = Alignment(
            horizontal="center"
        )

        # Usuário
        if auditoria.usuario:

            usuario = (
                auditoria.usuario.get_full_name()
                or auditoria.usuario.username
            )

        else:

            usuario = "-"

        ws.cell(
            linha,
            2
        ).value = usuario

        # Módulo
        ws.cell(
            linha,
            3
        ).value = auditoria.modulo

        # Ação
        ws.cell(
            linha,
            4
        ).value = auditoria.acao

        # Descrição
        ws.cell(
            linha,
            5
        ).value = auditoria.descricao

        # IP
        ws.cell(
            linha,
            6
        ).value = auditoria.ip

        # Nível
        celula = ws.cell(
            linha,
            7
        )

        celula.value = auditoria.nivel

        celula.alignment = Alignment(
            horizontal="center"
        )

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 22,

        "B": 30,

        "C": 22,

        "D": 18,

        "E": 55,

        "F": 18,

        "G": 15,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Auditoria_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response

# =========================================
# EXCEL RELATÓRIO DE PACIENTES
# =========================================

@login_required(login_url='/')
def exportar_excel_relatorio_pacientes(request):

    pacientes = Paciente.objects.all()

    # =========================================
    # FILTRO POR NOME
    # =========================================

    q = request.GET.get('q', '').strip()

    if q:

        pacientes = pacientes.filter(
            nome__icontains=q
        )

    # =========================================
    # FILTRO POR CONVÊNIO
    # =========================================

    convenio = request.GET.get('convenio', '').strip()

    if convenio:

        try:

            convenio_obj = Convenio.objects.get(
                id=convenio
            )

            pacientes = pacientes.filter(
                convenio=convenio_obj.nome
            )

        except Convenio.DoesNotExist:

            pass

    # =========================================
    # FILTRO POR STATUS
    # =========================================

    status = request.GET.get('status', '').strip()

    if status == "1":

        pacientes = pacientes.filter(
            ativo=True
        )

    elif status == "0":

        pacientes = pacientes.filter(
            ativo=False
        )

    pacientes = pacientes.order_by("nome")

    # =========================================
    # CRIA PLANILHA
    # =========================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Pacientes"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:G1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE PACIENTES"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(horizontal="center")

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:G2")

    data = ws["A2"]

    data.value = (
        f"Gerado em: "
        f"{timezone.localtime().strftime('%d/%m/%Y %H:%M')}"
    )

    data.alignment = Alignment(horizontal="center")

    # =========================================
    # RESUMO
    # =========================================

    total = pacientes.count()

    ativos = pacientes.filter(
        ativo=True
    ).count()

    inativos = pacientes.filter(
        ativo=False
    ).count()

    ws["A3"] = f"Total: {total}"
    ws["C3"] = f"Ativos: {ativos}"
    ws["E3"] = f"Inativos: {inativos}"

    for celula in ("A3", "C3", "E3"):
        ws[celula].font = Font(bold=True)



    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Nome",
        "CPF",
        "Telefone",
        "E-mail",
        "Convênio",
        "Nascimento",
        "Status"

    ]

    fill = PatternFill(
        start_color="0D6EFD",
        end_color="0D6EFD",
        fill_type="solid"
    )

    for coluna, texto in enumerate(cabecalho, start=1):

        cell = ws.cell(
            row=4,
            column=coluna
        )

        cell.value = texto

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    # =========================================
    # DADOS
    # =========================================

    linha = 5

    for paciente in pacientes:

        ws.cell(
            linha,
            1
        ).value = paciente.nome

        ws.cell(
            linha,
            2
        ).value = paciente.cpf

        ws.cell(
            linha,
            3
        ).value = paciente.telefone

        ws.cell(
            linha,
            4
        ).value = paciente.email

        ws.cell(
            linha,
            5
        ).value = paciente.convenio

        ws.cell(
            linha,
            6
        ).value = (
            paciente.nascimento.strftime("%d/%m/%Y")
            if paciente.nascimento
            else ""
        )

        ws.cell(
            linha,
            7
        ).value = (
            "Ativo"
            if paciente.ativo
            else "Inativo"
        )

        linha += 1

    # =========================================
    # AJUSTA LARGURA
    # =========================================

    larguras = {

        "A": 40,
        "B": 18,
        "C": 18,
        "D": 35,
        "E": 25,
        "F": 18,
        "G": 15,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[coluna].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response[
        "Content-Disposition"
    ] = (

        f'attachment; filename="Relatorio_Pacientes_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response


# =========================================
# EXCEL - PRODUÇÃO MENSAL
# =========================================

@login_required
def excel_producao_mensal(request):

    context = obter_dados_producao_mensal(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Produção Mensal"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:E1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE PRODUÇÃO MENSAL"

    titulo.font = Font(
        size=16,
        bold=True
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:E2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Produção Total"

    ws["B4"] = float(
        context["total_producao"]
    )
    ws["B4"].number_format = 'R$ #,##0.00'

    ws["A5"] = "Procedimentos"

    ws["B5"] = context["total_procedimentos"]

    ws["A6"] = "Pacientes"

    ws["B6"] = context["total_pacientes"]

    ws["A7"] = "Ticket Médio"

    ws["B7"] = float(
        context["ticket_medio"]
    )
    ws["B7"].number_format = 'R$ #,##0.00'

    for linha in range(4, 8):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Mês",

        "Procedimentos",

        "Pacientes",

        "Produção",

        "Ticket Médio",

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 10

    for coluna, texto in enumerate(
        cabecalho,
        start=1
    ):

        cell = ws.cell(
            row=linha_inicio,
            column=coluna
        )

        cell.value = texto

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    # =========================================
    # DADOS
    # =========================================

    linha = linha_inicio + 1

    for mes in context["meses"]:

        # Mês
        ws.cell(
            linha,
            1
        ).value = mes["nome"]

        # Procedimentos
        celula = ws.cell(
            linha,
            2
        )

        celula.value = mes["procedimentos"]

        celula.number_format = "#,##0"

        # Pacientes
        celula = ws.cell(
            linha,
            3
        )

        celula.value = mes["pacientes"]

        celula.number_format = "#,##0"

        # Produção
        celula = ws.cell(
            linha,
            4
        )

        celula.value = float(
            mes["producao"]
        )

        celula.number_format = 'R$ #,##0.00'

        # Ticket Médio
        celula = ws.cell(
            linha,
            5
        )

        celula.value = float(
            mes["ticket_medio"]
        )

        celula.number_format = 'R$ #,##0.00'

        linha += 1
    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 22,

        "B": 18,

        "C": 18,

        "D": 20,

        "E": 20,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Producao_Mensal_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response

# =========================================
# DADOS - TRATAMENTOS
# =========================================

def obter_dados_tratamentos(request):

    # =========================================
    # FILTROS
    # =========================================

    paciente = request.GET.get(
        "paciente",
        ""
    )

    status = request.GET.get(
        "status",
        ""
    )

    data_inicio = request.GET.get(
        "data_inicio",
        ""
    )

    data_final = request.GET.get(
        "data_final",
        ""
    )

    # =========================================
    # QUERY
    # =========================================

    tratamentos = (
        Tratamento.objects
        .select_related("paciente")
        .all()
    )

    # =========================================
    # FILTRO PACIENTE
    # =========================================

    if paciente:

        tratamentos = tratamentos.filter(
            paciente__nome__icontains=paciente
        )

    # =========================================
    # FILTRO STATUS
    # =========================================

    if status:

        tratamentos = tratamentos.filter(
            status=status
        )

    # =========================================
    # FILTRO DATA
    # =========================================

    if data_inicio:

        tratamentos = tratamentos.filter(
            data_inicio__gte=data_inicio
        )

    if data_final:

        tratamentos = tratamentos.filter(
            data_inicio__lte=data_final
        )

    # =========================================
    # ORDENAÇÃO
    # =========================================

    tratamentos = tratamentos.order_by(
        "-data_inicio",
        "paciente__nome",
    )

    # =========================================
    # INDICADORES
    # =========================================

    total_tratamentos = tratamentos.count()

    tratamentos_ativos = tratamentos.filter(
        status="ATIVO"
    ).count()

    tratamentos_encerrados = tratamentos.filter(
        status="ENCERRADO"
    ).count()

    pacientes_em_tratamento = (
        tratamentos
        .filter(status="ATIVO")
        .values("paciente")
        .distinct()
        .count()
    )

    # =========================================
    # CONTEXT
    # =========================================

    return {

        "tratamentos": tratamentos,

        "paciente": paciente,

        "status": status,

        "data_inicio": data_inicio,

        "data_final": data_final,

        "total_tratamentos": total_tratamentos,

        "tratamentos_ativos": tratamentos_ativos,

        "tratamentos_encerrados": tratamentos_encerrados,

        "pacientes_em_tratamento": pacientes_em_tratamento,

    }

# =========================================
# RELATÓRIO DE TRATAMENTOS
# =========================================

@login_required
def relatorio_tratamentos(request):

    context = obter_dados_tratamentos(request)

    context.update({

        "config": ConfiguracaoClinica.objects.first(),

        "titulo_relatorio": "RELATÓRIO DE TRATAMENTOS",

    })

    return render(

        request,

        "accounts/relatorios/tratamentos.html",

        context,

    )

# =========================================
# PDF - RELATÓRIO DE TRATAMENTOS
# =========================================

@login_required
def pdf_relatorio_tratamentos(request):

    context = obter_dados_tratamentos(request)

    config = ConfiguracaoClinica.objects.first()

    context.update({

        "config": config,

        "titulo_relatorio": "RELATÓRIO DE TRATAMENTOS",

        "now": timezone.now(),

    })

    return render(

        request,

        "accounts/pdf/tratamentos_pdf.html",

        context,

    )

    # =========================================
# EXCEL - RELATÓRIO DE TRATAMENTOS
# =========================================

@login_required
def excel_tratamentos(request):

    context = obter_dados_tratamentos(request)

    wb = Workbook()

    ws = wb.active

    ws.title = "Tratamentos"

    # =========================================
    # TÍTULO
    # =========================================

    ws.merge_cells("A1:F1")

    titulo = ws["A1"]

    titulo.value = "RELATÓRIO DE TRATAMENTOS"

    titulo.font = Font(
        bold=True,
        size=16
    )

    titulo.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # DATA
    # =========================================

    ws.merge_cells("A2:F2")

    data = ws["A2"]

    data.value = (
        "Gerado em: "
        + timezone.localtime().strftime("%d/%m/%Y %H:%M")
    )

    data.alignment = Alignment(
        horizontal="center"
    )

    # =========================================
    # RESUMO
    # =========================================

    ws["A4"] = "Total de Tratamentos"

    ws["B4"] = context["total_tratamentos"]

    ws["B4"].number_format = "#,##0"

    ws["A5"] = "Tratamentos Ativos"

    ws["B5"] = context["tratamentos_ativos"]

    ws["B5"].number_format = "#,##0"

    ws["A6"] = "Tratamentos Encerrados"

    ws["B6"] = context["tratamentos_encerrados"]

    ws["B6"].number_format = "#,##0"

    ws["A7"] = "Pacientes em Tratamento"

    ws["B7"] = context["pacientes_em_tratamento"]

    ws["B7"].number_format = "#,##0"

    for linha in range(4, 8):

        ws[f"A{linha}"].font = Font(
            bold=True
        )

    # =========================================
    # CABEÇALHO
    # =========================================

    cabecalho = [

        "Paciente",

        "Tratamento",

        "Data Início",

        "Data Encerramento",

        "Status",

        "Dias"

    ]

    fill = PatternFill(

        start_color="0D6EFD",

        end_color="0D6EFD",

        fill_type="solid",

    )

    linha_inicio = 10

    for coluna, texto in enumerate(

        cabecalho,

        start=1

    ):

        cell = ws.cell(

            row=linha_inicio,

            column=coluna

        )

        cell.value = texto

        cell.font = Font(

            bold=True,

            color="FFFFFF"

        )

        cell.fill = fill

        cell.alignment = Alignment(

            horizontal="center"

        )

    linha = linha_inicio + 1

        # =========================================
    # DADOS
    # =========================================

    for tratamento in context["tratamentos"]:

        # Paciente
        ws.cell(
            linha,
            1
        ).value = tratamento.paciente.nome

        # Tratamento
        ws.cell(
            linha,
            2
        ).value = tratamento.titulo

        # Data de Início
        celula = ws.cell(
            linha,
            3
        )

        if tratamento.data_inicio:

            celula.value = tratamento.data_inicio

            celula.number_format = "DD/MM/YYYY"

        # Data de Encerramento
        celula = ws.cell(
            linha,
            4
        )

        if tratamento.data_encerramento:

            celula.value = tratamento.data_encerramento

            celula.number_format = "DD/MM/YYYY"

        # Status
        ws.cell(
            linha,
            5
        ).value = tratamento.status

        # Dias de Tratamento
        celula = ws.cell(
            linha,
            6
        )

        if tratamento.data_inicio:

            if tratamento.data_encerramento:

                dias = (
                    tratamento.data_encerramento -
                    tratamento.data_inicio
                ).days

            else:

                dias = (
                    timezone.localdate() -
                    tratamento.data_inicio
                ).days

            celula.value = dias

        else:

            celula.value = 0

        celula.number_format = "#,##0"

        linha += 1

    # =========================================
    # LARGURA DAS COLUNAS
    # =========================================

    larguras = {

        "A": 35,

        "B": 35,

        "C": 18,

        "D": 18,

        "E": 18,

        "F": 12,

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[
            coluna
        ].width = largura

    # =========================================
    # DOWNLOAD
    # =========================================

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Relatorio_Tratamentos_{timezone.now().strftime("%d-%m-%Y")}.xlsx"'

    )

    wb.save(response)

    return response


# =========================================
# FUNCIONÁRIOS DO PERFIL
# =========================================

@login_required
def funcionarios_perfil(request, perfil_id):

    perfil = get_object_or_404(

        Perfil,

        id=perfil_id

    )

    funcionarios = PerfilUsuario.objects.filter(

        perfil_acesso=perfil

    ).select_related(

        "usuario"

    )

    context = {

        "perfil": perfil,

        "funcionarios": funcionarios,

    }

    return render(

        request,

        "accounts/funcionarios_perfil.html",

        context

    )